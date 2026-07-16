import os
import sys
import json
import base64
import sqlite3
import webbrowser
import mimetypes
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from threading import Timer
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_from_directory, jsonify, g, has_request_context

# Load .env (GOOGLE_API_KEY, LICENSE_SERVER_URL, GOOGLE_OAUTH_*, FLASK_SECRET_KEY)
# — override=True so an empty shell var doesn't shadow the file
try:
    from dotenv import load_dotenv
    load_dotenv(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'),
        override=True,
    )
except ImportError:
    pass

app = Flask(__name__)
# Max upload size: 25 MB (one big photo of a multi-row ledger page is ~3 MB)
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024
# Stay signed in for 30 days when "session.permanent = True" is set on login
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
app.config['TEMPLATES_AUTO_RELOAD'] = True   # always re-read templates on each request
app.jinja_env.auto_reload = True

# APP_DIR resolution:
#   - In source-run mode (`python3 app.py`): the script's directory.
#   - In PyInstaller bundle mode: the directory of the EXECUTABLE (next to the
#     customer's bills.db / uploads / backups). NOT sys._MEIPASS, which is a
#     throwaway temp dir that gets a new random name every launch — using
#     _MEIPASS would silently wipe the customer's data on every restart.
def _resolve_app_dir():
    if getattr(sys, 'frozen', False):
        # Running inside a PyInstaller bundle. sys.executable points at the
        # Munshi binary; its directory is where the customer keeps their data.
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


APP_DIR = _resolve_app_dir()
DB_PATH = os.path.join(APP_DIR, 'bills.db')
UPLOAD_DIR = os.path.join(APP_DIR, 'uploads')
BACKUP_DIR = os.path.join(APP_DIR, 'backups')
os.makedirs(UPLOAD_DIR, exist_ok=True)


def _resolve_secret_key():
    """Per-install Flask secret_key. Three sources, first hit wins:
       1) FLASK_SECRET_KEY env var (production deploy, explicit control)
       2) APP_DIR/.flask_secret file (persisted random key, auto-generated
          on first run of a fresh install — different per customer)
       3) Fall back to a generated key + warn (the file write must have failed)
       This is critical for Munshi: every customer's session cookies must be
       signed with a UNIQUE secret, otherwise stealing one .exe + bills.db
       lets an attacker forge logins on every other customer's install."""
    env_val = os.environ.get('FLASK_SECRET_KEY', '').strip()
    if env_val:
        return env_val
    secret_path = os.path.join(APP_DIR, '.flask_secret')
    try:
        if os.path.exists(secret_path):
            with open(secret_path, 'r') as f:
                val = f.read().strip()
                if val and len(val) >= 32:
                    return val
        # Generate a fresh per-install secret and persist it (0600 perms)
        import secrets as _secrets
        val = _secrets.token_urlsafe(48)
        with open(secret_path, 'w') as f:
            f.write(val)
        try:
            os.chmod(secret_path, 0o600)
        except OSError:
            pass     # best-effort on Windows
        return val
    except Exception as e:
        # Filesystem read-only or other rare error — fall back to in-memory
        # random key. This means sessions reset on every restart (annoying but
        # secure).
        import secrets as _secrets
        app.logger.warning(f'Could not persist .flask_secret: {e}; using ephemeral key (sessions reset on restart)')
        return _secrets.token_urlsafe(48)


app.secret_key = _resolve_secret_key()
# Session-cookie hardening (matches the license server). SameSite=Lax blocks
# cross-site POSTs (a first line of CSRF defence for the POST forms); HttpOnly
# hides the cookie from JS. Secure is intentionally OFF because Munshi runs over
# http://127.0.0.1, not HTTPS — marking Secure would make the browser drop the
# cookie and log everyone out. (Set SESSION_COOKIE_SECURE=True only if you ever
# serve Munshi over HTTPS.)
app.config.update(
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_HTTPONLY=True,
)
os.makedirs(BACKUP_DIR, exist_ok=True)

ALLOWED_EXTS = {'.jpg', '.jpeg', '.png', '.pdf', '.webp', '.gif', '.heic'}


# ── App-level error handlers (no raw Python tracebacks to user) ────────────────

@app.errorhandler(413)
def err_too_large(e):
    return render_template('error.html',
        code=413, title='File too large',
        message='That file is bigger than 25 MB. Try a smaller image, or compress it before uploading.'
    ), 413


@app.errorhandler(404)
def err_not_found(e):
    return render_template('error.html',
        code=404, title='Page not found',
        message='That URL does not exist. Use the navbar to get back to where you were.'
    ), 404


@app.errorhandler(Exception)
def err_unhandled(e):
    # Re-raise HTTPException so Flask renders 4xx/5xx normally; only catch true crashes
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e
    # Log full traceback to server log
    import traceback
    app.logger.error('Unhandled exception:\n' + traceback.format_exc())
    return render_template('error.html',
        code=500, title='Something went wrong',
        message='An unexpected error happened. Your data is safe — nothing was changed. '
                'Please try again. If it keeps happening, take a screenshot and share with support.',
        detail=str(e)[:300]
    ), 500

# ── DB helpers ────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    # WAL lets the AI-extraction reader and a background writer coexist without
    # "database is locked"; the busy_timeout absorbs brief write locks.
    try:
        conn.execute('PRAGMA journal_mode=WAL')
        conn.execute('PRAGMA busy_timeout=8000')
    except sqlite3.Error:
        pass
    # Safety net against connection leaks: if we're inside a request, register
    # this connection so it is GUARANTEED to be closed at request teardown —
    # even when an exception skips the caller's explicit conn.close(). A leaked
    # SQLite connection keeps its WAL write-lock, which is the root cause of the
    # intermittent "database is locked" failures that build up over time.
    # Callers still close their own connections on the happy path (close() is
    # idempotent, so the teardown sweep is a no-op then). Background threads run
    # without a request context and keep managing their own connections as before.
    if has_request_context():
        conns = g.get('_db_conns')
        if conns is None:
            conns = []
            g._db_conns = conns
        conns.append(conn)
    return conn


@app.teardown_request
def _close_request_db_conns(exc):
    """Close any DB connection opened via get_db() during this request that the
       handler didn't already close (e.g. an exception unwound past its close()).
       close() is idempotent, so connections closed normally are unaffected."""
    for conn in g.pop('_db_conns', []) or []:
        try:
            conn.close()
        except Exception:
            pass


def _bootstrap_from_seed_if_missing():
    """On a fresh install (no bills.db next to the app), copy the seed DB into
       place. We ship BLANK by default: prefer data/seed_blank.db (no real firm
       identity, no users, no diesel vendors) so a fresh copy lands on the
       /setup wizard. Falls back to data/seed.db if the blank seed is absent.
       Either way the operational tables (bills, challans, ledger, payments,
       audit) start empty and the customer configures their own firm."""
    if os.path.exists(DB_PATH):
        return  # already initialised on a previous run

    # Prefer the blank seed. Look in two places for each: alongside the script
    # (dev runs) or inside the PyInstaller _MEIPASS bundle (production runs).
    seed_names = ['seed_blank.db', 'seed.db']
    candidates = []
    try:
        bundle_dir = getattr(sys, '_MEIPASS', None)
    except Exception:
        bundle_dir = None
    for name in seed_names:
        if bundle_dir:
            candidates.append(os.path.join(bundle_dir, 'data', name))
        candidates.append(os.path.join(APP_DIR, 'data', name))

    for candidate in candidates:
        if os.path.exists(candidate):
            try:
                import shutil
                shutil.copy2(candidate, DB_PATH)
                # Genuinely-useful one-time startup notice (fresh install only):
                # kept on stdout so it shows without any logging config.
                print(f'[startup] Bootstrapped bills.db from seed: {candidate}', flush=True)
                return
            except Exception as e:
                app.logger.warning(f'Could not copy seed ({candidate}): {e}')
    # No seed found — init_db() will create empty tables from scratch.


def init_db():
    _bootstrap_from_seed_if_missing()
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS bills (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_no     TEXT    UNIQUE NOT NULL,
            bill_date   TEXT,
            recipient_name    TEXT,
            recipient_address TEXT,
            recipient_gstin   TEXT,
            state_code        TEXT,
            trip_type         TEXT,
            vehicle_no        TEXT,
            freight_type      TEXT,
            delivery_month    TEXT,
            client_name       TEXT,
            total_amount      REAL DEFAULT 0,
            deliveries        TEXT DEFAULT '[]',
            created_at        TEXT
        );
        CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT
        );
        CREATE TABLE IF NOT EXISTS recipients (
            name         TEXT PRIMARY KEY,
            address      TEXT,
            gstin        TEXT,
            state_code   TEXT,
            freight_rate REAL,
            updated_at   TEXT
        );
        CREATE TABLE IF NOT EXISTS vehicles (
            vehicle_no TEXT PRIMARY KEY,
            updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS extractions (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at   TEXT,
            mode         TEXT,    -- 'combine' | 'split'
            status       TEXT,    -- 'pending' | 'extracted' | 'reviewed' | 'used'
            note         TEXT
        );
        CREATE TABLE IF NOT EXISTS extracted_invoices (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            extraction_id   INTEGER,
            file_name       TEXT,    -- relative path under uploads/
            seq             INTEGER, -- order uploaded
            raw_json        TEXT,    -- full Gemini response JSON
            edited_json     TEXT,    -- after user edits (null until saved)
            error           TEXT,
            FOREIGN KEY(extraction_id) REFERENCES extractions(id)
        );
        CREATE TABLE IF NOT EXISTS challans (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            lr_no               TEXT UNIQUE,
            challan_date        TEXT,
            consignor_name      TEXT,
            consignor_address   TEXT,
            consignee_name      TEXT,
            consignee_address   TEXT,
            from_city_state     TEXT,
            to_city_state       TEXT,
            invoice_no          TEXT,
            invoice_date        TEXT,
            consignment_value   REAL,
            gst_number          TEXT,
            no_of_articles      TEXT,
            description         TEXT,
            value_of_goods      REAL,
            weight_kg           REAL,
            del_no              TEXT,
            shipment_no         TEXT,
            cost_no             TEXT,
            seal_no             TEXT,
            driver_name         TEXT,
            driver_mobile       TEXT,
            truck_no            TEXT,
            gate_in_time        TEXT,
            gate_out_time       TEXT,
            lane_transit_time   TEXT,
            expected_arrival    TEXT,
            source_image        TEXT,    -- relative path under uploads/
            raw_extraction      TEXT,    -- raw JSON from Gemini
            confidence_json     TEXT,    -- {field: 'low'|'medium'} for flagged fields
            status              TEXT DEFAULT 'open',  -- open | pod_received | billed
            notes               TEXT,
            created_at          TEXT,
            updated_at          TEXT
        );
        CREATE TABLE IF NOT EXISTS drivers (
            mobile     TEXT PRIMARY KEY,
            name       TEXT,
            updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS freight_rates (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_name TEXT,
            party_code    TEXT,
            location      TEXT,
            dist_twy_km   INTEGER,
            dist_owy_km   INTEGER,
            lp_owy        REAL,
            lp_twy        REAL,
            trolla_owy    REAL,
            trolla_twy    REAL,
            updated_at    TEXT,
            UNIQUE(customer_name, location)
        );

        CREATE TABLE IF NOT EXISTS transporters (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            name         TEXT UNIQUE NOT NULL,
            mobile       TEXT,
            bank_details TEXT,
            notes        TEXT,
            created_at   TEXT,
            updated_at   TEXT
        );

        CREATE TABLE IF NOT EXISTS diesel_vendors (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT UNIQUE NOT NULL,
            location    TEXT,
            notes       TEXT,
            created_at  TEXT,
            updated_at  TEXT
        );

        CREATE TABLE IF NOT EXISTS ledger_extractions (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            source_image  TEXT,
            page_date     TEXT,
            raw_json      TEXT,
            edited_json   TEXT,
            status        TEXT DEFAULT 'pending',
            created_at    TEXT
        );
        CREATE TABLE IF NOT EXISTS ledger_entries (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            challan_id      INTEGER,
            entry_date      TEXT,
            gr_no           TEXT,
            vehicle_no      TEXT,
            station         TEXT,
            shipment_no     TEXT,
            trip_type       TEXT DEFAULT 'One Way',
            mt_qty          REAL,
            freight         REAL DEFAULT 0,
            advance_cash    REAL DEFAULT 0,
            advance_account REAL DEFAULT 0,
            diesel          REAL DEFAULT 0,
            diesel_vendor_id INTEGER,
            transporter_id  INTEGER,
            pod_received    INTEGER DEFAULT 0,
            pod_date        TEXT,
            pod_image       TEXT,
            paid            INTEGER DEFAULT 0,
            paid_date       TEXT,
            paid_mode       TEXT,
            paid_amount     REAL,
            paid_reference  TEXT,
            remarks         TEXT,
            created_at      TEXT,
            updated_at      TEXT,
            FOREIGN KEY(transporter_id)   REFERENCES transporters(id),
            FOREIGN KEY(diesel_vendor_id) REFERENCES diesel_vendors(id),
            FOREIGN KEY(challan_id)       REFERENCES challans(id)
        );

        -- Diesel vendors are NOT seeded: a fresh install starts blank and the
        -- customer adds their own fuel pumps via Masters. (Removing this seed
        -- does not delete rows that already exist in an existing bills.db.)

        INSERT OR IGNORE INTO settings VALUES ('next_bill_number', '1');
        INSERT OR IGNORE INTO settings VALUES ('next_lr_number',   '13453');
        INSERT OR IGNORE INTO settings VALUES ('pod_overdue_days', '10');
        -- ── Default consignor (the entity hiring JL — i.e. our paying client) ──
        -- Blank on a fresh install; the customer fills these in as they work.
        INSERT OR IGNORE INTO settings VALUES ('default_consignor_name', '');
        INSERT OR IGNORE INTO settings VALUES ('default_consignor_address', '');
        INSERT OR IGNORE INTO settings VALUES ('default_consignor_gstin', '');
        INSERT OR IGNORE INTO settings VALUES ('default_consignor_state', '');
        INSERT OR IGNORE INTO settings VALUES ('client_name',      '');
        INSERT OR IGNORE INTO settings VALUES ('vehicle_type',     'LP');
        INSERT OR IGNORE INTO settings VALUES ('freight_type',     'A1/A4');
        INSERT OR IGNORE INTO settings VALUES ('clients',          '[]');

        -- ── Supplier identity (the firm that ISSUES bills) ──
        -- Blank on a fresh install; the setup wizard collects these.
        INSERT OR IGNORE INTO settings VALUES ('supplier_name', '');
        INSERT OR IGNORE INTO settings VALUES ('supplier_address', '');
        INSERT OR IGNORE INTO settings VALUES ('supplier_gstin', '');
        INSERT OR IGNORE INTO settings VALUES ('supplier_state_code', '');
        INSERT OR IGNORE INTO settings VALUES ('supplier_pan', '');
        INSERT OR IGNORE INTO settings VALUES ('supplier_phone', '');
        INSERT OR IGNORE INTO settings VALUES ('supplier_email', '');
        INSERT OR IGNORE INTO settings VALUES ('supplier_bank_name', '');
        INSERT OR IGNORE INTO settings VALUES ('supplier_bank_account', '');
        INSERT OR IGNORE INTO settings VALUES ('supplier_bank_ifsc', '');
        -- GST defaults for new bills
        INSERT OR IGNORE INTO settings VALUES ('default_hsn_sac', '996511');  -- Goods transport by road
        INSERT OR IGNORE INTO settings VALUES ('default_reverse_charge', '1');  -- GTA most common
        INSERT OR IGNORE INTO settings VALUES ('default_gst_pct', '5');  -- 5% if forward charge for GTA
        -- Generic GTA invoice fields (were hardcoded in the old FIXED dict)
        INSERT OR IGNORE INTO settings VALUES ('description_of_service', 'GOODS TRANSPORT AGENCY SERVICE');
        INSERT OR IGNORE INTO settings VALUES ('sac_code', '996791');
        INSERT OR IGNORE INTO settings VALUES ('place_of_supply', 'UTTAR PRADESH');  -- supplier's state
    ''')

    # Idempotent column adds (safe across upgrades)
    _add_column_if_missing(conn, 'recipients', 'freight_rate', 'REAL')
    _add_column_if_missing(conn, 'vehicles',   'transporter_id', 'INTEGER')
    # Phase D — cross-stage links
    _add_column_if_missing(conn, 'challans',        'ledger_entry_id', 'INTEGER')
    _add_column_if_missing(conn, 'ledger_entries',  'bill_id',         'INTEGER')
    _add_column_if_missing(conn, 'bills',           'ledger_entry_id', 'INTEGER')
    # Phase E — client payment tracking on bills
    _add_column_if_missing(conn, 'bills', 'client_paid',           'INTEGER DEFAULT 0')
    _add_column_if_missing(conn, 'bills', 'client_paid_date',      'TEXT')
    _add_column_if_missing(conn, 'bills', 'client_paid_amount',    'REAL')
    _add_column_if_missing(conn, 'bills', 'client_paid_mode',      'TEXT')
    _add_column_if_missing(conn, 'bills', 'client_paid_reference', 'TEXT')
    # Phase F — Trip 360 cross-doc identifiers
    # del_no is what JL writes on the challan; pod_doc_no is what the buyer's PoD slip says.
    # In normal flow they match; the gap surfaces typos / wrong PoDs.
    _add_column_if_missing(conn, 'challans',       'pod_doc_no', 'TEXT')
    _add_column_if_missing(conn, 'ledger_entries', 'weight_kg',  'REAL')

    # Phase G — GST-compliant invoicing
    # `total_amount` stays the GRAND total (incl. tax) for backward compatibility.
    # `taxable_value` is the pre-tax amount = freight earned. For RCM bills, tax fields
    # are all zero and total_amount = taxable_value.
    _add_column_if_missing(conn, 'bills', 'hsn_sac',         "TEXT DEFAULT '996511'")
    _add_column_if_missing(conn, 'bills', 'taxable_value',   'REAL DEFAULT 0')
    _add_column_if_missing(conn, 'bills', 'reverse_charge',  'INTEGER DEFAULT 1')
    _add_column_if_missing(conn, 'bills', 'place_of_supply', 'TEXT')   # 2-digit state code
    _add_column_if_missing(conn, 'bills', 'igst_pct',        'REAL DEFAULT 0')
    _add_column_if_missing(conn, 'bills', 'cgst_pct',        'REAL DEFAULT 0')
    _add_column_if_missing(conn, 'bills', 'sgst_pct',        'REAL DEFAULT 0')
    _add_column_if_missing(conn, 'bills', 'igst_amount',     'REAL DEFAULT 0')
    _add_column_if_missing(conn, 'bills', 'cgst_amount',     'REAL DEFAULT 0')
    _add_column_if_missing(conn, 'bills', 'sgst_amount',     'REAL DEFAULT 0')
    _add_column_if_missing(conn, 'bills', 'irn',             'TEXT')   # e-invoice IRN after upload to GSTN
    _add_column_if_missing(conn, 'bills', 'irn_qr',          'TEXT')   # base64 PNG of e-invoice QR

    # ── Payments (unified AR/AP ledger) ──
    conn.execute('''
        CREATE TABLE IF NOT EXISTS payments (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            party_type    TEXT NOT NULL,    -- 'client' | 'transporter' | 'diesel_vendor'
            party_key     TEXT NOT NULL,    -- recipient name for client; id for others
            payment_date  TEXT NOT NULL,
            amount        REAL NOT NULL,    -- always positive; direction implied by party_type
            mode          TEXT,             -- Cash / UPI / Bank / Cheque
            reference     TEXT,
            notes         TEXT,
            source        TEXT DEFAULT 'manual',  -- 'manual' | 'migrated'
            created_at    TEXT,
            created_by    TEXT
        )
    ''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_payments_party ON payments(party_type, party_key)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_payments_date  ON payments(payment_date DESC)')

    # ── One-time migration: convert old per-record "paid" flags to payment rows ──
    _payments_done = conn.execute("SELECT value FROM settings WHERE key='payments_migration_done'").fetchone()
    if not _payments_done or _payments_done['value'] != '1':
        # Migrate client payments from bills.client_paid
        for r in conn.execute(
            '''SELECT id, recipient_name, total_amount, client_paid_date, client_paid_mode,
                      client_paid_amount, client_paid_reference
               FROM bills WHERE COALESCE(client_paid,0) = 1'''
        ).fetchall():
            amt = r['client_paid_amount'] or r['total_amount'] or 0
            if amt > 0 and r['recipient_name']:
                conn.execute('''
                    INSERT INTO payments (party_type, party_key, payment_date, amount,
                                          mode, reference, notes, source, created_at)
                    VALUES (?,?,?,?,?,?,?,?,?)
                ''', ('client', r['recipient_name'].strip(),
                      r['client_paid_date'] or datetime.now().strftime('%Y-%m-%d'),
                      amt, r['client_paid_mode'], r['client_paid_reference'],
                      f'(migrated from bill #{r["id"]})', 'migrated', datetime.now().isoformat()))

        # Migrate transporter payments from ledger_entries.paid
        for r in conn.execute(
            '''SELECT id, transporter_id, freight, advance_cash, advance_account, diesel,
                      paid_date, paid_mode, paid_amount, paid_reference
               FROM ledger_entries WHERE COALESCE(paid,0) = 1 AND transporter_id IS NOT NULL'''
        ).fetchall():
            balance = (r['freight'] or 0) - (r['advance_cash'] or 0) \
                    - (r['advance_account'] or 0) - (r['diesel'] or 0)
            amt = r['paid_amount'] or balance
            if amt and amt > 0:
                conn.execute('''
                    INSERT INTO payments (party_type, party_key, payment_date, amount,
                                          mode, reference, notes, source, created_at)
                    VALUES (?,?,?,?,?,?,?,?,?)
                ''', ('transporter', str(r['transporter_id']),
                      r['paid_date'] or datetime.now().strftime('%Y-%m-%d'),
                      amt, r['paid_mode'], r['paid_reference'],
                      f'(migrated from ledger #{r["id"]})', 'migrated', datetime.now().isoformat()))

        # Use the same connection (avoid the "database is locked" issue from a 2nd conn)
        conn.execute('INSERT OR REPLACE INTO settings VALUES (?,?)',
                     ('payments_migration_done', '1'))

    # ── One-time migration: existing bills had consignee in recipient_name (bug).
    # Move them to the default consignor (the actual paying client).
    cur = conn.execute("SELECT value FROM settings WHERE key='recipient_name_consignor_fix'").fetchone()
    if not cur or cur['value'] != '1':
        def _s(k, default=''):
            r = conn.execute('SELECT value FROM settings WHERE key=?', (k,)).fetchone()
            return r['value'] if r else default
        default_name  = _s('default_consignor_name',  '')
        default_addr  = _s('default_consignor_address')
        default_gstin = _s('default_consignor_gstin')
        default_state = _s('default_consignor_state')
        if default_name:
            conn.execute('''
                UPDATE bills
                SET recipient_name=?, recipient_address=?, recipient_gstin=?, state_code=?
                WHERE COALESCE(recipient_name,'') != ?
            ''', (default_name, default_addr, default_gstin, default_state, default_name))
            # Also redirect payments — old client_key was consignee names; merge into consignor
            conn.execute('''
                UPDATE payments SET party_key=?
                WHERE party_type='client' AND party_key != ?
            ''', (default_name, default_name))
        conn.execute('INSERT OR REPLACE INTO settings VALUES (?,?)',
                     ('recipient_name_consignor_fix', '1'))

    # ── One-time reconciliation: payments table becomes the SINGLE SOURCE ──
    # For any bill/ledger row still carrying a "paid" flag that does NOT already
    # have a payment row (neither an auto marker nor the one the earlier
    # _payments_migration created), insert the payment row now, so an existing
    # install becomes consistent on the first boot after this upgrade.
    # Guard against double-counting: skip if a row already exists with our
    # 'auto-paid:*' marker OR with the migration's '(migrated from …)' note.
    _ss_done = conn.execute("SELECT value FROM settings WHERE key='payments_singlesource_done'").fetchone()
    if not _ss_done or _ss_done['value'] != '1':
        for r in conn.execute(
            '''SELECT id, recipient_name, total_amount, client_paid_amount,
                      client_paid_date, client_paid_mode
               FROM bills WHERE COALESCE(client_paid,0) = 1'''
        ).fetchall():
            name = (r['recipient_name'] or '').strip()
            if not name:
                continue
            ref = f'auto-paid:bill:{r["id"]}'
            exists = conn.execute(
                'SELECT 1 FROM payments WHERE reference=? OR notes=? LIMIT 1',
                (ref, f'(migrated from bill #{r["id"]})')
            ).fetchone()
            if exists:
                continue
            amt = r['client_paid_amount'] or r['total_amount'] or 0
            if amt and amt > 0:
                conn.execute('''
                    INSERT INTO payments (party_type, party_key, payment_date, amount,
                                          mode, reference, notes, source, created_at)
                    VALUES ('client',?,?,?,?,?,?, 'auto', ?)
                ''', (name,
                      r['client_paid_date'] or datetime.now().strftime('%Y-%m-%d'),
                      amt, r['client_paid_mode'], ref,
                      '(auto: reconciled marked-paid bill)', datetime.now().isoformat()))

        for r in conn.execute(
            '''SELECT id, transporter_id, freight, advance_cash, advance_account, diesel,
                      paid_amount, paid_date, paid_mode
               FROM ledger_entries WHERE COALESCE(paid,0) = 1 AND transporter_id IS NOT NULL'''
        ).fetchall():
            ref = f'auto-paid:ledger:{r["id"]}'
            exists = conn.execute(
                'SELECT 1 FROM payments WHERE reference=? OR notes=? LIMIT 1',
                (ref, f'(migrated from ledger #{r["id"]})')
            ).fetchone()
            if exists:
                continue
            net = (r['freight'] or 0) - (r['advance_cash'] or 0) \
                - (r['advance_account'] or 0) - (r['diesel'] or 0)
            amt = r['paid_amount'] or net
            if amt and amt > 0:
                conn.execute('''
                    INSERT INTO payments (party_type, party_key, payment_date, amount,
                                          mode, reference, notes, source, created_at)
                    VALUES ('transporter',?,?,?,?,?,?, 'auto', ?)
                ''', (str(r['transporter_id']),
                      r['paid_date'] or datetime.now().strftime('%Y-%m-%d'),
                      amt, r['paid_mode'], ref,
                      '(auto: reconciled marked-paid trip)', datetime.now().isoformat()))

        conn.execute('INSERT OR REPLACE INTO settings VALUES (?,?)',
                     ('payments_singlesource_done', '1'))

    # Audit trail
    conn.execute('''
        CREATE TABLE IF NOT EXISTS audit_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            occurred_at TEXT NOT NULL,
            user_name   TEXT,
            action      TEXT,
            entity      TEXT,
            entity_id   INTEGER,
            summary     TEXT,
            changes     TEXT
        )
    ''')
    # ── Audit indexes
    conn.execute('CREATE INDEX IF NOT EXISTS idx_audit_entity     ON audit_log(entity, entity_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_audit_when       ON audit_log(occurred_at DESC)')
    # ── Hot-query indexes (for performance with thousands of rows)
    conn.execute('CREATE INDEX IF NOT EXISTS idx_bills_date       ON bills(bill_date DESC)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_bills_recipient  ON bills(recipient_name)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_bills_paid       ON bills(client_paid)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_ledger_date      ON ledger_entries(entry_date DESC)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_ledger_pod       ON ledger_entries(pod_received, bill_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_ledger_vehicle   ON ledger_entries(vehicle_no)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_ledger_transp    ON ledger_entries(transporter_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_challans_lr      ON challans(lr_no)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_challans_status  ON challans(status)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_rates_customer   ON freight_rates(customer_name)')

    # ── License state (singleton row, id=1) ──
    # Caches the latest /verify response from the Munshi license server so the
    # app keeps working between phone-homes (which only happen daily).
    conn.execute('''
        CREATE TABLE IF NOT EXISTS license_state (
            id              INTEGER PRIMARY KEY CHECK (id = 1),
            license_key     TEXT,
            status          TEXT,            -- active | grace | locked | expired | suspended | not_found | unconfigured
            tier            TEXT,
            max_trucks      INTEGER,
            expires_at      TEXT,
            days_to_expiry  INTEGER,
            message         TEXT,
            last_checked_at TEXT,
            last_error      TEXT
        )
    ''')
    conn.execute('''
        INSERT OR IGNORE INTO license_state (id, status, message)
        VALUES (1, 'unconfigured', 'No license key configured. Munshi will run in trial mode.')
    ''')

    # ── Google Drive backup state (singleton row, id=1) ──
    # Caches OAuth tokens + last-sync metadata for per-customer Drive backups.
    # Tokens never leave the customer's machine; only used to upload bills.db to
    # the customer's OWN Google Drive folder.
    conn.execute('''
        CREATE TABLE IF NOT EXISTS google_drive_state (
            id                    INTEGER PRIMARY KEY CHECK (id = 1),
            refresh_token         TEXT,
            access_token          TEXT,
            access_token_expiry   TEXT,        -- ISO UTC; refresh 60s before
            folder_id             TEXT,        -- Drive file ID of "Munshi Backups"
            folder_name           TEXT,        -- display only
            connected_email       TEXT,        -- display only
            last_sync_at          TEXT,
            last_sync_status      TEXT,        -- ok | reauth_required | offline | quota | error | not_configured
            last_sync_error       TEXT,
            last_uploaded_file    TEXT
        )
    ''')
    conn.execute('''
        INSERT OR IGNORE INTO google_drive_state (id, last_sync_status)
        VALUES (1, 'not_configured')
    ''')

    # ── Users (auth) ──
    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            username       TEXT PRIMARY KEY,
            password_hash  TEXT NOT NULL,
            full_name      TEXT,
            role           TEXT DEFAULT 'operator',  -- 'admin' | 'operator'
            is_active      INTEGER DEFAULT 1,
            must_change_password INTEGER DEFAULT 1,
            created_at     TEXT,
            last_login     TEXT
        )
    ''')
    # NOTE: No user accounts are seeded here. A fresh install starts with an
    # EMPTY users table and lands on the /setup wizard, where the buyer creates
    # their own owner login. (Removing this seed does NOT delete users that
    # already exist in an existing bills.db.)

    # ── Migration: auto-mark an already-configured install as "set up" ──
    # If this DB already has at least one user AND a non-empty supplier_name,
    # it's the founder's existing (or any previously-configured) install — it
    # must NEVER be sent to the first-run wizard. Flag it complete once.
    _setup_row = conn.execute("SELECT value FROM settings WHERE key='setup_complete'").fetchone()
    if not _setup_row or _setup_row['value'] != '1':
        _has_user = conn.execute('SELECT 1 FROM users LIMIT 1').fetchone() is not None
        _sup_row = conn.execute("SELECT value FROM settings WHERE key='supplier_name'").fetchone()
        _has_supplier = bool(_sup_row and (_sup_row['value'] or '').strip())
        if _has_user and _has_supplier:
            conn.execute('INSERT OR REPLACE INTO settings VALUES (?,?)',
                         ('setup_complete', '1'))

    # ── Crash recovery: if Flask was killed while an extraction was running in
    # a background thread, the thread is gone but the row is still 'pending'.
    # Mark such orphans as failed so the polling UI shows a clear error instead
    # of spinning forever.
    conn.execute('''
        UPDATE extractions
        SET status = 'failed',
            note   = COALESCE(note, '') || ' [interrupted by server restart]'
        WHERE status = 'pending'
    ''')

    # ── Recycle Bin: archive tables for soft-delete (undo) ───────────────────
    # Deleted bills/challans/ledger entries/payments are MOVED here (not
    # destroyed), so the user can restore them. See _ensure_archive_table.
    _ensure_archive_table(conn, 'bills')
    _ensure_archive_table(conn, 'challans')
    _ensure_archive_table(conn, 'ledger_entries')
    _ensure_archive_table(conn, 'payments')

    conn.commit()
    conn.close()


def _add_column_if_missing(conn, table, col, coltype):
    cols = [r['name'] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
    if col not in cols:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {coltype}")


def _ensure_archive_table(conn, src):
    """Create/maintain a `<src>_archive` table that mirrors `<src>`'s columns,
    for the Recycle Bin (soft-delete by row-move).

    The LIVE table never holds deleted rows, so every existing list / report /
    lookup query keeps working unchanged — deleted records can't leak into
    totals or lists, and we never had to add `deleted_at IS NULL` filters across
    ~80 query sites. Restore moves the row back with its original id, so
    foreign-key links (ledger_entries.bill_id, challans.ledger_entry_id, …)
    become valid again automatically.

    The column set is re-synced on every init_db run, so columns added to the
    source in a future upgrade are picked up here automatically (no separate
    migration to remember)."""
    arch = src + '_archive'
    conn.execute(f'CREATE TABLE IF NOT EXISTS {arch} AS SELECT * FROM {src} WHERE 0')
    src_cols = {r['name']: (r['type'] or 'TEXT')
                for r in conn.execute(f"PRAGMA table_info({src})").fetchall()}
    arch_cols = {r['name'] for r in conn.execute(f"PRAGMA table_info({arch})").fetchall()}
    for col, ctype in src_cols.items():
        if col not in arch_cols:
            conn.execute(f'ALTER TABLE {arch} ADD COLUMN {col} {ctype}')


def get_setting(key):
    conn = get_db()
    row = conn.execute('SELECT value FROM settings WHERE key=?', (key,)).fetchone()
    conn.close()
    return row['value'] if row else ''


def set_setting(key, value):
    conn = get_db()
    conn.execute('INSERT OR REPLACE INTO settings VALUES (?,?)', (key, value))
    conn.commit()
    conn.close()


def _setup_complete():
    """True once the first-run setup wizard has been finished (or the install
       was auto-marked as already-configured in init_db)."""
    return (get_setting('setup_complete') or '') == '1'


def next_bill_no():
    n = int(get_setting('next_bill_number') or 1)
    return f'JL-{n:04d}', n


def _alloc_bill_no(conn):
    """Collision-proof next bill number for INSERTs. Uses the HIGHER of the
       stored counter and (max existing JL-#### number + 1). This prevents the
       UNIQUE-constraint crash that otherwise happens whenever the counter has
       drifted behind the real data — e.g. after loading demo/sample data,
       importing/restoring bills, manual DB edits, or a genuinely concurrent
       save. Returns (bill_no, n). Call inside the same transaction as the
       INSERT, and retry on IntegrityError to cover the racing case."""
    counter = int(get_setting('next_bill_number') or 1)
    row = conn.execute(
        "SELECT MAX(CAST(SUBSTR(bill_no, 4) AS INTEGER)) FROM bills "
        "WHERE bill_no LIKE 'JL-%'").fetchone()
    max_existing = (row[0] or 0) if row else 0
    n = max(counter, max_existing + 1)
    return f'JL-{n:04d}', n


# ─────────────────────────────────────────────────────────────────────────────
# GST helpers
# ─────────────────────────────────────────────────────────────────────────────

# 2-digit state-code → state-name lookup (for the Place-of-Supply line on bills)
GST_STATE_NAMES = {
    '01': 'JAMMU & KASHMIR', '02': 'HIMACHAL PRADESH', '03': 'PUNJAB',
    '04': 'CHANDIGARH', '05': 'UTTARAKHAND', '06': 'HARYANA',
    '07': 'DELHI', '08': 'RAJASTHAN', '09': 'UTTAR PRADESH',
    '10': 'BIHAR', '11': 'SIKKIM', '12': 'ARUNACHAL PRADESH',
    '13': 'NAGALAND', '14': 'MANIPUR', '15': 'MIZORAM',
    '16': 'TRIPURA', '17': 'MEGHALAYA', '18': 'ASSAM',
    '19': 'WEST BENGAL', '20': 'JHARKHAND', '21': 'ODISHA',
    '22': 'CHHATTISGARH', '23': 'MADHYA PRADESH', '24': 'GUJARAT',
    '25': 'DAMAN & DIU', '26': 'DADRA & NAGAR HAVELI', '27': 'MAHARASHTRA',
    '28': 'ANDHRA PRADESH (OLD)', '29': 'KARNATAKA', '30': 'GOA',
    '31': 'LAKSHADWEEP', '32': 'KERALA', '33': 'TAMIL NADU',
    '34': 'PUDUCHERRY', '35': 'ANDAMAN & NICOBAR', '36': 'TELANGANA',
    '37': 'ANDHRA PRADESH', '38': 'LADAKH', '97': 'OTHER TERRITORY',
}


def validate_gstin(gstin):
    """Returns (is_valid, normalized_or_none, error_msg).
       GSTIN format: 15 chars = 2 (state) + 10 (PAN) + 1 (entity) + 1 ('Z') + 1 (checksum).
       We validate the structural pattern but skip the checksum digit (rarely typed wrong)."""
    import re
    if not gstin:
        return False, None, 'GSTIN is empty'
    g = gstin.strip().upper()
    if len(g) != 15:
        return False, None, f'GSTIN must be 15 characters (got {len(g)})'
    if not re.match(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[0-9A-Z]{1}Z[0-9A-Z]{1}$', g):
        return False, None, 'GSTIN format is invalid (expected: 09AFFFS7446N1Z6 pattern)'
    return True, g, ''


def compute_gst_split(taxable_value, gst_pct, supplier_state, place_of_supply, reverse_charge):
    """Return a dict with the tax line items.
       Rule: same state → CGST + SGST (gst_pct split equally);
             different states → IGST (full gst_pct).
       Reverse-charge bills carry 0 tax (recipient pays via RCM).
       All amounts are rounded to whole rupees — freight invoices don't carry paise."""
    tv = round(float(taxable_value or 0))    # taxable value itself is rounded too
    out = {'igst_pct': 0, 'cgst_pct': 0, 'sgst_pct': 0,
           'igst_amount': 0, 'cgst_amount': 0, 'sgst_amount': 0,
           'total_tax': 0, 'grand_total': tv}
    if reverse_charge:
        return out
    pct = float(gst_pct or 0)
    if pct <= 0:
        return out
    same_state = (supplier_state or '').strip() == (place_of_supply or '').strip()
    if same_state:
        half = round(tv * pct / 200)         # CGST = SGST = half of gst_pct, rounded
        out['cgst_pct'] = pct / 2
        out['sgst_pct'] = pct / 2
        out['cgst_amount'] = half
        out['sgst_amount'] = half
        out['total_tax'] = half * 2
    else:
        amt = round(tv * pct / 100)
        out['igst_pct'] = pct
        out['igst_amount'] = amt
        out['total_tax'] = amt
    out['grand_total'] = tv + out['total_tax']
    return out


def amount_in_words_inr(amount):
    """Convert a rupee amount to words in Indian numbering (Lakh / Crore).
       Returns: 'Rupees One Lakh Forty Seven Thousand Nine Hundred Ninety Seven Only'."""
    try:
        amt = float(amount or 0)
    except (TypeError, ValueError):
        return 'Rupees Zero Only'
    if amt < 0:
        return 'Minus ' + amount_in_words_inr(-amt)

    rupees = int(amt)
    paise  = round((amt - rupees) * 100)

    ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine',
            'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen',
            'Seventeen', 'Eighteen', 'Nineteen']
    tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy',
            'Eighty', 'Ninety']

    def two_digit(n):
        if n < 20:
            return ones[n]
        return (tens[n // 10] + (' ' + ones[n % 10] if n % 10 else '')).strip()

    def three_digit(n):
        # Returns up to 999 in words
        parts = []
        if n >= 100:
            parts.append(ones[n // 100] + ' Hundred')
            n %= 100
        if n:
            parts.append(two_digit(n))
        return ' '.join(parts)

    # Indian numbering: ones (0-999), thousand (1K-99K), lakh (1L-99L), crore (1Cr+)
    parts = []
    crore = rupees // 10000000
    rupees %= 10000000
    lakh  = rupees // 100000
    rupees %= 100000
    thousand = rupees // 1000
    rupees %= 1000
    hundreds = rupees

    if crore:
        parts.append((two_digit(crore) if crore > 0 else '') + ' Crore')
    if lakh:
        parts.append(two_digit(lakh) + ' Lakh')
    if thousand:
        parts.append(two_digit(thousand) + ' Thousand')
    if hundreds:
        parts.append(three_digit(hundreds))

    out = 'Rupees ' + (' '.join(parts).strip() or 'Zero')
    if paise:
        out += ' and ' + two_digit(paise) + ' Paise'
    out += ' Only'
    return out


def get_supplier_identity():
    """Bundle the JL supplier identity from settings (cached on the request)."""
    state_code = get_setting('supplier_state_code') or '09'
    state_name = GST_STATE_NAMES.get(state_code, '')
    return {
        'name':         get_setting('supplier_name') or '',
        'address':      get_setting('supplier_address') or '',
        'gstin':        get_setting('supplier_gstin') or '',
        'state_code':   state_code,
        'state_name':   state_name,
        'pan':          get_setting('supplier_pan') or '',
        'phone':        get_setting('supplier_phone') or '',
        'email':        get_setting('supplier_email') or '',
        'bank_name':    get_setting('supplier_bank_name') or '',
        'bank_account': get_setting('supplier_bank_account') or '',
        'bank_ifsc':    get_setting('supplier_bank_ifsc') or '',
        # Generic GTA invoice fields (moved out of the old hardcoded FIXED dict)
        'description_of_service': get_setting('description_of_service') or 'GOODS TRANSPORT AGENCY SERVICE',
        'sac_code':               get_setting('sac_code') or '996791',
        'place_of_supply':        get_setting('place_of_supply') or state_name,
    }


MONTHS = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']


def build_delivery_month(month_select, bill_date):
    """Combine month abbrev (e.g. 'MAY') with last 2 digits of bill_date year → 'MAY/25'."""
    if not month_select:
        return ''
    yr = bill_date[2:4] if bill_date and len(bill_date) >= 4 else ''
    return f'{month_select}/{yr}' if yr else month_select


def split_delivery_month(combined):
    """Reverse: 'MAY/25' → 'MAY' (just the month part for the dropdown)."""
    if not combined:
        return ''
    return combined.split('/')[0].strip().upper()


def remember_recipient(conn, name, address, gstin, state_code, freight_rate=None):
    if not name:
        return
    # Only update freight_rate if a non-empty value was provided
    rate_val = None
    if freight_rate not in (None, '', 0, '0'):
        try:
            rate_val = float(freight_rate)
        except (TypeError, ValueError):
            rate_val = None
    if rate_val is not None:
        conn.execute(
            '''INSERT INTO recipients (name, address, gstin, state_code, freight_rate, updated_at)
               VALUES (?,?,?,?,?,?)
               ON CONFLICT(name) DO UPDATE SET
                 address=excluded.address, gstin=excluded.gstin,
                 state_code=excluded.state_code,
                 freight_rate=excluded.freight_rate,
                 updated_at=excluded.updated_at''',
            (name.strip(), address or '', gstin or '', state_code or '',
             rate_val, datetime.now().isoformat())
        )
    else:
        conn.execute(
            '''INSERT INTO recipients (name, address, gstin, state_code, updated_at)
               VALUES (?,?,?,?,?)
               ON CONFLICT(name) DO UPDATE SET
                 address=excluded.address, gstin=excluded.gstin,
                 state_code=excluded.state_code, updated_at=excluded.updated_at''',
            (name.strip(), address or '', gstin or '', state_code or '',
             datetime.now().isoformat())
        )


def remember_vehicle(conn, vehicle_no):
    if not vehicle_no:
        return
    conn.execute(
        '''INSERT INTO vehicles (vehicle_no, updated_at) VALUES (?,?)
           ON CONFLICT(vehicle_no) DO UPDATE SET updated_at=excluded.updated_at''',
        (vehicle_no.strip().upper(), datetime.now().isoformat())
    )


def get_recipients():
    conn = get_db()
    rows = conn.execute(
        'SELECT name, address, gstin, state_code, freight_rate FROM recipients ORDER BY updated_at DESC'
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_vehicles():
    conn = get_db()
    rows = conn.execute('SELECT vehicle_no FROM vehicles ORDER BY updated_at DESC').fetchall()
    conn.close()
    return [r['vehicle_no'] for r in rows]


# ── Freight rate list (uploaded Excel master rates) ────────────────────────────

def get_rate_list():
    conn = get_db()
    rows = conn.execute(
        '''SELECT customer_name, party_code, location, dist_twy_km, dist_owy_km,
                  lp_owy, lp_twy, trolla_owy, trolla_twy
           FROM freight_rates ORDER BY customer_name'''
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def import_rate_list_from_xlsx(xlsx_path):
    """Parse a rate-list workbook matching the Book1.xlsx layout and upsert into freight_rates.
       Header row with column names is at row 4 (1-based). Data starts at row 5.
       Columns: A CUSTOMER NAME | B PARTY CODE | C Location | D Dis TWY | E Dis OWY |
                F LP OWY | G LP TWY | H Trolla OWY | I Trolla TWY
       Returns (added_or_updated_count, error_string_or_None)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return 0, "openpyxl not installed (pip install openpyxl)"
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        return 0, f"Could not open workbook: {e}"

    conn = get_db()
    count = 0
    skipped = 0
    # Find the header row dynamically — look for a row containing "CUSTOMER NAME"
    header_row = None
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and any(isinstance(c, str) and 'CUSTOMER NAME' in c.upper() for c in row if c):
            header_row = ri
            break
    if header_row is None:
        conn.close()
        return 0, "Could not find header row with 'CUSTOMER NAME'"

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        # Need at least the customer name (col A) populated
        name = (row[0] or '').strip() if isinstance(row[0], str) else (str(row[0]).strip() if row[0] is not None else '')
        if not name:
            skipped += 1
            continue
        try:
            party_code = str(row[1]).strip() if row[1] is not None else ''
            location   = (row[2] or '').strip() if isinstance(row[2], str) else (str(row[2]).strip() if row[2] is not None else '')
            dist_twy   = int(row[3]) if row[3] not in (None, '') else None
            dist_owy   = int(row[4]) if row[4] not in (None, '') else None
            lp_owy     = float(row[5]) if row[5] not in (None, '') else None
            lp_twy     = float(row[6]) if row[6] not in (None, '') else None
            trl_owy    = float(row[7]) if row[7] not in (None, '') else None
            trl_twy    = float(row[8]) if row[8] not in (None, '') else None
        except (ValueError, TypeError, IndexError):
            skipped += 1
            continue

        conn.execute(
            '''INSERT INTO freight_rates
                 (customer_name, party_code, location, dist_twy_km, dist_owy_km,
                  lp_owy, lp_twy, trolla_owy, trolla_twy, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?)
               ON CONFLICT(customer_name, location) DO UPDATE SET
                 party_code=excluded.party_code,
                 dist_twy_km=excluded.dist_twy_km,
                 dist_owy_km=excluded.dist_owy_km,
                 lp_owy=excluded.lp_owy, lp_twy=excluded.lp_twy,
                 trolla_owy=excluded.trolla_owy, trolla_twy=excluded.trolla_twy,
                 updated_at=excluded.updated_at''',
            (name.upper(), party_code, location.upper(),
             dist_twy, dist_owy, lp_owy, lp_twy, trl_owy, trl_twy,
             datetime.now().isoformat())
        )
        count += 1
    conn.commit()
    conn.close()
    msg = None if skipped == 0 else f"(skipped {skipped} blank/invalid rows)"
    return count, msg


def fmt_amount(v):
    try:
        f = float(v)
        return f'{f:,.2f}' if f else ''
    except Exception:
        return str(v) if v else ''


def fmt_date(v):
    """Convert YYYY-MM-DD → DD/MM/YY (Indian print format)."""
    if not v:
        return ''
    try:
        d = datetime.strptime(str(v)[:10], '%Y-%m-%d')
        return d.strftime('%d/%m/%y')
    except Exception:
        return str(v)


def _from_json_filter(s):
    if not s:
        return None
    try:
        return json.loads(s)
    except Exception:
        return None


def fmt_int(v):
    """Format a number as a rounded integer with Indian thousands separators.
       Empty / non-numeric / zero values render as empty string (so blank cells
       on the printed bill stay blank rather than showing '0')."""
    if v in (None, '', 0, '0'):
        return ''
    try:
        f = float(v)
        if abs(f) < 0.5:
            return ''
        return f'{round(f):,}'
    except (TypeError, ValueError):
        return str(v)


def fmt_int0(v):
    """Same as fmt_int but renders 0 (not blank) — for totals row that must show '0' on empty bills."""
    try:
        return f'{round(float(v or 0)):,}'
    except (TypeError, ValueError):
        return '0'


app.jinja_env.filters['fmt']       = fmt_amount
app.jinja_env.filters['date']      = fmt_date
app.jinja_env.filters['from_json'] = _from_json_filter
app.jinja_env.filters['inr']       = fmt_int      # blank for zero/empty
app.jinja_env.filters['inr0']      = fmt_int0     # always shows '0' for empty


# ── i18n: English by default, optional Hindi (transporter register) ──────────
# Simple, robust, server-side. The whole app frame stays English unless the
# session language is 'hi' AND a Hindi translation exists for the exact English
# source string. Missing keys fall back to English — never a crash, never blank.
SUPPORTED_LANGS = ('en', 'hi')
_TRANSLATIONS = {}


def _load_translations():
    """Load translations/hi.json (a flat English→Hindi map). Looks alongside
       the script (dev runs) and inside the PyInstaller _MEIPASS bundle
       (packaged runs). Any failure just leaves the map empty → all-English."""
    global _TRANSLATIONS
    candidates = []
    bundle_dir = getattr(sys, '_MEIPASS', None)
    if bundle_dir:
        candidates.append(os.path.join(bundle_dir, 'translations', 'hi.json'))
    candidates.append(os.path.join(APP_DIR, 'translations', 'hi.json'))
    candidates.append(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                   'translations', 'hi.json'))
    for path in candidates:
        try:
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    _TRANSLATIONS = data
                    return
        except Exception as e:
            app.logger.warning(f'Could not load translations from {path}: {e}')
    _TRANSLATIONS = {}


_load_translations()


def t(s):
    """Translate English source string `s` to Hindi when the current session
       language is 'hi' and a translation exists; otherwise return `s`
       unchanged. Safe to call outside a request context."""
    if not s:
        return s
    try:
        if session.get('lang') == 'hi':
            return _TRANSLATIONS.get(s, s)
    except Exception:
        pass
    return s


# Expose t() to every template as a Jinja global: {{ t('Dashboard') }}
app.jinja_env.globals['t'] = t


@app.context_processor
def inject_lang():
    """Make the current language code available to every template as {{ lang }}
       (default 'en') — used for <html lang=…> and the toggle's active state."""
    try:
        return {'lang': session.get('lang', 'en')}
    except Exception:
        return {'lang': 'en'}


# ── Payments / AR / AP helpers ────────────────────────────────────────────────

def _client_charges(conn, name):
    """Total billed to a client (sum of all bills with that recipient_name)."""
    r = conn.execute(
        'SELECT COALESCE(SUM(total_amount),0) FROM bills WHERE recipient_name=?', (name,)
    ).fetchone()
    return float(r[0]) if r else 0.0


def _transporter_charges_net(conn, transporter_id):
    """For a transporter: gross freight earned MINUS advances already paid at trip time
       (cash, account, diesel). Result is what would be owed if we hadn't made any
       direct settlement payments yet."""
    r = conn.execute('''
        SELECT COALESCE(SUM(freight - COALESCE(advance_cash,0) - COALESCE(advance_account,0)
                            - COALESCE(diesel,0)), 0)
        FROM ledger_entries WHERE transporter_id=?''', (transporter_id,)
    ).fetchone()
    return float(r[0]) if r else 0.0


def _diesel_vendor_charges(conn, vendor_id):
    """Total diesel value drawn from this pump across all ledger entries."""
    r = conn.execute(
        'SELECT COALESCE(SUM(diesel),0) FROM ledger_entries WHERE diesel_vendor_id=? AND diesel > 0',
        (vendor_id,)
    ).fetchone()
    return float(r[0]) if r else 0.0


def _payments_total(conn, party_type, party_key):
    r = conn.execute(
        'SELECT COALESCE(SUM(amount),0) FROM payments WHERE party_type=? AND party_key=?',
        (party_type, str(party_key))
    ).fetchone()
    return float(r[0]) if r else 0.0


def get_party_balance(party_type, party_key):
    """Positive number = the balance still pending.
       For client: how much they owe us.
       For transporter / diesel: how much we owe them."""
    conn = get_db()
    if party_type == 'client':
        charges = _client_charges(conn, party_key)
    elif party_type == 'transporter':
        charges = _transporter_charges_net(conn, party_key)
    elif party_type == 'diesel_vendor':
        charges = _diesel_vendor_charges(conn, party_key)
    else:
        conn.close()
        return 0.0
    paid = _payments_total(conn, party_type, party_key)
    conn.close()
    return charges - paid


# ── Auto payment rows for the "marked paid" flags ────────────────────────────
# The payments table is the SINGLE SOURCE OF TRUTH for balances. When a user
# flips a "paid" flag (a bill's client_paid, a ledger entry's paid), we mirror
# that into a payments row so every screen (dashboard, reports, hub) agrees.
# Each auto row carries a stable `reference` marker so we can find/refresh/remove
# exactly the row that belongs to that flag:
#     reference = 'auto-paid:bill:<bill_id>'      (client payment for a bill)
#     reference = 'auto-paid:ledger:<le_id>'      (transporter payment for a trip)

def _auto_payment_upsert(conn, party_type, party_key, amount, ref,
                         when=None, mode=None, created_by=None,
                         note='(auto: marked paid)'):
    """Insert (or refresh) the auto payment row identified by `ref`.
       Idempotent: any existing row with this marker is replaced, so repeated
       submits never double-count. A non-positive amount just clears the marker."""
    conn.execute('DELETE FROM payments WHERE reference=?', (ref,))
    if not party_key or amount is None or amount <= 0:
        return
    conn.execute('''
        INSERT INTO payments (party_type, party_key, payment_date, amount,
                              mode, reference, notes, source, created_at, created_by)
        VALUES (?,?,?,?,?,?,?, 'auto', ?, ?)
    ''', (party_type, str(party_key),
          when or datetime.now().strftime('%Y-%m-%d'),
          float(amount), mode, ref, note,
          datetime.now().isoformat(), created_by))


def _auto_payment_remove(conn, ref):
    """Reverse (delete) the auto payment row for a flag that was un-marked."""
    conn.execute('DELETE FROM payments WHERE reference=?', (ref,))


def get_party_transactions(party_type, party_key, limit=200):
    """Return interleaved chronological list of charges (bills/trips/diesel-given)
       and payments for a given party. Each item: dict with date, kind, amount, ref, link."""
    conn = get_db()
    items = []

    if party_type == 'client':
        for r in conn.execute(
            '''SELECT id, bill_no, bill_date, total_amount FROM bills
               WHERE recipient_name=? ORDER BY bill_date DESC, id DESC LIMIT ?''',
            (party_key, limit)
        ).fetchall():
            items.append({
                'date': r['bill_date'], 'kind': 'charge', 'label': f'Bill {r["bill_no"]}',
                'amount': r['total_amount'] or 0, 'link': url_for('view_bill', bill_id=r['id']),
            })
    elif party_type == 'transporter':
        for r in conn.execute(
            '''SELECT id, gr_no, entry_date, freight, advance_cash, advance_account, diesel
               FROM ledger_entries WHERE transporter_id=?
               ORDER BY entry_date DESC, id DESC LIMIT ?''',
            (party_key, limit)
        ).fetchall():
            net = (r['freight'] or 0) - (r['advance_cash'] or 0) \
                - (r['advance_account'] or 0) - (r['diesel'] or 0)
            if net != 0:
                items.append({
                    'date': r['entry_date'], 'kind': 'charge',
                    'label': f'Trip GR-{r["gr_no"] or r["id"]} (net after on-trip advances)',
                    'amount': net, 'link': url_for('ledger_view', le_id=r['id']),
                })
    elif party_type == 'diesel_vendor':
        for r in conn.execute(
            '''SELECT id, gr_no, entry_date, diesel, vehicle_no FROM ledger_entries
               WHERE diesel_vendor_id=? AND diesel > 0
               ORDER BY entry_date DESC, id DESC LIMIT ?''',
            (party_key, limit)
        ).fetchall():
            items.append({
                'date': r['entry_date'], 'kind': 'charge',
                'label': f'Diesel for GR-{r["gr_no"] or r["id"]} ({r["vehicle_no"] or "—"})',
                'amount': r['diesel'] or 0, 'link': url_for('ledger_view', le_id=r['id']),
            })

    # Payments
    for r in conn.execute(
        '''SELECT id, payment_date, amount, mode, reference, notes
           FROM payments WHERE party_type=? AND party_key=?
           ORDER BY payment_date DESC, id DESC LIMIT ?''',
        (party_type, str(party_key), limit)
    ).fetchall():
        label = 'Payment ' + ('received' if party_type == 'client' else 'paid')
        if r['mode']: label += f' ({r["mode"]})'
        if r['reference']: label += f' • ref {r["reference"]}'
        items.append({
            'date': r['payment_date'], 'kind': 'payment',
            'label': label + (' — ' + r['notes'] if r['notes'] else ''),
            'amount': -(r['amount'] or 0),   # negative = reduces balance
            'payment_id': r['id'],
        })

    conn.close()
    items.sort(key=lambda x: (x.get('date') or '', x.get('payment_id', 0)), reverse=True)
    return items


def list_clients_with_balance():
    """Distinct clients (recipient_name) ever billed, with their current balance."""
    conn = get_db()
    names = [r[0] for r in conn.execute(
        'SELECT DISTINCT recipient_name FROM bills WHERE recipient_name IS NOT NULL'
    ).fetchall()]
    conn.close()
    out = []
    for n in sorted(names):
        if not n:
            continue
        bal = get_party_balance('client', n)
        out.append({'name': n, 'balance': bal})
    return out


# ── Auth (session-based, real passwords) ─────────────────────────────────────

import hashlib
import secrets
from functools import wraps

# Endpoints that don't require auth. 'setup'/'setup_demo' are the first-run
# wizard, reached before any user exists.
_PUBLIC_ENDPOINTS = {'login', 'static', 'health', 'setup', 'setup_demo', 'set_lang'}


def _hash_password(password):
    """PBKDF2-HMAC-SHA256 with 16-byte salt and 200k iterations (OWASP 2023)."""
    salt = secrets.token_bytes(16)
    h = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 200_000)
    return base64.b64encode(salt + h).decode()


def _verify_password(password, stored):
    try:
        raw = base64.b64decode(stored.encode())
    except Exception:
        return False
    if len(raw) < 32:
        return False
    salt, h = raw[:16], raw[16:]
    new_h = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 200_000)
    return secrets.compare_digest(h, new_h)


# ── Login brute-force throttling (in-memory; resets on app restart) ───────────
# After _LOGIN_MAX_FAILURES wrong passwords within the window, the account is
# locked for the lockout period. In-memory is fine for Munshi's threat model
# (a single laptop, not exposed to the internet); every failed attempt is also
# written to the audit log, which DOES persist across restarts.
_LOGIN_MAX_FAILURES = 5
_LOGIN_WINDOW_SECONDS = 15 * 60       # count failures within this rolling window
_LOGIN_LOCKOUT_SECONDS = 15 * 60      # lock the account this long once threshold is hit
_login_failures = {}  # username -> list of datetime failure timestamps


def _login_lockout_remaining(username):
    """Seconds left in the lockout for this username, or 0 if not locked."""
    fails = _login_failures.get(username)
    if not fails:
        return 0
    now = datetime.now()
    recent = [t for t in fails if (now - t).total_seconds() < _LOGIN_WINDOW_SECONDS]
    _login_failures[username] = recent
    if len(recent) >= _LOGIN_MAX_FAILURES:
        # The lockout clock starts at the most recent failure.
        locked_until = recent[-1] + timedelta(seconds=_LOGIN_LOCKOUT_SECONDS)
        return max(0, int((locked_until - now).total_seconds()))
    return 0


def _record_login_failure(username):
    if not username:
        return
    _login_failures.setdefault(username, []).append(datetime.now())


def _clear_login_failures(username):
    _login_failures.pop(username, None)


def current_user():
    return session.get('user') or ''


def current_user_role():
    return session.get('role') or ''


@app.context_processor
def inject_auth():
    return {
        'current_user': current_user(),
        'current_user_role': current_user_role(),
        'is_admin': current_user_role() == 'admin',
    }


@app.context_processor
def inject_supplier_and_helpers():
    """Make {{ supplier }} (JL identity) + {{ amount_in_words }} +
       {{ gst_state_name }} available to every template."""
    return {
        'supplier': get_supplier_identity(),
        'amount_in_words': amount_in_words_inr,
        'gst_state_name': lambda code: GST_STATE_NAMES.get((code or '').strip(), ''),
    }


@app.context_processor
def inject_csrf_token():
    """Expose {{ csrf_token() }} to every template (including standalone
       login.html, which doesn't extend base.html)."""
    return {'csrf_token': lambda: session.get('csrf_token', '')}


@app.before_request
def _seed_csrf_token():
    """Ensure every session has a CSRF token, even before login — GET /login
       renders the form, which needs the token. Runs FIRST among the
       before_request hooks (defined before _require_login) so the token exists
       before any form renders. Idempotent: once set, we never rotate it within
       the session, so multiple open tabs keep working."""
    if not session.get('csrf_token'):
        session['csrf_token'] = secrets.token_urlsafe(32)


@app.before_request
def _require_setup():
    """First-run gate. Runs BEFORE _require_login (registered earlier). On a
       fresh, unconfigured install every page redirects to the /setup wizard —
       a blank install must land on /setup, NOT /login. Once setup is complete
       this guard is a no-op and never interferes."""
    ep = request.endpoint
    if ep in ('setup', 'setup_demo', 'static', 'health', 'set_lang'):
        return None
    if ep is None:
        return None
    if _setup_complete():
        return None
    # Not set up yet → send everything to the wizard.
    return redirect(url_for('setup'))


@app.before_request
def _require_login():
    # Skip the guard for public endpoints and any unrouted request
    if request.endpoint is None or request.endpoint in _PUBLIC_ENDPOINTS:
        return None
    if not session.get('user'):
        # Don't redirect form/JSON POSTs, return 401-ish — but for normal GETs send to login
        if request.method == 'GET':
            return redirect(url_for('login', next=request.url))
        return redirect(url_for('login'))
    # First-login password change: force the change page until done
    if session.get('must_change_password') and request.endpoint != 'change_password':
        return redirect(url_for('change_password'))


@app.route('/setup', methods=['GET', 'POST'])
def setup():
    """First-run wizard: the buyer names their firm and creates the owner login.
       If setup is already complete, there's nothing to do here — go home."""
    if _setup_complete():
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        # Company / firm identity
        company    = (request.form.get('company_name') or '').strip()
        gstin      = (request.form.get('gstin') or '').strip().upper()
        pan        = (request.form.get('pan') or '').strip().upper()
        address    = (request.form.get('address') or '').strip()
        state_code = (request.form.get('state_code') or '').strip()
        phone      = (request.form.get('phone') or '').strip()
        # Owner login
        username   = (request.form.get('username') or '').strip()
        password   = request.form.get('password') or ''
        confirm    = request.form.get('confirm_password') or ''

        # ── Validate (friendly, minimal) ──
        error = None
        if not company:
            error = 'Please enter your company / firm name.'
        elif not username:
            error = 'Please choose a username for your owner login.'
        elif len(password) < 4:
            error = 'Password must be at least 4 characters.'
        elif password != confirm:
            error = 'The two passwords do not match.'

        # Soft note on GSTIN length (don't hard-block — some firms aren't GST-registered)
        gstin_note = None
        if gstin and len(gstin) != 15:
            gstin_note = 'Heads up: a GSTIN is normally 15 characters. Saved as entered.'

        if error:
            return render_template('setup.html', error=error, form=request.form)

        # ── Save firm identity ──
        set_setting('supplier_name',       company)
        set_setting('supplier_gstin',      gstin)
        set_setting('supplier_pan',        pan)
        set_setting('supplier_address',    address)
        set_setting('supplier_state_code', state_code)
        set_setting('supplier_phone',      phone)

        # ── Create the owner account (admin, ready to use immediately) ──
        conn = get_db()
        existing = conn.execute('SELECT 1 FROM users WHERE username=?', (username,)).fetchone()
        if existing:
            conn.close()
            return render_template('setup.html',
                                   error='That username already exists. Pick another.',
                                   form=request.form)
        conn.execute(
            '''INSERT INTO users (username, password_hash, full_name, role,
                                  is_active, must_change_password, created_at)
               VALUES (?,?,?,?,1,0,?)''',
            (username, _hash_password(password), username, 'admin',
             datetime.now().isoformat()))
        try:
            log_audit(conn, 'setup', 'user', None,
                      summary=f'Initial setup completed by {username}', user=username)
        except Exception:
            pass
        conn.commit()
        conn.close()

        # ── Mark setup done + sign the owner in ──
        set_setting('setup_complete', '1')
        session.clear()
        session['user'] = username
        session['role'] = 'admin'
        session['must_change_password'] = False
        session.permanent = True
        if gstin_note:
            flash(gstin_note)
        flash(f'Welcome to Munshi, {company}! Your account is ready.')
        return redirect(url_for('dashboard'))

    return render_template('setup.html', error=None, form={})


@app.route('/setup/demo', methods=['POST'])
def setup_demo():
    """Load the sample demo dataset for a sales pitch, then hand off to login
       (demo user is 'Demo' / 'Demo'). Copies rows out of data/seed_demo.db into
       the live DB. Wrapped so a failure lands the user back on /setup safely."""
    if _setup_complete():
        return redirect(url_for('login'))

    demo_path = os.path.join(APP_DIR, 'data', 'seed_demo.db')
    try:
        bundle_dir = getattr(sys, '_MEIPASS', None)
    except Exception:
        bundle_dir = None
    if bundle_dir:
        bundled = os.path.join(bundle_dir, 'data', 'seed_demo.db')
        if os.path.exists(bundled):
            demo_path = bundled

    if not os.path.exists(demo_path):
        flash('Sample demo data is not available in this copy.')
        return redirect(url_for('setup'))

    # Tables to copy from the demo DB into the live DB. settings + users are
    # REPLACE-merged (demo identity wins); operational tables are appended.
    tables = ['settings', 'users', 'freight_rates', 'transporters',
              'diesel_vendors', 'recipients', 'vehicles', 'drivers',
              'bills', 'challans', 'ledger_entries', 'payments']
    try:
        src = sqlite3.connect(demo_path)
        src.row_factory = sqlite3.Row
        dst = get_db()
        for t in tables:
            # Only copy tables that exist in BOTH DBs.
            src_has = src.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (t,)
            ).fetchone()
            dst_has = dst.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (t,)
            ).fetchone()
            if not (src_has and dst_has):
                continue
            src_cols = [r['name'] for r in src.execute(f'PRAGMA table_info({t})').fetchall()]
            dst_cols = {r['name'] for r in dst.execute(f'PRAGMA table_info({t})').fetchall()}
            cols = [c for c in src_cols if c in dst_cols]
            if not cols:
                continue
            col_list = ','.join(f'"{c}"' for c in cols)
            placeholders = ','.join('?' for _ in cols)
            verb = 'INSERT OR REPLACE' if t in ('settings', 'users') else 'INSERT OR IGNORE'
            for row in src.execute(f'SELECT {col_list} FROM "{t}"').fetchall():
                dst.execute(
                    f'{verb} INTO "{t}" ({col_list}) VALUES ({placeholders})',
                    tuple(row[c] for c in cols))
        dst.execute('INSERT OR REPLACE INTO settings VALUES (?,?)', ('setup_complete', '1'))
        dst.execute('INSERT OR REPLACE INTO settings VALUES (?,?)', ('is_demo', '1'))
        dst.commit()
        dst.close()
        src.close()
    except Exception as e:
        app.logger.warning(f'setup_demo failed: {e}')
        flash('Sorry, loading the sample demo data failed. Please try again or set up your own firm.')
        return redirect(url_for('setup'))

    flash('Sample demo data loaded. Sign in with username "Demo" and password "Demo".')
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    raw_next = request.args.get('next') or request.form.get('next') or ''
    # Only allow relative in-site redirects (block open-redirect phishing):
    # a next URL must start with a single '/', not '//' (protocol-relative)
    # and not an absolute URL like https://evil.com.
    next_url = raw_next if (raw_next.startswith('/') and not raw_next.startswith('//')) else url_for('dashboard')
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''
        # Brute-force throttle: if this account has had too many recent wrong
        # attempts, refuse (and don't even check the password) until it cools down.
        wait = _login_lockout_remaining(username)
        if wait > 0:
            mins = max(1, wait // 60)
            flash(f'Too many wrong attempts. Please wait about {mins} minute(s) before trying again.')
            return render_template('login.html', next_url=next_url)
        conn = get_db()
        row = conn.execute(
            'SELECT * FROM users WHERE username=? AND is_active=1', (username,)).fetchone()
        if row and _verify_password(password, row['password_hash']):
            _lang = session.get('lang')          # preserve chosen language across the reset
            session.clear()
            if _lang:
                session['lang'] = _lang
            session['user'] = row['username']
            session['role'] = row['role'] or 'operator'
            session['must_change_password'] = bool(row['must_change_password'])
            session.permanent = True  # respect PERMANENT_SESSION_LIFETIME
            conn.execute('UPDATE users SET last_login=? WHERE username=?',
                         (datetime.now().isoformat(), row['username']))
            log_audit(conn, 'login', 'user', None,
                      summary=f'User {row["username"]} signed in', user=row['username'])
            conn.commit()
            conn.close()
            _clear_login_failures(username)
            return redirect(next_url)
        # Failed sign-in: record the attempt for throttling, audit it, then commit.
        _record_login_failure(username)
        log_audit(conn, 'login_failed', 'user', None,
                  summary=f'Failed sign-in for "{username or "unknown"}"',
                  user=(username or 'unknown'))
        conn.commit()
        conn.close()
        flash(t('Invalid username or password.'))
    return render_template('login.html', next_url=next_url)


@app.route('/logout', methods=['GET', 'POST'])
def logout():
    user = current_user()
    if user:
        try:
            conn = get_db()
            log_audit(conn, 'logout', 'user', None,
                      summary=f'User {user} signed out', user=user)
            conn.commit()
            conn.close()
        except Exception:
            pass
    _lang = session.get('lang')                  # keep the language choice after sign-out
    session.clear()
    if _lang:
        session['lang'] = _lang
    flash(t('Signed out.'))
    return redirect(url_for('login'))


@app.route('/lang/<code>')
def set_lang(code):
    """Flip the whole app frame between English and Hindi. Public + pre-setup so
       the language can be chosen on the login / setup screens too. Redirects
       back to where the user was (same-site only, to block open-redirects)."""
    if code in SUPPORTED_LANGS:
        session['lang'] = code
    ref = request.referrer
    if ref and ref.startswith(request.host_url):
        return redirect(ref)
    return redirect(url_for('dashboard'))


@app.route('/change-password', methods=['GET', 'POST'])
def change_password():
    if not session.get('user'):
        return redirect(url_for('login'))
    if request.method == 'POST':
        old = request.form.get('old_password') or ''
        new = request.form.get('new_password') or ''
        confirm = request.form.get('confirm_password') or ''
        if len(new) < 6:
            flash(t('New password must be at least 6 characters.'))
            return redirect(url_for('change_password'))
        if new.lower() == (session.get('user') or '').lower():
            flash('Your password cannot be the same as your username. Pick something different.')
            return redirect(url_for('change_password'))
        if new != confirm:
            flash(t('New passwords do not match.'))
            return redirect(url_for('change_password'))
        conn = get_db()
        row = conn.execute('SELECT password_hash FROM users WHERE username=?',
                           (session['user'],)).fetchone()
        if not row or not _verify_password(old, row['password_hash']):
            conn.close()
            flash(t('Current password is wrong.'))
            return redirect(url_for('change_password'))
        conn.execute('UPDATE users SET password_hash=?, must_change_password=0 WHERE username=?',
                     (_hash_password(new), session['user']))
        conn.commit()
        conn.close()
        session['must_change_password'] = False
        flash(t('Password updated.'))
        return redirect(url_for('dashboard'))
    return render_template('change_password.html',
                           must_change=session.get('must_change_password'))


# ── License enforcement (Munshi commercial — phone-home + kill-switch) ──────

LICENSE_SERVER_URL = os.environ.get('LICENSE_SERVER_URL', '').strip()
LICENSE_PHONE_HOME_HOURS = int(os.environ.get('LICENSE_PHONE_HOME_HOURS', '168'))
# Opt-in license enforcement (default OFF). When ON, the app blocks writes even
# if no license server/key is configured (status 'unconfigured') — so a paying
# customer can't bypass payment by simply emptying LICENSE_SERVER_URL. OFF by
# default so free-trial pilots and the in-house family install keep running
# unrestricted. Flip to ON (MUNSHI_REQUIRE_LICENSE=1) only when you start charging.
REQUIRE_LICENSE = os.environ.get('MUNSHI_REQUIRE_LICENSE', '').strip().lower() in ('1', 'true', 'yes', 'on')
# Default weekly (168h), not daily — offline-friendly and reinforces the privacy
# promise. The check only ever sends the license key + a truck count (no ledger
# data), and never locks the app while offline (cached state persists).


def _get_license_state():
    conn = get_db()
    row = conn.execute('SELECT * FROM license_state WHERE id=1').fetchone()
    conn.close()
    return dict(row) if row else {'status': 'unconfigured'}


def _save_license_state(updates):
    """Update license_state singleton with `updates` dict."""
    if not updates:
        return
    cols = ','.join(f'{k}=?' for k in updates)
    conn = get_db()
    conn.execute(f'UPDATE license_state SET {cols} WHERE id=1',
                 tuple(updates.values()))
    conn.commit()
    conn.close()


def _phone_home_to_license_server():
    """Call /verify on the license server and cache the result locally.
       Quiet best-effort — never raises; offline runs keep using cached state."""
    if not LICENSE_SERVER_URL:
        return None
    state = _get_license_state()
    key = (state.get('license_key') or get_setting('license_key') or '').strip().upper()
    if not key:
        _save_license_state({
            'status': 'unconfigured',
            'message': 'No license key configured.',
            'last_checked_at': datetime.now().isoformat(),
            'last_error': None,
        })
        return None
    try:
        import urllib.request, urllib.error
        payload = json.dumps({
            'license_key': key,
            'truck_count': _count_active_vehicles(),
            'client_version': 'munshi-1.0',
        }).encode('utf-8')
        req = urllib.request.Request(
            LICENSE_SERVER_URL.rstrip('/') + '/verify',
            data=payload, method='POST',
            headers={'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        # 404 = not found; we still record it
        try:
            data = json.loads(e.read().decode('utf-8'))
        except Exception:
            data = {'status': 'unreachable', 'message': f'License server error: HTTP {e.code}'}
    except Exception as e:
        # Network down / DNS fail / server unreachable → keep cached state intact,
        # just bump last_error & last_checked_at
        _save_license_state({
            'last_checked_at': datetime.now().isoformat(),
            'last_error': str(e)[:200],
        })
        return None
    # Cache the response
    _save_license_state({
        'license_key':    key,
        'status':         data.get('status') or 'unknown',
        'tier':           data.get('tier'),
        'max_trucks':     data.get('max_trucks'),
        'expires_at':     data.get('expires_at'),
        'days_to_expiry': data.get('days_to_expiry'),
        'message':        data.get('message') or '',
        'last_checked_at': datetime.now().isoformat(),
        'last_error':     None,
    })
    return data


def _count_active_vehicles():
    """Simple truck-count metric for license server analytics."""
    try:
        conn = get_db()
        n = conn.execute('SELECT COUNT(*) FROM vehicles').fetchone()[0]
        conn.close()
        return int(n or 0)
    except Exception:
        return 0


# DENY-BY-DEFAULT: when license status is locked/expired/suspended, EVERY POST
# request is blocked except those in this explicit allow-list. This is safer
# than an allow-list of "to-block" endpoints because new POST routes added
# later are automatically protected.
#
# Allowed POSTs when locked:
#   - auth flow (so the user can sign in, change pwd, sign out)
#   - license management (so they can enter a renewed key)
#   - report generation (read-only despite being POST — summary/PDF export)
_POSTS_ALLOWED_WHEN_LOCKED = {
    'login', 'logout', 'change_password',
    'license_set', 'license_check_now',
    'summary',          # POST /summary just renders a date-range report, no DB writes
    'set_user',         # legacy, no-op-ish
    # Drive backup endpoints — data preservation > license enforcement.
    # Customer in license-lockout can still recover their data via Drive.
    'drive_disconnect', 'drive_sync_now',
}

# Endpoints exempted from the lockout check entirely (every method).
_LICENSE_CHECK_EXEMPT = {'license_page', 'license_check_now', 'license_set',
                        'login', 'logout', 'change_password',
                        'setup', 'setup_demo',   # first-run wizard
                        'health', 'static', 'serve_upload',
                        # Drive flows skip lockout — see above rationale
                        'drive_connect', 'drive_oauth_callback',
                        'drive_disconnect', 'drive_sync_now'}


@app.before_request
def _enforce_license_lockout():
    """When license is locked/expired/suspended/not_found, block every POST/PUT/
       DELETE except the explicit auth + license-management allow-list.
       GET requests always pass — users keep read access to their own data."""
    ep = request.endpoint
    if ep is None or ep in _LICENSE_CHECK_EXEMPT:
        return None
    if request.method in ('GET', 'HEAD', 'OPTIONS'):
        return None
    state = _get_license_state()
    status = state.get('status') or 'unconfigured'
    # 'not_found' is included: if the server doesn't recognise the key, we
    # treat it as a forged/wiped license and block writes too.
    # With MUNSHI_REQUIRE_LICENSE=1 (opt-in), also block when there's no license
    # configured at all — closes the "delete the env line to bypass" hole for
    # paying customers. Default off so pilots/in-house stay unrestricted.
    blocked = status in ('locked', 'expired', 'suspended', 'not_found') or \
              (REQUIRE_LICENSE and status == 'unconfigured')
    if blocked:
        if ep not in _POSTS_ALLOWED_WHEN_LOCKED:
            if status == 'unconfigured':
                flash('Munshi is in read-only mode. Enter a license key to continue making changes.')
            else:
                flash('Munshi is in read-only mode. Renew your license to continue making changes.')
            return redirect(url_for('license_page'))


@app.context_processor
def inject_license_state():
    return {'license_state': _get_license_state()}


@app.context_processor
def inject_drive_state():
    """Make {{ drive_state }} available everywhere (used by Settings card +
       global re-auth banner)."""
    return {
        'drive_state':   _get_drive_state(),
        'drive_configured': bool(GOOGLE_OAUTH_CLIENT_ID and GOOGLE_OAUTH_CLIENT_SECRET),
    }


# Periodic phone-home: trigger from any request when last_checked_at is stale.
@app.before_request
def _maybe_phone_home():
    if not LICENSE_SERVER_URL:
        return None
    if request.endpoint in ('static', 'extract_status'):
        return None
    state = _get_license_state()
    last = state.get('last_checked_at')
    needs_check = True
    if last:
        try:
            then = datetime.fromisoformat(last)
            if (datetime.now() - then).total_seconds() < LICENSE_PHONE_HOME_HOURS * 3600:
                needs_check = False
        except Exception:
            pass
    if needs_check:
        # Fire in a background thread so the request doesn't wait on the network
        import threading
        threading.Thread(target=_phone_home_to_license_server, daemon=True).start()


@app.before_request
def _maybe_daily_backup():
    """Cheap check on every request: if today's local backup doesn't exist OR
       Drive sync is connected but didn't run today, dispatch backup_db_if_needed
       (which handles both local + Drive in a thread). Cost: 1 os.path.exists()
       + (rarely) 1 DB row read.

       Backoff: if the previous Drive sync errored, retry at most every 5
       minutes. Prevents a request burst (e.g. the extraction-polling page)
       from spawning a Drive-upload thread storm when Drive is unreachable."""
    if request.endpoint in ('static', 'extract_status'):
        return None
    today = datetime.now().strftime('%Y-%m-%d')
    backup_path = os.path.join(BACKUP_DIR, f'bills-{today}.db')
    if os.path.exists(backup_path):
        # Local backup done. Did Drive sync today? If not (and connected), trigger.
        if _drive_is_connected():
            try:
                state = _get_drive_state()
                last = state.get('last_uploaded_file') or ''
                already_today = (last == f'bills-{today}.db' and
                                 state.get('last_sync_status') == 'ok')
                if already_today:
                    return None
                # Backoff on repeated failures: don't retry sooner than 5 min
                last_at = state.get('last_sync_at') or ''
                if (state.get('last_sync_status') != 'ok' and last_at):
                    try:
                        then = datetime.fromisoformat(last_at)
                        if (datetime.utcnow() - then).total_seconds() < 300:
                            return None  # still in backoff window
                    except Exception:
                        pass
                import threading
                threading.Thread(target=_drive_sync_now, daemon=True).start()
            except Exception:
                pass
        return None
    # No local backup yet today — make it now (also dispatches Drive sync).
    try:
        backup_db_if_needed()
    except Exception as e:
        app.logger.warning(f'_maybe_daily_backup failed: {e}')


# ── CSRF protection ─────────────────────────────────────────────────────────
# Endpoints exempt from the CSRF check entirely. GET/HEAD/OPTIONS are always
# skipped by the method check below; this set is belt-and-suspenders for
# clarity (drive_oauth_callback is GET-only anyway).
_CSRF_EXEMPT = {'drive_oauth_callback', 'static', 'health'}


@app.before_request
def _enforce_csrf():
    """Reject POST/PUT/PATCH/DELETE without a valid CSRF token. Runs LAST
       among the before_request hooks (defined after _maybe_daily_backup) so
       _require_login redirects unauthenticated POSTs to /login (not a 400) and
       _enforce_license_lockout redirects locked-out users to /license (not a
       400). Auth/license POSTs that ARE allowed still get CSRF-checked — they
       carry tokens like every other form."""
    if request.method not in ('POST', 'PUT', 'PATCH', 'DELETE'):
        return None
    ep = request.endpoint
    if ep is None or ep in _CSRF_EXEMPT:
        return None
    sent = request.form.get('csrf_token') or request.headers.get('X-CSRFToken', '') or ''
    expected = session.get('csrf_token', '') or ''
    if not expected or not secrets.compare_digest(sent, expected):
        return render_template('error.html',
            code=400, title='Security check failed',
            message='Your login session expired or this form was submitted from '
                    'another page. Please go back, refresh the page, and try again.'
        ), 400


@app.route('/license')
def license_page():
    state = _get_license_state()
    return render_template('license.html', state=state,
                           server_url=LICENSE_SERVER_URL,
                           configured_key=(get_setting('license_key') or ''))


@app.route('/license/set', methods=['POST'])
def license_set():
    if current_user_role() != 'admin':
        flash('Only admins can change the license key.')
        return redirect(url_for('license_page'))
    key = (request.form.get('license_key') or '').strip().upper()
    set_setting('license_key', key)
    _save_license_state({'license_key': key, 'last_checked_at': None})
    # Synchronously phone home now so the user sees the result immediately
    _phone_home_to_license_server()
    flash('License key saved. Verifying with server…')
    return redirect(url_for('license_page'))


@app.route('/license/check', methods=['POST'])
def license_check_now():
    _phone_home_to_license_server()
    flash('License verification refreshed.')
    return redirect(url_for('license_page'))


# ── Google Drive backup (per-customer, OAuth 2.0) ───────────────────────────
#
# Each Munshi customer connects their OWN Google account; we upload the daily
# SQLite backup to a folder in their personal Drive. Our scope is `drive.file`
# — we can only see/manage files we created, never their other Drive content.
# Customer data never touches Munshi infrastructure.

GOOGLE_OAUTH_CLIENT_ID     = os.environ.get('GOOGLE_OAUTH_CLIENT_ID', '').strip()
GOOGLE_OAUTH_CLIENT_SECRET = os.environ.get('GOOGLE_OAUTH_CLIENT_SECRET', '').strip()
DRIVE_SCOPES               = ['https://www.googleapis.com/auth/drive.file']
DRIVE_BACKUP_FOLDER_NAME   = 'Munshi Backups'

# Locally-bound redirect; works with Google Cloud OAuth client of type "Desktop"
# OR "Web" (loopback addresses are allowed as redirect URIs for both).
def _drive_redirect_uri():
    """Build the OAuth redirect URI based on the request's host. Falls back to
       localhost:5056 if called outside a request context."""
    try:
        # Reconstruct from the live request so the port matches whatever we're
        # actually running on (port can differ per install).
        return f'{request.scheme}://{request.host}/oauth/google/callback'
    except RuntimeError:
        port = os.environ.get('PORT', '5056')
        return f'http://127.0.0.1:{port}/oauth/google/callback'


def _drive_client_config():
    """Synthesize a Flow.from_client_config dict from env vars.
       Uses the `installed` key because we create a Google Cloud OAuth client
       of type 'Desktop application' (per OPS_RUNBOOK Section 2.4)."""
    return {
        'installed': {
            'client_id':     GOOGLE_OAUTH_CLIENT_ID,
            'client_secret': GOOGLE_OAUTH_CLIENT_SECRET,
            'auth_uri':      'https://accounts.google.com/o/oauth2/auth',
            'token_uri':     'https://oauth2.googleapis.com/token',
            'redirect_uris': [_drive_redirect_uri()],
        }
    }


def _get_drive_state():
    conn = get_db()
    row = conn.execute('SELECT * FROM google_drive_state WHERE id=1').fetchone()
    conn.close()
    return dict(row) if row else {'last_sync_status': 'not_configured'}


def _save_drive_state(updates):
    if not updates:
        return
    cols = ','.join(f'{k}=?' for k in updates)
    conn = get_db()
    conn.execute(f'UPDATE google_drive_state SET {cols} WHERE id=1',
                 tuple(updates.values()))
    conn.commit()
    conn.close()


def _drive_oauth_flow(state=None):
    """Build a google_auth_oauthlib Flow object with PKCE enabled by default
       (PKCE comes built-in to google-auth-oauthlib >= 1.0)."""
    from google_auth_oauthlib.flow import Flow
    flow = Flow.from_client_config(
        _drive_client_config(),
        scopes=DRIVE_SCOPES,
        state=state,
    )
    flow.redirect_uri = _drive_redirect_uri()
    return flow


def _drive_credentials_from_state(state):
    """Build a google.oauth2.credentials.Credentials from the stored state dict.
       Returns None if no refresh_token (i.e. user not connected)."""
    from google.oauth2.credentials import Credentials
    if not state or not state.get('refresh_token'):
        return None
    creds = Credentials(
        token=state.get('access_token'),
        refresh_token=state['refresh_token'],
        token_uri='https://oauth2.googleapis.com/token',
        client_id=GOOGLE_OAUTH_CLIENT_ID,
        client_secret=GOOGLE_OAUTH_CLIENT_SECRET,
        scopes=DRIVE_SCOPES,
    )
    return creds


def _drive_service():
    """Returns an authorized Drive v3 service, refreshing the access_token if
       it's within 60 seconds of expiry. On RefreshError (invalid_grant) marks
       state='reauth_required' and returns None."""
    from googleapiclient.discovery import build
    from google.auth.transport.requests import Request as GAuthRequest
    from google.auth.exceptions import RefreshError

    state = _get_drive_state()
    creds = _drive_credentials_from_state(state)
    if creds is None:
        return None

    # Refresh proactively (60s skew) — Credentials.expired only flips when expired
    needs_refresh = True
    if creds.token and state.get('access_token_expiry'):
        try:
            expiry = datetime.fromisoformat(state['access_token_expiry'])
            if expiry - datetime.utcnow() > timedelta(seconds=60):
                needs_refresh = False
        except Exception:
            pass

    if needs_refresh:
        try:
            creds.refresh(GAuthRequest())
        except RefreshError as e:
            _save_drive_state({
                'last_sync_status': 'reauth_required',
                'last_sync_error':  f'Google revoked our access. Please reconnect. ({e})'[:300],
                'access_token':     None,
                'refresh_token':    None,
                'access_token_expiry': None,
            })
            return None
        # Cache the freshly-refreshed token
        _save_drive_state({
            'access_token':        creds.token,
            'access_token_expiry': creds.expiry.isoformat() if creds.expiry else None,
        })
    return build('drive', 'v3', credentials=creds, cache_discovery=False)


def _ensure_munshi_folder(svc, state):
    """Make sure the backup folder still exists & isn't trashed. With
       `drive.file` scope we can only see files our app created, so we cannot
       search by name across the user's Drive — if the folder is missing /
       trashed, we create a fresh one and persist the new ID."""
    folder_id = state.get('folder_id')

    if folder_id:
        try:
            meta = svc.files().get(fileId=folder_id, fields='id,name,trashed').execute()
            if meta and not meta.get('trashed'):
                return folder_id
        except Exception:
            pass  # 404 — folder gone; fall through to recreate

    # Create a fresh folder
    new = svc.files().create(body={
        'name':     DRIVE_BACKUP_FOLDER_NAME,
        'mimeType': 'application/vnd.google-apps.folder',
    }, fields='id,name').execute()
    _save_drive_state({
        'folder_id':   new['id'],
        'folder_name': new['name'],
    })
    return new['id']


def _drive_file_exists_in_folder(svc, folder_id, filename):
    """Idempotency check: returns True if a file with this name already exists
       in our backup folder (and isn't trashed). Prevents duplicate uploads."""
    q = (f"name='{filename}' and '{folder_id}' in parents "
         f"and trashed=false")
    try:
        res = svc.files().list(q=q, fields='files(id,name)', pageSize=1).execute()
        return bool(res.get('files'))
    except Exception:
        return False  # On error, allow upload to proceed; caller will see the failure


def _drive_upload_backup(file_path):
    """Upload one backup file to the customer's Drive folder. Resumable upload
       (handles >5 MB files cleanly). Idempotent — skips if same filename
       already in folder."""
    from googleapiclient.http import MediaFileUpload
    from googleapiclient.errors import HttpError

    if not os.path.exists(file_path):
        return False, 'File missing on disk'
    if not GOOGLE_OAUTH_CLIENT_ID or not GOOGLE_OAUTH_CLIENT_SECRET:
        return False, 'OAuth credentials not configured'

    svc = _drive_service()
    if svc is None:
        return False, 'Not connected or token revoked'

    state = _get_drive_state()
    try:
        folder_id = _ensure_munshi_folder(svc, state)
    except Exception as e:
        _save_drive_state({
            'last_sync_status': 'error',
            'last_sync_error':  f'Folder check failed: {e}'[:300],
            'last_sync_at':     datetime.utcnow().isoformat(),
        })
        return False, str(e)

    fname = os.path.basename(file_path)
    if _drive_file_exists_in_folder(svc, folder_id, fname):
        _save_drive_state({
            'last_sync_status':   'ok',
            'last_sync_at':       datetime.utcnow().isoformat(),
            'last_uploaded_file': fname,
            'last_sync_error':    None,
        })
        return True, 'Already uploaded'

    media = MediaFileUpload(file_path, mimetype='application/octet-stream',
                            resumable=True)
    try:
        svc.files().create(body={
            'name':    fname,
            'parents': [folder_id],
        }, media_body=media, fields='id,name').execute()
    except HttpError as e:
        is_quota = (e.resp.status == 403 and 'quota' in str(e).lower())
        _save_drive_state({
            'last_sync_status':  'quota' if is_quota else 'error',
            'last_sync_error':   f'Drive HTTP {e.resp.status}: {e}'[:300],
            'last_sync_at':      datetime.utcnow().isoformat(),
        })
        return False, str(e)
    except Exception as e:
        _save_drive_state({
            'last_sync_status':  'offline',
            'last_sync_error':   str(e)[:300],
            'last_sync_at':      datetime.utcnow().isoformat(),
        })
        return False, str(e)

    _save_drive_state({
        'last_sync_status':   'ok',
        'last_sync_at':       datetime.utcnow().isoformat(),
        'last_uploaded_file': fname,
        'last_sync_error':    None,
    })
    return True, fname


import threading as _threading_mod
# Module-level singleton — only one Drive sync may run at a time.
# Non-blocking acquire: if a sync is already in flight, the new caller bails
# immediately rather than queueing. Prevents a thread storm when many requests
# come in while Drive is unreachable.
_drive_sync_lock = _threading_mod.Lock()


def _drive_sync_now():
    """Upload today's local backup file (if any) to Drive. Safe to call from
       a daemon thread or directly. No-op when not configured / not connected.
       Re-entrant safe: if another sync is already running, this call exits
       immediately."""
    if not _drive_sync_lock.acquire(blocking=False):
        return  # another sync is already in flight; nothing to do
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        path = os.path.join(BACKUP_DIR, f'bills-{today}.db')
        if not os.path.exists(path):
            return  # No local backup yet — let backup_db_if_needed create it first
        try:
            _drive_upload_backup(path)
        except Exception as e:
            app.logger.warning(f'drive sync failed: {e}')
    finally:
        _drive_sync_lock.release()


def _drive_is_connected():
    """True if the customer has connected Drive at least once and tokens are
       still valid (or refreshable)."""
    state = _get_drive_state()
    return bool(state.get('refresh_token'))


# ── Drive routes ────────────────────────────────────────────────────────────

@app.route('/settings/drive/connect')
def drive_connect():
    """Start the OAuth flow: build auth URL with state, redirect to Google."""
    if current_user_role() != 'admin':
        flash('Only admins can connect Google Drive.')
        return redirect(url_for('settings'))
    if not GOOGLE_OAUTH_CLIENT_ID or not GOOGLE_OAUTH_CLIENT_SECRET:
        flash('Administrator hasn\'t configured Google OAuth credentials yet. See OPS_RUNBOOK.md.')
        return redirect(url_for('settings'))
    flow = _drive_oauth_flow()
    auth_url, state = flow.authorization_url(
        access_type='offline',          # required to get a refresh_token
        prompt='consent',               # force consent every time so we always get a refresh_token
        include_granted_scopes='true',
    )
    # PKCE: persist the code_verifier across requests. The verifier was
    # generated by authorization_url() and stored on the Flow object; without
    # this, the callback's fresh Flow has no verifier and Google rejects the
    # token exchange with 'code_verifier failed'.
    session['drive_oauth_state']    = state
    session['drive_code_verifier']  = getattr(flow, 'code_verifier', None)
    return redirect(auth_url)


@app.route('/oauth/google/callback')
def drive_oauth_callback():
    """Verify state, exchange code for tokens, create folder, persist state."""
    if current_user_role() != 'admin':
        flash('Only admins can complete the Drive connection.')
        return redirect(url_for('settings'))
    expected_state    = session.pop('drive_oauth_state', None)
    expected_verifier = session.pop('drive_code_verifier', None)
    actual_state      = request.args.get('state')
    if not expected_state or actual_state != expected_state:
        return ('OAuth state mismatch — possible CSRF. Try connecting again.', 400)
    if request.args.get('error'):
        flash(f'Google rejected the connection: {request.args.get("error")}')
        return redirect(url_for('settings'))

    flow = _drive_oauth_flow(state=expected_state)
    # Restore the PKCE verifier onto the new Flow so fetch_token sends the same
    # one we computed the challenge from in /drive/connect.
    if expected_verifier:
        flow.code_verifier = expected_verifier
    try:
        # fetch_token expects the full URL with the code + state appended
        flow.fetch_token(authorization_response=request.url)
    except Exception as e:
        flash(f'Token exchange failed: {e}')
        return redirect(url_for('settings'))

    creds = flow.credentials
    _save_drive_state({
        'refresh_token':       creds.refresh_token,
        'access_token':        creds.token,
        'access_token_expiry': creds.expiry.isoformat() if creds.expiry else None,
        'last_sync_status':    'ok',
        'last_sync_error':     None,
    })

    # Fetch the connected email + create the folder
    try:
        from googleapiclient.discovery import build
        svc = build('drive', 'v3', credentials=creds, cache_discovery=False)
        about = svc.about().get(fields='user(emailAddress)').execute()
        email = about.get('user', {}).get('emailAddress', '')
        _save_drive_state({'connected_email': email})
        _ensure_munshi_folder(svc, _get_drive_state())
    except Exception as e:
        app.logger.warning(f'post-connect setup failed: {e}')

    # Audit log — proper conn lifecycle: open, log, commit, close.
    conn = get_db()
    try:
        log_audit(conn, 'connect', 'google_drive', 0,
                  summary=f'Connected Google Drive ({_get_drive_state().get("connected_email") or "?"})')
        conn.commit()
    finally:
        conn.close()
    flash('Google Drive connected. Your daily backups will now sync automatically.')
    return redirect(url_for('settings') + '#drive')


@app.route('/settings/drive/disconnect', methods=['POST'])
def drive_disconnect():
    if current_user_role() != 'admin':
        flash('Only admins can disconnect Google Drive.')
        return redirect(url_for('settings'))
    state = _get_drive_state()
    token = state.get('refresh_token') or state.get('access_token')
    if token:
        # Revoke at Google so it disappears from the user's connected apps list
        try:
            import urllib.request, urllib.parse, urllib.error
            req = urllib.request.Request(
                'https://oauth2.googleapis.com/revoke',
                data=urllib.parse.urlencode({'token': token}).encode(),
                method='POST',
                headers={'Content-Type': 'application/x-www-form-urlencoded'})
            urllib.request.urlopen(req, timeout=8).read()
        except Exception as e:
            app.logger.warning(f'token revoke failed (continuing anyway): {e}')
    # Clear local state regardless of revoke success
    _save_drive_state({
        'refresh_token':       None,
        'access_token':        None,
        'access_token_expiry': None,
        'folder_id':           None,
        'folder_name':         None,
        'connected_email':     None,
        'last_sync_at':        None,
        'last_sync_status':    'not_configured',
        'last_sync_error':     None,
        'last_uploaded_file':  None,
    })
    conn = get_db()
    try:
        log_audit(conn, 'disconnect', 'google_drive', 0,
                  summary='Disconnected Google Drive')
        conn.commit()
    finally:
        conn.close()
    flash('Google Drive disconnected. Your backup files in Drive remain untouched.')
    return redirect(url_for('settings') + '#drive')


@app.route('/settings/drive/sync-now', methods=['POST'])
def drive_sync_now():
    """Trigger an immediate sync in a background thread."""
    if not _drive_is_connected():
        flash('Connect Google Drive first.')
        return redirect(url_for('settings') + '#drive')
    # Make sure today's local backup exists before we sync
    try:
        backup_db_if_needed()
    except Exception:
        pass
    import threading
    threading.Thread(target=_drive_sync_now, daemon=True).start()
    flash('Drive sync started — refresh in a few seconds.')
    return redirect(url_for('settings') + '#drive')


# ── Admin user management ────────────────────────────────────────────────────

def _require_admin():
    """Return None if current user is admin, else a redirect/403 response."""
    if current_user_role() != 'admin':
        flash('You need admin access for that action.')
        return redirect(url_for('dashboard'))
    return None


@app.route('/users')
def users_index():
    guard = _require_admin()
    if guard: return guard
    conn = get_db()
    rows = conn.execute(
        '''SELECT username, full_name, role, is_active, must_change_password,
                  created_at, last_login
           FROM users ORDER BY role DESC, username''').fetchall()
    conn.close()
    return render_template('users.html', users=[dict(r) for r in rows])


@app.route('/users/add', methods=['POST'])
def users_add():
    guard = _require_admin()
    if guard: return guard
    f = request.form
    username = (f.get('username') or '').strip()
    full_name = (f.get('full_name') or '').strip() or username
    role = f.get('role') or 'operator'
    if role not in ('admin', 'operator'):
        role = 'operator'
    if not username:
        flash('Username is required.')
        return redirect(url_for('users_index'))
    if len(username) < 2:
        flash('Username must be at least 2 characters.')
        return redirect(url_for('users_index'))
    conn = get_db()
    existing = conn.execute('SELECT 1 FROM users WHERE username=?', (username,)).fetchone()
    if existing:
        conn.close()
        flash(f'User "{username}" already exists.')
        return redirect(url_for('users_index'))
    # First-login password = username
    conn.execute(
        '''INSERT INTO users (username, password_hash, full_name, role,
                              is_active, must_change_password, created_at)
           VALUES (?,?,?,?,1,1,?)''',
        (username, _hash_password(username), full_name, role, datetime.now().isoformat()))
    log_audit(conn, 'create', 'user', 0,
              summary=f'Created user "{username}" ({role})')
    conn.commit()
    conn.close()
    flash(f'User "{username}" created. First-login password is "{username}" — they must change it.')
    return redirect(url_for('users_index'))


@app.route('/users/<username>/reset-password', methods=['POST'])
def users_reset_password(username):
    guard = _require_admin()
    if guard: return guard
    conn = get_db()
    row = conn.execute('SELECT username FROM users WHERE username=?', (username,)).fetchone()
    if not row:
        conn.close()
        flash('User not found.')
        return redirect(url_for('users_index'))
    # Reset to username — they'll be forced to change on next login
    conn.execute(
        'UPDATE users SET password_hash=?, must_change_password=1 WHERE username=?',
        (_hash_password(username), username))
    log_audit(conn, 'reset_password', 'user', 0,
              summary=f'Reset password for "{username}"')
    conn.commit()
    conn.close()
    flash(f'Password for "{username}" reset to their username. They will be forced to change it on next login.')
    return redirect(url_for('users_index'))


@app.route('/users/<username>/deactivate', methods=['POST'])
def users_deactivate(username):
    guard = _require_admin()
    if guard: return guard
    if username == current_user():
        flash('You cannot deactivate yourself.')
        return redirect(url_for('users_index'))
    conn = get_db()
    conn.execute('UPDATE users SET is_active=0 WHERE username=?', (username,))
    log_audit(conn, 'deactivate', 'user', 0, summary=f'Deactivated user "{username}"')
    conn.commit()
    conn.close()
    flash(f'User "{username}" deactivated. They can no longer sign in.')
    return redirect(url_for('users_index'))


@app.route('/users/<username>/activate', methods=['POST'])
def users_activate(username):
    guard = _require_admin()
    if guard: return guard
    conn = get_db()
    conn.execute('UPDATE users SET is_active=1 WHERE username=?', (username,))
    log_audit(conn, 'activate', 'user', 0, summary=f'Re-activated user "{username}"')
    conn.commit()
    conn.close()
    flash(f'User "{username}" re-activated.')
    return redirect(url_for('users_index'))


@app.route('/users/<username>/role', methods=['POST'])
def users_change_role(username):
    guard = _require_admin()
    if guard: return guard
    new_role = request.form.get('role') or 'operator'
    if new_role not in ('admin', 'operator'):
        new_role = 'operator'
    if username == current_user() and new_role != 'admin':
        flash('You cannot demote yourself — promote another admin first.')
        return redirect(url_for('users_index'))
    conn = get_db()
    conn.execute('UPDATE users SET role=? WHERE username=?', (new_role, username))
    log_audit(conn, 'change_role', 'user', 0,
              summary=f'Changed "{username}" role → {new_role}')
    conn.commit()
    conn.close()
    flash(f'"{username}" is now {new_role}.')
    return redirect(url_for('users_index'))


def _diff_dict(before, after, fields=None):
    """Return {field: [old, new]} for fields whose value changed.
       before/after are dict-like (e.g. sqlite3.Row dicts or plain dicts)."""
    if fields is None:
        fields = set(before.keys()) | set(after.keys())
    out = {}
    for k in fields:
        old = before.get(k) if before else None
        new = after.get(k)  if after  else None
        if (old or '') != (new or ''):
            out[k] = [old, new]
    return out


def log_audit(conn, action, entity, entity_id, summary='', changes=None, user=None):
    """Append a row to audit_log. `conn` is a live DB connection (commit handled by caller)."""
    user = user or current_user()
    conn.execute(
        '''INSERT INTO audit_log (occurred_at, user_name, action, entity, entity_id, summary, changes)
           VALUES (?,?,?,?,?,?,?)''',
        (datetime.now().isoformat(), user, action, entity, entity_id,
         (summary or '')[:500], json.dumps(changes) if changes else None)
    )


def get_audit_for(entity, entity_id, limit=50):
    conn = get_db()
    rows = conn.execute(
        '''SELECT * FROM audit_log
           WHERE entity = ? AND entity_id = ?
           ORDER BY occurred_at DESC LIMIT ?''',
        (entity, entity_id, limit)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    search = request.args.get('q', '').strip()
    conn = get_db()
    if search:
        rows = conn.execute(
            '''SELECT id, bill_no, bill_date, recipient_name, total_amount
               FROM bills WHERE bill_no LIKE ? OR recipient_name LIKE ?
               ORDER BY id DESC''',
            (f'%{search}%', f'%{search}%')
        ).fetchall()
    else:
        rows = conn.execute(
            'SELECT id, bill_no, bill_date, recipient_name, total_amount FROM bills ORDER BY id DESC'
        ).fetchall()
    conn.close()
    return render_template('index.html', bills=rows, search=search)


@app.route('/bill/new', methods=['GET', 'POST'])
def new_bill():
    if request.method == 'POST':
        f = request.form
        conn = get_db()

        count = max(1, min(20, _safe_int(f.get('delivery_count')) or 1))   # clamped 1..20
        bill_date_val = f.get('bill_date', '')
        delivery_month_val = build_delivery_month(f.get('delivery_month_select', ''), bill_date_val)
        vehicle_no_val = (f.get('vehicle_no') or '').strip().upper()
        recipient_name_val = (f.get('recipient_name') or '').strip()

        deliveries = []
        for i in range(count):
            deliveries.append({
                'sr_no':         i + 1,
                'gr_no':         f.get(f'd_gr_no_{i}', ''),
                'outward_no':    f.get(f'd_outward_no_{i}', ''),
                'outward_date':  f.get(f'd_outward_date_{i}', ''),
                'inward_no':     f.get(f'd_inward_no_{i}', ''),
                'inward_date':   f.get(f'd_inward_date_{i}', ''),
                'location':      f.get(f'd_location_{i}', ''),
                'consignee':     f.get(f'd_consignee_{i}', ''),
                'delivery_qty':  f.get(f'd_delivery_qty_{i}', ''),
                'converted_case':f.get(f'd_converted_case_{i}', ''),
                'inward_qty':    f.get(f'd_inward_qty_{i}', ''),
                'empty_qty':     f.get(f'd_empty_qty_{i}', ''),
                'weight':        f.get(f'd_weight_{i}', ''),
                'freight_rate':  f.get(f'd_freight_rate_{i}', ''),
                'overload':      f.get(f'd_overload_{i}', ''),
                'toll_tax':      f.get(f'd_toll_tax_{i}', ''),
                'excess_km':     f.get(f'd_excess_km_{i}', ''),
                'detention':     f.get(f'd_detention_{i}', ''),
                'unloading':     f.get(f'd_unloading_{i}', ''),
                'value_of_supply': f.get(f'd_value_of_supply_{i}', ''),
            })

        taxable_value = sum(_parse_amount(d['value_of_supply']) for d in deliveries)

        # GST computation
        hsn_sac         = (f.get('hsn_sac') or get_setting('default_hsn_sac') or '996511').strip()
        place_of_supply = (f.get('place_of_supply') or f.get('state_code') or '').strip()
        reverse_charge  = 1 if f.get('reverse_charge') else 0
        gst_pct         = max(0, min(28, _safe_num(f.get('gst_pct')) or 0)) if not reverse_charge else 0
        supplier_state  = get_setting('supplier_state_code') or '09'
        tax = compute_gst_split(taxable_value, gst_pct or 0,
                                supplier_state, place_of_supply, reverse_charge)
        total = tax['grand_total']

        # Allocate a collision-proof bill number and insert. If the number is
        # already taken (a racing save, or demo/imported data whose numbers ran
        # ahead of the counter), re-allocate from the current max and retry
        # rather than crashing with a UNIQUE-constraint 500.
        for _attempt in range(6):
            bill_no, n = _alloc_bill_no(conn)
            try:
                conn.execute(
                    '''INSERT INTO bills
                       (bill_no, bill_date, recipient_name, recipient_address, recipient_gstin,
                        state_code, trip_type, vehicle_no, freight_type, delivery_month,
                        client_name, total_amount, deliveries, created_at,
                        hsn_sac, taxable_value, reverse_charge, place_of_supply,
                        igst_pct, cgst_pct, sgst_pct,
                        igst_amount, cgst_amount, sgst_amount)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                    (bill_no,
                     bill_date_val, recipient_name_val, f.get('recipient_address'),
                     f.get('recipient_gstin'), f.get('state_code'), f.get('trip_type'),
                     vehicle_no_val, get_setting('freight_type'), delivery_month_val,
                     f.get('client_name', get_setting('client_name')),
                     total, json.dumps(deliveries), datetime.now().isoformat(),
                     hsn_sac, taxable_value, reverse_charge, place_of_supply,
                     tax['igst_pct'], tax['cgst_pct'], tax['sgst_pct'],
                     tax['igst_amount'], tax['cgst_amount'], tax['sgst_amount'])
                )
                break
            except sqlite3.IntegrityError:
                if _attempt == 5:
                    conn.rollback()
                    conn.close()
                    flash('Could not assign a unique bill number just now — please try saving again.')
                    return redirect(url_for('new_bill'))
                # else: loop re-allocates from the (now higher) max and retries
        bill_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.execute('UPDATE settings SET value=? WHERE key=?', (str(n + 1), 'next_bill_number'))
        remember_recipient(conn, recipient_name_val, f.get('recipient_address'),
                           f.get('recipient_gstin'), f.get('state_code'),
                           freight_rate=f.get('d_freight_rate_0'))
        remember_vehicle(conn, vehicle_no_val)

        # Phase D: link bill ↔ ledger if this bill came from a ledger entry
        from_le = f.get('from_ledger_id')
        if from_le and from_le.isdigit():
            le_id = int(from_le)
            conn.execute('UPDATE bills SET ledger_entry_id=? WHERE id=?', (le_id, bill_id))
            conn.execute('UPDATE ledger_entries SET bill_id=?, updated_at=? WHERE id=?',
                         (bill_id, datetime.now().isoformat(), le_id))
            log_audit(conn, 'update', 'ledger_entry', le_id,
                      summary=f'Linked to bill {bill_no}', changes={'bill_id': [None, bill_id]})

        log_audit(conn, 'create', 'bill', bill_id,
                  summary=f'Created bill {bill_no} for {recipient_name_val} · ₹{total:,.2f}')
        conn.commit()
        conn.close()
        return redirect(url_for('view_bill', bill_id=bill_id))

    bill_no, _ = next_bill_no()
    clients = json.loads(get_setting('clients') or '[]')

    # Optional prefill from an AI extraction OR from a ledger entry (Phase D)
    prefill_bill = None
    ext_id = request.args.get('from_extraction', type=int)
    le_id  = request.args.get('from_ledger',     type=int)
    if ext_id:
        bill_pre, deliveries = _build_prefill_from_extraction(ext_id)
        if bill_pre:
            bill_pre['deliveries']  = deliveries
            bill_pre['total_amount'] = 0
            prefill_bill = bill_pre
    elif le_id:
        bill_pre, deliveries = _build_prefill_from_ledger(le_id)
        if bill_pre:
            bill_pre['deliveries']        = deliveries
            bill_pre['total_amount']      = 0
            bill_pre['_from_ledger_id']   = le_id   # carried via hidden input below
            prefill_bill = bill_pre

    return render_template('new_bill.html',
        next_bill_no=bill_no,
        client_name=get_setting('client_name'),
        vehicle_type=get_setting('vehicle_type'),
        freight_type=get_setting('freight_type'),
        clients=clients,
        recipients=get_recipients(),
        vehicles=get_vehicles(),
        rate_list=get_rate_list(),
        months=MONTHS,
        today=datetime.now().strftime('%Y-%m-%d'),
        bill=prefill_bill,
        edit=False,
        prefilled=bool(prefill_bill),
        # GST defaults for new bills
        default_hsn_sac=get_setting('default_hsn_sac') or '996511',
        default_gst_pct=get_setting('default_gst_pct') or '5',
        default_reverse_charge=(get_setting('default_reverse_charge') or '1') == '1',
        gst_states=GST_STATE_NAMES,
    )


def _build_prefill_from_extraction(ext_id):
    """Read all extracted_invoices for this extraction and build a (bill_dict, deliveries_list)
       suitable for new_bill.html (which already supports an 'edit' / 'bill' object)."""
    conn = get_db()
    extraction = conn.execute('SELECT * FROM extractions WHERE id=?', (ext_id,)).fetchone()
    invs = conn.execute(
        'SELECT * FROM extracted_invoices WHERE extraction_id=? ORDER BY seq',
        (ext_id,)
    ).fetchall()
    conn.close()
    if not extraction or not invs:
        return None, []

    parsed_invs = []
    for inv in invs:
        try:
            parsed_invs.append(json.loads(inv['edited_json'] or inv['raw_json'] or '{}'))
        except Exception:
            pass
    if not parsed_invs:
        return None, []

    first = parsed_invs[0]
    mode = extraction['mode'] or 'combine'

    # Recipient (the entity who hires JL — i.e. the consignor on the VBL).
    # If consignor matches our default client, use the saved default address/GSTIN.
    # Otherwise use what the VBL provides (may be partial — user can fix on the form).
    default_consignor = (get_setting('default_consignor_name') or '').strip().upper()
    extracted_consignor = (first.get('consignor_name') or '').strip().upper()
    use_defaults = bool(default_consignor) and extracted_consignor == default_consignor

    if use_defaults:
        recipient_name    = get_setting('default_consignor_name')
        recipient_address = get_setting('default_consignor_address')
        recipient_gstin   = get_setting('default_consignor_gstin')
        state_code        = get_setting('default_consignor_state')
    else:
        recipient_name    = first.get('consignor_name') or default_consignor or ''
        recipient_address = first.get('consignor_address') or ''
        recipient_gstin   = first.get('consignor_gstin') or ''
        state_code        = (recipient_gstin[:2] if recipient_gstin else '')

    bill_pre = {
        'bill_date': first.get('date', ''),
        'recipient_name': recipient_name,
        'recipient_address': recipient_address,
        'recipient_gstin': recipient_gstin,
        'state_code': state_code,
        'trip_type': first.get('trip_type', 'One Way'),
        'vehicle_no': first.get('vehicle_reg_no', ''),
        'delivery_month': '',
        'delivery_month_select': '',
    }

    if mode == 'combine':
        # 1 row, totals summed, doc nos concatenated
        total_weight = sum(_num(p.get('total_weight_kg')) for p in parsed_invs)
        total_qty    = sum(_num(p.get('total_quantity'))   for p in parsed_invs)
        outward_nos  = ' / '.join(p.get('doc_no', '') for p in parsed_invs if p.get('doc_no'))
        outward_dt   = first.get('date', '')
        deliveries = [{
            'sr_no': 1,
            'gr_no': '',
            'outward_no': outward_nos,
            'outward_date': outward_dt,
            'inward_no': '',
            'inward_date': '',
            'location': '',
            'consignee': first.get('consignee_name', ''),
            'delivery_qty': total_qty,
            'converted_case': total_qty,
            'inward_qty': '',
            'empty_qty': '',
            'weight': total_weight,
            'freight_rate': '',
            'overload': '',
            'toll_tax': '',
            'excess_km': '',
            'detention': '',
            'unloading': '',
            'value_of_supply': '',
        }]
    else:
        # split: one row per invoice
        deliveries = []
        for i, p in enumerate(parsed_invs, start=1):
            deliveries.append({
                'sr_no': i,
                'gr_no': '',
                'outward_no': p.get('doc_no', ''),
                'outward_date': p.get('date', ''),
                'inward_no': '',
                'inward_date': '',
                'location': '',
                'consignee': p.get('consignee_name', ''),
                'delivery_qty': _num(p.get('total_quantity')),
                'converted_case': _num(p.get('total_quantity')),
                'inward_qty': '',
                'empty_qty': '',
                'weight': _num(p.get('total_weight_kg')),
                'freight_rate': '',
                'overload': '',
                'toll_tax': '',
                'excess_km': '',
                'detention': '',
                'unloading': '',
                'value_of_supply': '',
            })

    return bill_pre, deliveries


def _num(v):
    try:
        n = float(v)
        return int(n) if n == int(n) else n
    except (TypeError, ValueError):
        return ''


def _parse_amount(v):
    """Tolerantly coerce a hand-typed money value to a float — never raises.
       Accepts Indian-style commas, a ₹/Rs prefix, and stray spaces, so values
       like '1,200', '₹ 1,200.50' or '1200' all parse correctly. Blank or
       unparseable input becomes 0.0 (so a typo can't crash a bill save or
       silently vanish into a wrong total)."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    for junk in (',', '₹', 'Rs.', 'Rs', 'INR', ' '):
        s = s.replace(junk, '')
    try:
        return float(s)
    except (TypeError, ValueError):
        return 0.0


@app.route('/bill/<int:bill_id>')
def view_bill(bill_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM bills WHERE id=?', (bill_id,)).fetchone()
    conn.close()
    if not row:
        return 'Bill not found', 404
    bill = dict(row)
    bill['deliveries'] = json.loads(bill['deliveries'] or '[]')
    return render_template('bill_view.html', bill=bill,
                           vehicle_type=get_setting('vehicle_type'),
                           freight_type=get_setting('freight_type'),
                           audit_entries=get_audit_for('bill', bill_id))


@app.route('/bill/<int:bill_id>/einvoice.json')
def bill_einvoice_json(bill_id):
    """Export this bill as GSTN e-invoice schema 1.1 JSON.
       Customer uploads this to the IRP / e-invoice portal (or hands to their GSP)
       to generate the IRN + signed QR. Returns Content-Disposition: attachment."""
    from flask import Response
    conn = get_db()
    row = conn.execute('SELECT * FROM bills WHERE id=?', (bill_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Bill not found'}), 404

    bill = dict(row)
    deliveries = json.loads(bill['deliveries'] or '[]')
    sup = get_supplier_identity()
    taxable = float(bill.get('taxable_value') or bill.get('total_amount') or 0)
    igst = float(bill.get('igst_amount') or 0)
    cgst = float(bill.get('cgst_amount') or 0)
    sgst = float(bill.get('sgst_amount') or 0)
    total_tax = round(igst + cgst + sgst, 2)
    grand = float(bill.get('total_amount') or taxable + total_tax)
    rcm = bool(bill.get('reverse_charge'))
    pos = (bill.get('place_of_supply') or bill.get('state_code') or sup['state_code'] or '09')

    # Single-line-item invoice (transport service). Most GTAs file this way.
    # If a future bill needs per-delivery line items, expand here.
    item = {
        'SlNo':           '1',
        'PrdDesc':        sup['description_of_service'],            # "GOODS TRANSPORT AGENCY SERVICE"
        'IsServc':        'Y',
        'HsnCd':          bill.get('hsn_sac') or '996511',
        'Qty':            len(deliveries) or 1,
        'Unit':           'NOS',
        'UnitPrice':      round(taxable / max(1, len(deliveries) or 1), 2),
        'TotAmt':         round(taxable, 2),
        'Discount':       0,
        'AssAmt':         round(taxable, 2),
        'GstRt':          float(bill.get('igst_pct') or (bill.get('cgst_pct') or 0) * 2),
        'IgstAmt':        round(igst, 2),
        'CgstAmt':        round(cgst, 2),
        'SgstAmt':        round(sgst, 2),
        'CesRt':          0,
        'CesAmt':         0,
        'StateCesRt':     0,
        'StateCesAmt':    0,
        'OthChrg':        0,
        'TotItemVal':     round(taxable + igst + cgst + sgst, 2),
    }

    einv = {
        'Version': '1.1',
        'TranDtls': {
            'TaxSch':    'GST',
            'SupTyp':    'B2B',
            'RegRev':    'Y' if rcm else 'N',
            'IgstOnIntra': 'N',
        },
        'DocDtls': {
            'Typ': 'INV',
            'No':  bill['bill_no'],
            'Dt':  _format_dd_mm_yyyy(bill.get('bill_date')),    # DD/MM/YYYY
        },
        'SellerDtls': {
            'Gstin':  sup['gstin'] or '',
            'LglNm':  sup['name'] or '',
            'Addr1':  (sup['address'] or '')[:100],
            'Addr2':  (sup['address'] or '')[100:200] if len(sup['address'] or '') > 100 else None,
            'Loc':    GST_STATE_NAMES.get(sup['state_code'], '').title(),
            'Pin':    _extract_pincode(sup['address']),
            'Stcd':   sup['state_code'],
            'Ph':     (sup['phone'] or '').split(',')[0].strip()[:12] or None,
            'Em':     sup['email'] or None,
        },
        'BuyerDtls': {
            'Gstin':  bill.get('recipient_gstin') or '',
            'LglNm':  bill.get('recipient_name') or '',
            'Pos':    pos,
            'Addr1':  (bill.get('recipient_address') or '')[:100],
            'Addr2':  (bill.get('recipient_address') or '')[100:200] if len(bill.get('recipient_address') or '') > 100 else None,
            'Loc':    GST_STATE_NAMES.get(bill.get('state_code') or '', '').title(),
            'Pin':    _extract_pincode(bill.get('recipient_address')),
            'Stcd':   bill.get('state_code') or '',
        },
        'ItemList': [item],
        'ValDtls': {
            'AssVal':    round(taxable, 2),
            'CgstVal':   round(cgst, 2),
            'SgstVal':   round(sgst, 2),
            'IgstVal':   round(igst, 2),
            'CesVal':    0,
            'StCesVal':  0,
            'Discount':  0,
            'OthChrg':   0,
            'RndOffAmt': 0,
            'TotInvVal': round(grand, 2),
        },
    }

    # Drop None / empty optional fields (GSTN's schema validator is strict)
    def prune(d):
        if isinstance(d, dict):
            return {k: prune(v) for k, v in d.items() if v not in (None, '', [], {})}
        if isinstance(d, list):
            return [prune(x) for x in d]
        return d

    pruned = prune(einv)
    body = json.dumps(pruned, indent=2, ensure_ascii=False)
    filename = f'einvoice_{bill["bill_no"].replace("/", "-")}.json'
    return Response(body,
                    mimetype='application/json',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})


def _format_dd_mm_yyyy(iso_or_text):
    """Convert YYYY-MM-DD (ISO) → DD/MM/YYYY for e-invoice schema."""
    if not iso_or_text:
        return ''
    s = str(iso_or_text).strip()[:10]
    if len(s) == 10 and s[4] == '-' and s[7] == '-':
        return f'{s[8:10]}/{s[5:7]}/{s[0:4]}'
    return s


def _extract_pincode(addr):
    """Pull a 6-digit PIN from an address string. Returns None if not found."""
    if not addr:
        return None
    import re
    m = re.search(r'\b(\d{6})\b', addr)
    return m.group(1) if m else None


@app.route('/bill/<int:bill_id>/edit', methods=['GET', 'POST'])
def edit_bill(bill_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM bills WHERE id=?', (bill_id,)).fetchone()
    if not row:
        conn.close()
        return 'Bill not found', 404

    if request.method == 'POST':
        f = request.form
        count = max(1, min(20, _safe_int(f.get('delivery_count')) or 1))   # clamped 1..20
        bill_date_val = f.get('bill_date', '')
        delivery_month_val = build_delivery_month(f.get('delivery_month_select', ''), bill_date_val)
        vehicle_no_val = (f.get('vehicle_no') or '').strip().upper()
        recipient_name_val = (f.get('recipient_name') or '').strip()
        deliveries = []
        for i in range(count):
            deliveries.append({
                'sr_no':         i + 1,
                'gr_no':         f.get(f'd_gr_no_{i}', ''),
                'outward_no':    f.get(f'd_outward_no_{i}', ''),
                'outward_date':  f.get(f'd_outward_date_{i}', ''),
                'inward_no':     f.get(f'd_inward_no_{i}', ''),
                'inward_date':   f.get(f'd_inward_date_{i}', ''),
                'location':      f.get(f'd_location_{i}', ''),
                'consignee':     f.get(f'd_consignee_{i}', ''),
                'delivery_qty':  f.get(f'd_delivery_qty_{i}', ''),
                'converted_case':f.get(f'd_converted_case_{i}', ''),
                'inward_qty':    f.get(f'd_inward_qty_{i}', ''),
                'empty_qty':     f.get(f'd_empty_qty_{i}', ''),
                'weight':        f.get(f'd_weight_{i}', ''),
                'freight_rate':  f.get(f'd_freight_rate_{i}', ''),
                'overload':      f.get(f'd_overload_{i}', ''),
                'toll_tax':      f.get(f'd_toll_tax_{i}', ''),
                'excess_km':     f.get(f'd_excess_km_{i}', ''),
                'detention':     f.get(f'd_detention_{i}', ''),
                'unloading':     f.get(f'd_unloading_{i}', ''),
                'value_of_supply': f.get(f'd_value_of_supply_{i}', ''),
            })
        taxable_value = sum(_parse_amount(d['value_of_supply']) for d in deliveries)

        hsn_sac         = (f.get('hsn_sac') or get_setting('default_hsn_sac') or '996511').strip()
        place_of_supply = (f.get('place_of_supply') or f.get('state_code') or '').strip()
        reverse_charge  = 1 if f.get('reverse_charge') else 0
        gst_pct         = max(0, min(28, _safe_num(f.get('gst_pct')) or 0)) if not reverse_charge else 0
        supplier_state  = get_setting('supplier_state_code') or '09'
        tax = compute_gst_split(taxable_value, gst_pct or 0,
                                supplier_state, place_of_supply, reverse_charge)
        total = tax['grand_total']

        conn.execute(
            '''UPDATE bills SET bill_date=?, recipient_name=?, recipient_address=?,
               recipient_gstin=?, state_code=?, trip_type=?, vehicle_no=?,
               freight_type=?, delivery_month=?, client_name=?,
               total_amount=?, deliveries=?,
               hsn_sac=?, taxable_value=?, reverse_charge=?, place_of_supply=?,
               igst_pct=?, cgst_pct=?, sgst_pct=?,
               igst_amount=?, cgst_amount=?, sgst_amount=?
               WHERE id=?''',
            (bill_date_val, recipient_name_val, f.get('recipient_address'),
             f.get('recipient_gstin'), f.get('state_code'), f.get('trip_type'),
             vehicle_no_val, get_setting('freight_type'), delivery_month_val,
             f.get('client_name'), total, json.dumps(deliveries),
             hsn_sac, taxable_value, reverse_charge, place_of_supply,
             tax['igst_pct'], tax['cgst_pct'], tax['sgst_pct'],
             tax['igst_amount'], tax['cgst_amount'], tax['sgst_amount'],
             bill_id)
        )
        remember_recipient(conn, recipient_name_val, f.get('recipient_address'),
                           f.get('recipient_gstin'), f.get('state_code'),
                           freight_rate=f.get('d_freight_rate_0'))
        remember_vehicle(conn, vehicle_no_val)

        # Audit: capture changes to scalar fields
        before = dict(row)
        after = {
            'bill_date': bill_date_val, 'recipient_name': recipient_name_val,
            'recipient_address': f.get('recipient_address'),
            'recipient_gstin': f.get('recipient_gstin'),
            'state_code': f.get('state_code'), 'trip_type': f.get('trip_type'),
            'vehicle_no': vehicle_no_val, 'delivery_month': delivery_month_val,
            'client_name': f.get('client_name'), 'total_amount': total,
        }
        changes = _diff_dict(before, after, fields=after.keys())
        if changes:
            log_audit(conn, 'update', 'bill', bill_id,
                      summary=f'Edited bill {row["bill_no"]} ({len(changes)} field{"s" if len(changes)!=1 else ""} changed)',
                      changes=changes)
        conn.commit()
        conn.close()
        return redirect(url_for('view_bill', bill_id=bill_id))

    bill = dict(row)
    bill['deliveries'] = json.loads(bill['deliveries'] or '[]')
    bill['delivery_month_select'] = split_delivery_month(bill.get('delivery_month'))
    conn.close()
    clients = json.loads(get_setting('clients') or '[]')
    return render_template('new_bill.html',
        bill=bill,
        next_bill_no=bill['bill_no'],
        client_name=bill['client_name'],
        vehicle_type=get_setting('vehicle_type'),
        freight_type=get_setting('freight_type'),
        clients=clients,
        recipients=get_recipients(),
        vehicles=get_vehicles(),
        rate_list=get_rate_list(),
        months=MONTHS,
        today=bill['bill_date'],
        edit=True,
        default_hsn_sac=get_setting('default_hsn_sac') or '996511',
        default_gst_pct=get_setting('default_gst_pct') or '5',
        default_reverse_charge=(get_setting('default_reverse_charge') or '1') == '1',
        gst_states=GST_STATE_NAMES,
    )


@app.route('/bill/<int:bill_id>/delete', methods=['POST'])
def delete_bill(bill_id):
    conn = get_db()
    row = conn.execute('SELECT bill_no, recipient_name FROM bills WHERE id=?', (bill_id,)).fetchone()
    if not row:
        conn.close()
        flash('Bill not found.')
        return redirect(url_for('index'))
    summary_text = f'Deleted bill {row["bill_no"]} ({row["recipient_name"]})'
    # Soft-delete: move the row to bills_archive (Recycle Bin) instead of
    # destroying it. We deliberately do NOT null ledger_entries.bill_id — on
    # restore the bill returns with its original id and those links re-validate
    # automatically. (Purging from the bin is what frees the trips to re-bill.)
    conn.execute('INSERT INTO bills_archive SELECT * FROM bills WHERE id=?', (bill_id,))
    conn.execute('DELETE FROM bills WHERE id=?', (bill_id,))
    log_audit(conn, 'delete', 'bill', bill_id, summary=summary_text + ' (→ Recycle Bin)')
    conn.commit()
    conn.close()
    flash(t('Bill moved to the Recycle Bin. Restore it from there if this was a mistake.'))
    return redirect(url_for('index'))


@app.route('/summary', methods=['GET', 'POST'])
def summary():
    if request.method == 'POST':
        ids = request.form.getlist('bill_ids')
        if not ids:
            flash('Please select at least one bill.')
            return redirect(url_for('summary'))
        conn = get_db()
        placeholders = ','.join('?' * len(ids))
        rows = conn.execute(
            f'SELECT * FROM bills WHERE id IN ({placeholders}) ORDER BY bill_no', ids
        ).fetchall()
        conn.close()
        bills = []
        for r in rows:
            b = dict(r)
            b['deliveries'] = json.loads(b['deliveries'] or '[]')
            bills.append(b)
        grand_total = sum(_parse_amount(b['total_amount']) for b in bills)
        return render_template('summary_view.html', bills=bills, grand_total=grand_total)

    conn = get_db()
    rows = conn.execute(
        'SELECT id, bill_no, bill_date, recipient_name, total_amount FROM bills ORDER BY id DESC'
    ).fetchall()
    conn.close()
    return render_template('summary.html', bills=rows)


@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if request.method == 'POST':
        f = request.form
        set_setting('next_bill_number', f.get('next_bill_number', '1'))
        set_setting('next_lr_number',   f.get('next_lr_number', '1'))
        set_setting('default_consignor_name',    f.get('default_consignor_name', '').strip())
        set_setting('default_consignor_address', f.get('default_consignor_address', '').strip())
        set_setting('default_consignor_gstin',   f.get('default_consignor_gstin', '').strip())
        set_setting('default_consignor_state',   f.get('default_consignor_state', '').strip())
        set_setting('client_name',      f.get('default_client', 'VARUN BEVERAGES LIMITED'))
        set_setting('vehicle_type',     f.get('vehicle_type', 'A1/A4'))
        set_setting('freight_type',     f.get('freight_type', 'LP'))
        raw = f.get('clients', '')
        clients = [c.strip() for c in raw.splitlines() if c.strip()]
        set_setting('clients', json.dumps(clients))
        # ── Supplier identity (JL — the transporter) ──
        set_setting('supplier_name',         f.get('supplier_name', '').strip())
        set_setting('supplier_address',      f.get('supplier_address', '').strip())
        # GSTIN: normalise + validate. Empty allowed (not all customers have GSTIN yet).
        sup_gstin_raw = (f.get('supplier_gstin') or '').strip().upper()
        if sup_gstin_raw:
            ok, norm, err = validate_gstin(sup_gstin_raw)
            if not ok:
                flash(f'Supplier GSTIN invalid: {err}')
                return redirect(url_for('settings'))
            set_setting('supplier_gstin', norm)
        else:
            set_setting('supplier_gstin', '')
        set_setting('supplier_state_code',   f.get('supplier_state_code', '09').strip())
        set_setting('supplier_pan',          f.get('supplier_pan', '').strip().upper())
        set_setting('supplier_phone',        f.get('supplier_phone', '').strip())
        set_setting('supplier_email',        f.get('supplier_email', '').strip())
        set_setting('supplier_bank_name',    f.get('supplier_bank_name', '').strip())
        set_setting('supplier_bank_account', f.get('supplier_bank_account', '').strip())
        set_setting('supplier_bank_ifsc',    f.get('supplier_bank_ifsc', '').strip().upper())
        # ── GST defaults ──
        set_setting('default_hsn_sac',        f.get('default_hsn_sac', '996511').strip())
        set_setting('default_reverse_charge', '1' if f.get('default_reverse_charge') else '0')
        set_setting('default_gst_pct',        f.get('default_gst_pct', '5').strip())
        flash(t('Settings saved.'))
        return redirect(url_for('settings'))

    clients_raw = '\n'.join(json.loads(get_setting('clients') or '[]'))
    rate_list = get_rate_list()
    return render_template('settings.html',
        next_bill_number=get_setting('next_bill_number'),
        next_lr_number=get_setting('next_lr_number'),
        default_client=get_setting('client_name'),
        default_consignor_name=get_setting('default_consignor_name'),
        default_consignor_address=get_setting('default_consignor_address'),
        default_consignor_gstin=get_setting('default_consignor_gstin'),
        default_consignor_state=get_setting('default_consignor_state'),
        vehicle_type=get_setting('vehicle_type'),
        freight_type=get_setting('freight_type'),
        clients_raw=clients_raw,
        rate_list=rate_list,
        # GST / supplier identity
        supplier=get_supplier_identity(),
        default_hsn_sac=get_setting('default_hsn_sac') or '996511',
        default_reverse_charge=(get_setting('default_reverse_charge') or '1') == '1',
        default_gst_pct=get_setting('default_gst_pct') or '5',
        gst_states=GST_STATE_NAMES,
    )


@app.route('/settings/rate-list/upload', methods=['POST'])
def upload_rate_list():
    f = request.files.get('rate_file')
    if not f or not f.filename:
        flash('Please choose an Excel file.')
        return redirect(url_for('settings'))
    if not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        flash('File must be an Excel workbook (.xlsx).')
        return redirect(url_for('settings'))
    tmp_path = os.path.join(UPLOAD_DIR, f'_rate_upload_{uuid.uuid4().hex[:8]}.xlsx')
    f.save(tmp_path)
    try:
        count, msg = import_rate_list_from_xlsx(tmp_path)
        if count == 0 and msg:
            flash(f'Rate list import failed: {msg}')
        else:
            extra = f' {msg}' if msg else ''
            flash(f'Imported {count} rate rows.{extra}')
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
    return redirect(url_for('settings'))


@app.route('/settings/rate-list/clear', methods=['POST'])
def clear_rate_list():
    conn = get_db()
    conn.execute('DELETE FROM freight_rates')
    conn.commit()
    conn.close()
    flash('Rate list cleared.')
    return redirect(url_for('settings'))


# ── Rate-list editor (full CRUD) ─────────────────────────────────────────────

@app.route('/rate-list')
def rate_list_editor():
    f = request.args
    where, params = ['1=1'], []
    if f.get('q'):
        where.append('(customer_name LIKE ? OR location LIKE ? OR party_code LIKE ?)')
        q = f'%{f["q"]}%'; params += [q, q, q]
    rows = get_db().execute(
        f'SELECT * FROM freight_rates WHERE {" AND ".join(where)} ORDER BY customer_name LIMIT 1000',
        params
    ).fetchall()
    return render_template('rate_list_editor.html',
                           rates=[dict(r) for r in rows],
                           filters=dict(f))


@app.route('/rate-list/add', methods=['POST'])
def rate_list_add():
    f = request.form
    name = (f.get('customer_name') or '').strip().upper()
    if not name:
        flash('Customer name required.')
        return redirect(url_for('rate_list_editor'))
    location = (f.get('location') or '').strip().upper()
    conn = get_db()
    try:
        conn.execute(
            '''INSERT INTO freight_rates
                 (customer_name, party_code, location, dist_twy_km, dist_owy_km,
                  lp_owy, lp_twy, trolla_owy, trolla_twy, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?)''',
            (name, (f.get('party_code') or '').strip(), location,
             _safe_int(f.get('dist_twy_km')), _safe_int(f.get('dist_owy_km')),
             _safe_num(f.get('lp_owy')),    _safe_num(f.get('lp_twy')),
             _safe_num(f.get('trolla_owy')),_safe_num(f.get('trolla_twy')),
             datetime.now().isoformat())
        )
        new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        log_audit(conn, 'create', 'rate_row', new_id, summary=f'Added rate {name} ({location})')
        conn.commit()
        flash(f'Added rate row for {name} ({location}).')
    except sqlite3.IntegrityError:
        flash(f'A rate row already exists for {name} at {location}. Use Edit instead.')
    conn.close()
    return redirect(url_for('rate_list_editor'))


@app.route('/rate-list/<int:rid>/update', methods=['POST'])
def rate_list_update(rid):
    f = request.form
    conn = get_db()
    before_row = conn.execute('SELECT * FROM freight_rates WHERE id=?', (rid,)).fetchone()
    after_vals = {
        'customer_name': (f.get('customer_name') or '').strip().upper(),
        'party_code':    (f.get('party_code') or '').strip(),
        'location':      (f.get('location') or '').strip().upper(),
        'dist_twy_km':   _safe_int(f.get('dist_twy_km')),
        'dist_owy_km':   _safe_int(f.get('dist_owy_km')),
        'lp_owy':        _safe_num(f.get('lp_owy')),
        'lp_twy':        _safe_num(f.get('lp_twy')),
        'trolla_owy':    _safe_num(f.get('trolla_owy')),
        'trolla_twy':    _safe_num(f.get('trolla_twy')),
    }
    conn.execute('''
        UPDATE freight_rates SET
          customer_name=?, party_code=?, location=?,
          dist_twy_km=?, dist_owy_km=?,
          lp_owy=?, lp_twy=?, trolla_owy=?, trolla_twy=?,
          updated_at=?
        WHERE id=?''',
        (after_vals['customer_name'], after_vals['party_code'], after_vals['location'],
         after_vals['dist_twy_km'], after_vals['dist_owy_km'],
         after_vals['lp_owy'], after_vals['lp_twy'],
         after_vals['trolla_owy'], after_vals['trolla_twy'],
         datetime.now().isoformat(), rid))
    if before_row:
        changes = _diff_dict(dict(before_row), after_vals, fields=after_vals.keys())
        if changes:
            log_audit(conn, 'update', 'rate_row', rid,
                      summary=f'Edited rate {after_vals["customer_name"]} ({after_vals["location"]})',
                      changes=changes)
    conn.commit()
    conn.close()
    flash('Rate row updated.')
    return redirect(url_for('rate_list_editor', q=request.args.get('q', '')))


@app.route('/rate-list/<int:rid>/delete', methods=['POST'])
def rate_list_delete(rid):
    conn = get_db()
    row = conn.execute('SELECT customer_name, location FROM freight_rates WHERE id=?', (rid,)).fetchone()
    summary = f'Deleted rate {row["customer_name"]} ({row["location"]})' if row else f'Deleted rate #{rid}'
    conn.execute('DELETE FROM freight_rates WHERE id=?', (rid,))
    log_audit(conn, 'delete', 'rate_row', rid, summary=summary)
    conn.commit()
    conn.close()
    flash('Rate row deleted.')
    return redirect(url_for('rate_list_editor', q=request.args.get('q', '')))


def _safe_int(v):
    try:
        return int(float(v)) if v not in (None, '') else None
    except (ValueError, TypeError):
        return None


# ─────────────────────────────────────────────────────────────────────────────
# VBL EXTRACTION — Google Gemini 2.5 Flash
# ─────────────────────────────────────────────────────────────────────────────
# Sends each VBL Proof-of-Delivery image/PDF to Gemini 2.5 Flash with a strict
# JSON schema. Free tier (1500 req/day, 10 RPM) covers the user's volume.
# Internet required.

VBL_EXTRACTION_PROMPT = """You are extracting structured data from a Proof-of-Delivery invoice
issued by a consignor company and dispatched via a transporter.
Read the document carefully — it may be photographed at an angle, partially obscured by a stamp,
rotated 90/180/270 degrees, or have ink smudges. Extract values exactly as they appear on the document.

Return ONLY a JSON object matching this exact schema (no prose, no markdown fences):

{
  "doc_no":           "string — top-right Doc No., digits only, e.g. 893358371",
  "date":             "YYYY-MM-DD — parse the Date: field; convert DD.MM.YYYY → YYYY-MM-DD",
  "consignor_name":    "string — the consignor/header company name exactly as printed",
  "consignor_address": "string — full address line directly under the consignor company name (e.g. 'GATA NO-84 & 85, VILL-GADAN KHEDA, CHAKARPUR RANIYAN, Pin:-209304 Dist:- KANPUR DEHAT')",
  "consignor_gstin":   "string — GSTIN under consignor header",
  "consignee_name":   "string — Party's Name (do NOT include the trailing /code number)",
  "consignee_address":"string — full Address line, single line, commas preserved",
  "consignee_gstin":  "string — GSTIN: under party block (15 alphanumeric chars)",
  "consignee_mobile": "string — Cust. Mobile No., 10-12 digits",
  "consignee_fssai":  "string — FSSAI/Valid Date number if present",
  "vehicle_reg_no":   "string — Vehicle Reg. No., uppercase, NO spaces (e.g. UP77BN0421). Strip the truck capacity suffix.",
  "truck_capacity":   "string — e.g. 15T, 9T (the value after slash on Vehicle Reg line)",
  "trip_type":        "One Way | Two Way | Round Trip — value of the Inco Term field",
  "transporter":      "string — the Transporter field value, the transporter's own pre-printed name as it appears",
  "form":             "string — Form: value if present, else empty",
  "total_weight_kg":  number,
  "total_quantity":   number,
  "quantity_unit":    "CS | BS | PCS — unit shown next to totals (usually CS)",
  "line_items": [
    {
      "document_number": "string — 3-part code like 893355481/880207629/RV0096401853",
      "product_code":    "string — short numeric code at start of product name, e.g. 5550197",
      "product_name":    "string — full product description on that row",
      "batch_or_date":   "string — small line below product name like 7956F01C26 28.08.2026",
      "mrp":             number,
      "quantity":        number,
      "unit":            "CS | BS | PCS"
    }
  ],
  "confidence_per_field": {},
  "notes": "string — anything unusual (smudges, partial cuts, stamp overlap)"
}

Rules:
- If a field is unreadable, use empty string "" / 0 / [].
- Numbers: never include thousand separators or units inside the number value.
- Dates: convert to YYYY-MM-DD strictly.
- Vehicle reg numbers: uppercase, no spaces, no truck capacity suffix.
- For confidence_per_field: include ONLY fields where you are uncertain, with values "low" or "medium". High-confidence fields are omitted.
- For line_items: include EVERY visible row in order, even if some columns are partially hidden.
"""


# Lazy-loaded Gemini client (initialized on first call)
_gemini_client = None

def _get_gemini_client():
    global _gemini_client

    if _gemini_client is not None:
        return _gemini_client, None

    api_key = os.environ.get("GOOGLE_API_KEY", "").strip()

    if not api_key or api_key.startswith("PASTE_"):
        return None, (
            "🔑 Google API key not set. Open .env in this folder and paste your "
            "key from https://aistudio.google.com/apikey, then restart the app."
        )

    try:
        from google import genai
        from google.genai import types
        # Without an explicit timeout, a hung/stalled network connection blocks
        # the extraction thread FOREVER — the retry logic in
        # _gemini_call_with_retry only ever triggers on an exception, and a
        # plain hang never raises one. This bounds every Gemini call to 90s so
        # a stuck request always surfaces as a normal (retryable/friendly)
        # error instead of leaving the extraction stuck in 'pending' forever.
        _gemini_client = genai.Client(
            api_key=api_key,
            http_options=types.HttpOptions(timeout=90_000),  # milliseconds
        )
        return _gemini_client, None
    except ImportError:
        return None, ("📦 Gemini SDK not installed. Restart the app — run_mac.command will "
                      "auto-install it on next launch.")
    except Exception as e:
        return None, f"Could not initialize Gemini client: {e}"


def ai_available():
    """Quick check used to disable the upload buttons if AI isn't ready."""
    return _get_gemini_client()[0] is not None


@app.context_processor
def inject_ai_status():
    """Make {{ ai_available }} available to all templates."""
    return {'ai_available': ai_available()}


# ── Notifications feed for the bell icon ──────────────────────────────────────
@app.context_processor
def inject_notifications():
    """Compute pending alerts on every page render so the bell badge stays live."""
    try:
        conn = get_db()
        items = []
        # 1) Overdue PODs
        overdue_days = int(get_setting('pod_overdue_days') or 10)
        from datetime import timedelta
        cutoff = (datetime.now().date() - timedelta(days=overdue_days)).isoformat()
        n_overdue = conn.execute(
            'SELECT COUNT(*) FROM ledger_entries WHERE pod_received=0 AND entry_date < ?',
            (cutoff,)
        ).fetchone()[0]
        if n_overdue > 0:
            items.append({
                'icon': 'bi-exclamation-triangle-fill',
                'color': 'danger',
                'title': f'{n_overdue} overdue POD' + ('s' if n_overdue != 1 else ''),
                'meta': f'no POD received after {overdue_days} days',
                'link': url_for('ledger_index', status='pod_pending'),
            })

        # 2) Ready to bill
        n_ready = conn.execute(
            'SELECT COUNT(*) FROM ledger_entries WHERE pod_received=1 AND (bill_id IS NULL OR bill_id=0)'
        ).fetchone()[0]
        if n_ready > 0:
            items.append({
                'icon': 'bi-receipt-cutoff',
                'color': 'success',
                'title': f'{n_ready} entr{"ies" if n_ready != 1 else "y"} ready to bill',
                'meta': 'POD received — create bills',
                'link': url_for('ready_to_bill'),
            })

        # 3) Unpaid bills (positive client balance)
        unpaid = conn.execute(
            '''SELECT COUNT(DISTINCT recipient_name) FROM bills
               WHERE recipient_name IS NOT NULL'''
        ).fetchone()[0]
        # We only show this card if there's actual balance pending (computed below)
        clients_with_balance = list_clients_with_balance() if unpaid else []
        n_unpaid = sum(1 for c in clients_with_balance if c['balance'] > 0.005)
        if n_unpaid > 0:
            items.append({
                'icon': 'bi-cash-coin',
                'color': 'warning',
                'title': f'{n_unpaid} client{"s" if n_unpaid != 1 else ""} with pending balance',
                'meta': 'collect outstanding payments',
                'link': url_for('payments_hub'),
            })

        # 4) Draft challans (saved but not finalized)
        n_drafts = conn.execute(
            "SELECT COUNT(*) FROM challans WHERE status='draft'"
        ).fetchone()[0]
        if n_drafts > 0:
            items.append({
                'icon': 'bi-clipboard2-pulse',
                'color': 'info',
                'title': f'{n_drafts} draft challan{"s" if n_drafts != 1 else ""}',
                'meta': 'review and save',
                'link': url_for('challans_index'),
            })

        conn.close()
        return {'notifications': {'entries': items, 'total': len(items)}}
    except Exception:
        return {'notifications': {'entries': [], 'total': 0}}


# ── Global search API ─────────────────────────────────────────────────────────
@app.route('/api/vehicle-history/<path:vehicle_no>')
def api_vehicle_history(vehicle_no):
    """Return defaults for a vehicle from its most recent ledger entry +
       most-frequent station/transporter for that vehicle. JSON, used by
       the ledger form to suggest auto-fill values."""
    v = (vehicle_no or '').strip().upper().replace(' ', '')
    if len(v) < 4:
        return jsonify({'found': False})
    conn = get_db()
    last = conn.execute(
        '''SELECT entry_date, station, freight, advance_cash, advance_account,
                  diesel, mt_qty, trip_type, transporter_id, gr_no
           FROM ledger_entries
           WHERE UPPER(REPLACE(vehicle_no, ' ', '')) = ?
           ORDER BY entry_date DESC, id DESC LIMIT 1''', (v,)).fetchone()
    # Most frequent station for this vehicle
    freq_station = conn.execute(
        '''SELECT station, COUNT(*) c FROM ledger_entries
           WHERE UPPER(REPLACE(vehicle_no, ' ', '')) = ? AND station != ''
           GROUP BY station ORDER BY c DESC LIMIT 1''', (v,)).fetchone()
    # Linked challan info if last trip had one
    challan = None
    if last:
        cr = conn.execute(
            '''SELECT consignee_name, invoice_no, weight_kg
               FROM challans WHERE id IN
                 (SELECT challan_id FROM ledger_entries
                  WHERE UPPER(REPLACE(vehicle_no, ' ', '')) = ? AND challan_id IS NOT NULL
                  ORDER BY entry_date DESC LIMIT 1)''', (v,)).fetchone()
        if cr:
            challan = dict(cr)
    conn.close()
    if not last:
        return jsonify({'found': False})
    out = dict(last)
    out['found'] = True
    if freq_station:
        out['frequent_station'] = freq_station['station']
    if challan:
        out['last_consignee'] = challan.get('consignee_name')
        out['last_invoice'] = challan.get('invoice_no')
    return jsonify(out)


@app.route('/api/search')
def api_search():
    """JSON search across bills, challans, ledger entries, and clients."""
    q = (request.args.get('q') or '').strip()
    if len(q) < 2:
        return jsonify({'bills': [], 'challans': [], 'ledger': [], 'clients': []})
    like = f'%{q}%'
    conn = get_db()
    try:
        bills = [dict(r) for r in conn.execute(
            '''SELECT id, bill_no, recipient_name, total_amount, bill_date FROM bills
               WHERE bill_no LIKE ? OR recipient_name LIKE ? OR vehicle_no LIKE ?
               ORDER BY id DESC LIMIT 6''',
            (like, like, like)).fetchall()]
        challans = [dict(r) for r in conn.execute(
            '''SELECT id, lr_no, consignee_name, truck_no FROM challans
               WHERE lr_no LIKE ? OR consignee_name LIKE ? OR truck_no LIKE ? OR invoice_no LIKE ?
               ORDER BY id DESC LIMIT 6''',
            (like, like, like, like)).fetchall()]
        ledger = [dict(r) for r in conn.execute(
            '''SELECT id, gr_no, vehicle_no, station, entry_date FROM ledger_entries
               WHERE gr_no LIKE ? OR vehicle_no LIKE ? OR station LIKE ? OR shipment_no LIKE ?
               ORDER BY id DESC LIMIT 6''',
            (like, like, like, like)).fetchall()]
        clients = [r[0] for r in conn.execute(
            '''SELECT DISTINCT recipient_name FROM bills
               WHERE recipient_name LIKE ? LIMIT 6''',
            (like,)).fetchall() if r[0]]
    finally:
        conn.close()

    return jsonify({
        'bills':    [{'title': f'Bill {b["bill_no"]}',
                      'subtitle': f'{b["recipient_name"] or ""} · ₹{b["total_amount"]:,.0f}' if b["total_amount"] else (b["recipient_name"] or ""),
                      'link': url_for('view_bill', bill_id=b['id'])} for b in bills],
        'challans': [{'title': f'LR-{c["lr_no"]}',
                      'subtitle': f'{c["consignee_name"] or ""} · {c["truck_no"] or ""}',
                      'link': url_for('challan_review', challan_id=c['id'])} for c in challans],
        'ledger':   [{'title': f'GR-{le["gr_no"] or le["id"]}',
                      'subtitle': f'{le["vehicle_no"] or ""} · {le["station"] or ""}',
                      'link': url_for('ledger_view', le_id=le['id'])} for le in ledger],
        'clients':  [{'title': name, 'subtitle': 'open ledger card',
                      'link': url_for('party_detail', party_type='client', party_key=name)}
                     for name in clients],
    })


# ─────────────────────────────────────────────────────────────────────────────
# Trip 360 — find a trip by any natural key (LR, Del No, Invoice, Vehicle…)
# and show the entire lifecycle: challan → ledger → bill → payments.
# ─────────────────────────────────────────────────────────────────────────────

def _load_trip_summary(conn, challan_id, le_id):
    s = {'challan_id': challan_id, 'le_id': le_id,
         'lr_no': '', 'date': '', 'vehicle': '', 'consignee': ''}
    if challan_id:
        r = conn.execute(
            'SELECT lr_no, challan_date, truck_no, consignee_name FROM challans WHERE id=?',
            (challan_id,)).fetchone()
        if r:
            s.update({'lr_no': r['lr_no'] or '', 'date': r['challan_date'] or '',
                      'vehicle': r['truck_no'] or '', 'consignee': r['consignee_name'] or ''})
    elif le_id:
        r = conn.execute(
            'SELECT gr_no, entry_date, vehicle_no, station FROM ledger_entries WHERE id=?',
            (le_id,)).fetchone()
        if r:
            s.update({'lr_no': r['gr_no'] or '', 'date': r['entry_date'] or '',
                      'vehicle': r['vehicle_no'] or '', 'consignee': r['station'] or ''})
    return s


def _load_trip_360(conn, challan_id, le_id):
    challan = None
    ledger = None
    bill = None

    if challan_id:
        r = conn.execute('SELECT * FROM challans WHERE id=?', (challan_id,)).fetchone()
        if r:
            challan = dict(r)
            if not le_id and challan.get('ledger_entry_id'):
                le_id = challan['ledger_entry_id']

    if le_id:
        r = conn.execute('SELECT * FROM ledger_entries WHERE id=?', (le_id,)).fetchone()
        if r:
            ledger = dict(r)
            ledger['balance'] = _ledger_balance(
                ledger['freight'], ledger['advance_cash'],
                ledger['advance_account'], ledger['diesel'])
            if ledger.get('bill_id'):
                br = conn.execute('SELECT * FROM bills WHERE id=?', (ledger['bill_id'],)).fetchone()
                if br:
                    bill = dict(br)
            if not challan and ledger.get('challan_id'):
                cr = conn.execute('SELECT * FROM challans WHERE id=?',
                                  (ledger['challan_id'],)).fetchone()
                if cr:
                    challan = dict(cr)

    payments_to_transporter = []
    if ledger and ledger.get('transporter_id'):
        payments_to_transporter = [dict(r) for r in conn.execute(
            '''SELECT * FROM payments WHERE party_type='transporter' AND party_key=?
               ORDER BY payment_date DESC LIMIT 5''',
            (str(ledger['transporter_id']),)).fetchall()]

    payments_from_client = []
    if bill and bill.get('recipient_name'):
        payments_from_client = [dict(r) for r in conn.execute(
            '''SELECT * FROM payments WHERE party_type='client' AND party_key=?
               ORDER BY payment_date DESC LIMIT 5''',
            (bill['recipient_name'],)).fetchall()]

    warnings = []
    if challan and ledger:
        if challan.get('lr_no') and ledger.get('gr_no') and \
           str(challan['lr_no']).strip() != str(ledger['gr_no']).strip():
            warnings.append(
                f'LR No. on challan ({challan["lr_no"]}) ≠ GR No. on ledger ({ledger["gr_no"]})')
        if challan.get('truck_no') and ledger.get('vehicle_no') and \
           challan['truck_no'].upper().replace(' ', '') != ledger['vehicle_no'].upper().replace(' ', ''):
            warnings.append(
                f'Vehicle mismatch: challan {challan["truck_no"]} vs ledger {ledger["vehicle_no"]}')
        cw = challan.get('weight_kg')
        lw = ledger.get('weight_kg') or (
            (ledger.get('mt_qty') or 0) * 1000 if ledger.get('mt_qty') else None)
        if cw and lw and abs(cw - lw) > 100:
            warnings.append(f'Weight mismatch: challan {cw:.0f} kg vs ledger {lw:.0f} kg')

    if challan and challan.get('pod_doc_no') and challan.get('del_no') and \
       str(challan['pod_doc_no']).strip() != str(challan['del_no']).strip():
        warnings.append(
            f'Del No. ({challan["del_no"]}) ≠ PoD Doc No. ({challan["pod_doc_no"]}) — typo?')

    return {
        'challan': challan, 'ledger': ledger, 'bill': bill,
        'payments_to_transporter': payments_to_transporter,
        'payments_from_client': payments_from_client,
        'warnings': warnings,
    }


@app.route('/trip')
def trip_lookup():
    q = (request.args.get('q') or '').strip()
    conn = get_db()
    matches = []
    seen = set()  # (challan_id, le_id) tuples we've already added

    def add(challan_id, le_id, reason, conf):
        key = (challan_id, le_id)
        if key in seen:
            return
        seen.add(key)
        matches.append({'challan_id': challan_id, 'le_id': le_id,
                        'reason': reason, 'confidence': conf})

    if q:
        # 1. Strongest: lr_no exact
        for r in conn.execute(
                'SELECT id, ledger_entry_id FROM challans WHERE lr_no=?',
                (q,)).fetchall():
            add(r['id'], r['ledger_entry_id'], f'LR No. = "{q}"', 100)

        # 2. del_no or pod_doc_no exact
        for r in conn.execute(
                '''SELECT id, ledger_entry_id FROM challans
                   WHERE del_no=? OR pod_doc_no=?''', (q, q)).fetchall():
            add(r['id'], r['ledger_entry_id'], f'Del No. / PoD Doc No. = "{q}"', 100)

        # 3. invoice_no exact
        for r in conn.execute(
                'SELECT id, ledger_entry_id FROM challans WHERE invoice_no=?',
                (q,)).fetchall():
            add(r['id'], r['ledger_entry_id'], f'Invoice No. = "{q}"', 95)

        # 4. ledger gr_no exact (catches manually-entered ledger rows w/o challan)
        for r in conn.execute(
                'SELECT id, challan_id FROM ledger_entries WHERE gr_no=?',
                (q,)).fetchall():
            add(r['challan_id'], r['id'], f'Ledger GR No. = "{q}"', 95)

        # 5. ledger shipment_no (= invoice no in many flows)
        for r in conn.execute(
                'SELECT id, challan_id FROM ledger_entries WHERE shipment_no=?',
                (q,)).fetchall():
            add(r['challan_id'], r['id'], f'Ledger Shipment / Invoice = "{q}"', 90)

        # 6. fall back to fuzzy LIKE if nothing exact
        if not matches:
            like = f'%{q}%'
            for r in conn.execute(
                    '''SELECT id, ledger_entry_id FROM challans
                       WHERE lr_no LIKE ? OR del_no LIKE ? OR pod_doc_no LIKE ?
                          OR invoice_no LIKE ? OR truck_no LIKE ? OR consignee_name LIKE ?
                       ORDER BY id DESC LIMIT 8''',
                    (like, like, like, like, like, like)).fetchall():
                add(r['id'], r['ledger_entry_id'], 'Fuzzy match in challan', 60)
            for r in conn.execute(
                    '''SELECT id, challan_id FROM ledger_entries
                       WHERE gr_no LIKE ? OR vehicle_no LIKE ? OR shipment_no LIKE ?
                          OR station LIKE ?
                       ORDER BY id DESC LIMIT 8''',
                    (like, like, like, like)).fetchall():
                add(r['challan_id'], r['id'], 'Fuzzy match in ledger', 55)

    trip = None
    summaries = []
    if len(matches) == 1:
        m = matches[0]
        trip = _load_trip_360(conn, m['challan_id'], m['le_id'])
        trip['match_reason'] = m['reason']
        trip['confidence'] = m['confidence']
    elif len(matches) > 1:
        for m in matches[:10]:
            s = _load_trip_summary(conn, m['challan_id'], m['le_id'])
            s['reason'] = m['reason']
            s['confidence'] = m['confidence']
            summaries.append(s)

    conn.close()
    return render_template('trip_view.html', q=q, trip=trip, matches=summaries)


def _try_parse_gemini_json(raw):
    """Try to parse Gemini output as JSON. If it fails, attempt to repair
       common truncation issues (unterminated string, missing brackets, etc.).
       Returns (parsed_dict_or_None, error_msg_or_None)."""
    if not raw:
        return None, "Empty response from Gemini"

    # Strip code fences first
    s = raw.strip()
    if s.startswith('```'):
        s = s.split('\n', 1)[1] if '\n' in s else s
        if s.endswith('```'):
            s = s[:-3]
        if s.startswith('json'):
            s = s[4:].lstrip()
    s = s.strip()

    # First attempt: parse as-is
    try:
        return json.loads(s), None
    except json.JSONDecodeError:
        pass

    # Repair: walk through chars tracking string state and bracket nesting,
    # then close everything in correct order at the end.
    stack = []          # 'O' for open object, 'A' for open array
    in_string = False
    last_was_value_start = False   # detect bare ", value, ", value patterns we may need to truncate
    i = 0
    repaired = s
    n = len(repaired)
    last_complete_value_end = -1   # position right after the last successfully-closed value/pair
    while i < n:
        c = repaired[i]
        if in_string:
            if c == '\\' and i + 1 < n:
                i += 2; continue
            if c == '"':
                in_string = False
        else:
            if c == '"':
                in_string = True
            elif c == '{':
                stack.append('O')
            elif c == '[':
                stack.append('A')
            elif c == '}' or c == ']':
                if stack: stack.pop()
                last_complete_value_end = i + 1
            elif c == ',':
                last_complete_value_end = i   # cut here if needed
        i += 1

    # If unterminated string at the end, drop everything from the open quote forward
    # and trim back to the last comma so we don't leave a "key": (no value) artifact.
    if in_string:
        # Find the start of the open string
        last_quote = repaired.rfind('"')
        if last_complete_value_end > -1 and last_complete_value_end < last_quote:
            repaired = repaired[:last_complete_value_end].rstrip(',').rstrip()
        else:
            # Just close the string
            repaired += '"'

    # Re-walk just the brackets to know what's still open after repair above
    open_brace = repaired.count('{') - repaired.count('}')
    open_bracket = 0
    in_str = False
    bracket_stack = []
    j = 0
    while j < len(repaired):
        c = repaired[j]
        if in_str:
            if c == '\\' and j + 1 < len(repaired):
                j += 2; continue
            if c == '"':
                in_str = False
        else:
            if c == '"':   in_str = True
            elif c == '{': bracket_stack.append('}')
            elif c == '[': bracket_stack.append(']')
            elif c == '}' or c == ']':
                if bracket_stack: bracket_stack.pop()
        j += 1
    # Close in reverse order
    repaired = repaired.rstrip().rstrip(',')
    while bracket_stack:
        repaired += bracket_stack.pop()

    try:
        return json.loads(repaired), 'Output was truncated by Gemini — partial fields may be missing'
    except json.JSONDecodeError as e:
        return None, f"Gemini returned invalid JSON we couldn't repair: {e}. Try uploading the photo again."


def _friendly_gemini_error(e):
    """Turn raw Gemini API exceptions into user-friendly messages."""
    import re
    s = str(e)
    if 'RESOURCE_EXHAUSTED' in s or '429' in s:
        m = re.search(r"['\"]retryDelay['\"]:\s*['\"](\d+)s['\"]", s)
        delay = m.group(1) if m else '60'
        # Extract quota limit if present
        m2 = re.search(r"['\"]quotaValue['\"]:\s*['\"]?(\d+)['\"]?", s)
        limit = m2.group(1) if m2 else '5'
        return (f"⏳ Gemini free-tier limit reached ({limit} requests/minute). "
                f"Wait about {delay} seconds and click Extract again. "
                f"This is Google's API throttle — your data is fine.")
    if 'UNAVAILABLE' in s or '503' in s:
        return ("🌐 Gemini is temporarily overloaded on Google's side (not your fault, not a quota issue). "
                "We auto-retried 3 times and still couldn't get through. "
                "Try again in 30–60 seconds — it's usually a brief spike.")
    if 'API_KEY_INVALID' in s or 'invalid api key' in s.lower():
        return "🔑 Gemini API key is invalid. Open .env and paste a fresh key from https://aistudio.google.com/apikey, then restart."
    if 'PERMISSION_DENIED' in s:
        return ("🔒 Gemini API permission denied. Check that your API key has access to the "
                f"configured model ({os.environ.get('GEMINI_MODEL', 'gemini-flash-latest')}).")
    if 'NOT_FOUND' in s and 'model' in s.lower():
        return ("🚫 The configured Gemini model is no longer available for this API key "
                "(Google periodically retires model versions). Open .env, set "
                "GEMINI_MODEL=gemini-flash-latest (or remove the line to use the built-in "
                "default), then restart the app.")
    if 'DEADLINE_EXCEEDED' in s or 'timeout' in s.lower() or 'timed out' in s.lower():
        return "⏱️ Gemini took too long to respond (network stall or slow connection). Try again."
    return f"Gemini API error: {e}"


def _gemini_call_with_retry(call_fn, max_attempts=3):
    """Run a Gemini call, automatically retrying on transient errors.
       Handles:
         - 503 / UNAVAILABLE / 500 / DEADLINE_EXCEEDED → exponential backoff (1s, 3s, 7s)
         - 429 / RESOURCE_EXHAUSTED  → wait the retryDelay value Gemini sends, retry once
                                       (capped at 60s so a stuck request doesn't hang forever)
    """
    import time, re
    last_err = None
    for attempt in range(max_attempts):
        try:
            return call_fn()
        except Exception as e:
            last_err = e
            s = str(e)
            is_503 = ('UNAVAILABLE' in s or '503' in s
                      or 'INTERNAL' in s or '500' in s
                      or 'DEADLINE_EXCEEDED' in s)
            is_429 = ('RESOURCE_EXHAUSTED' in s or '429' in s)

            if not (is_503 or is_429) or attempt == max_attempts - 1:
                raise

            if is_429:
                # Honor Gemini's suggested retryDelay (e.g. '34s')
                m = re.search(r"['\"]retryDelay['\"]:\s*['\"]?(\d+)s['\"]?", s)
                delay = int(m.group(1)) if m else 35
                delay = min(delay + 2, 60)   # +2 s buffer, hard cap 60 s
            else:
                delay = [1, 3, 7][min(attempt, 2)]
            time.sleep(delay)
    raise last_err  # unreachable but defensive


def _load_image_for_gemini(file_path):
    """Load any uploaded file (JPG/PNG/PDF) into a PIL Image, applying EXIF rotation."""
    import io
    from PIL import Image, ImageOps
    ext = Path(file_path).suffix.lower()
    if ext == '.pdf':
        # Rasterize first page of PDF to PNG
        img_bytes = _rasterize_pdf_to_image(file_path)
        return Image.open(io.BytesIO(img_bytes)).convert('RGB')
    img = Image.open(file_path)
    img = ImageOps.exif_transpose(img).convert('RGB')
    return img


def _rasterize_pdf_to_image(pdf_path, page_num=0, dpi=200):
    """Convert one PDF page to PNG bytes using PyMuPDF."""
    import fitz
    doc = fitz.open(pdf_path)
    page = doc.load_page(page_num)
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page.get_pixmap(matrix=mat)
    out = pix.tobytes('png')
    doc.close()
    return out


def extract_vbl_invoice(file_path):
    """Send one VBL file to Gemini 2.5 Flash and return parsed JSON dict.
       Returns: {"ok": bool, "data": dict, "error": str | None, "raw": str}"""
    client, err = _get_gemini_client()
    if err:
        return {"ok": False, "data": {}, "error": err, "raw": ""}

    try:
        pil_img = _load_image_for_gemini(file_path)
    except Exception as e:
        return {"ok": False, "data": {}, "error": f"Could not load image: {e}", "raw": ""}

    try:
        from google.genai import types
        response = _gemini_call_with_retry(lambda: client.models.generate_content(
            model=os.environ.get('GEMINI_MODEL', 'gemini-flash-latest'),
            contents=[pil_img, VBL_EXTRACTION_PROMPT],
            config=types.GenerateContentConfig(
                response_mime_type='application/json',
                temperature=0.0,
                max_output_tokens=16384,
            ),
        ))
    except Exception as e:
        return {"ok": False, "data": {}, "error": _friendly_gemini_error(e), "raw": ""}

    raw = (response.text or '').strip()
    data, parse_warn = _try_parse_gemini_json(raw)
    if data is not None:
        return {"ok": True, "data": data, "error": parse_warn, "raw": raw}
    return {"ok": False, "data": {}, "error": parse_warn or "Could not parse Gemini output", "raw": raw}


# ─────────────────────────────────────────────────────────────────────────────
# (Old RapidOCR + regex parser removed — Gemini Flash above does all the work)
# ─────────────────────────────────────────────────────────────────────────────


# ─── /extract — upload page + handler ─────────────────────────────────────────

def _run_extraction_async(ext_id, saved_files):
    """Background worker: calls Gemini for each saved file, updates progress on
       the extractions row, and inserts results into extracted_invoices.
       Runs in a daemon thread; uses its own DB connection (sqlite3 doesn't
       share connections across threads)."""
    import time
    total = len(saved_files)
    # Free-tier quota guardrail kept for parity; on paid tier this no-ops effectively.
    throttle_seconds = 13 if total >= 5 else 0

    def write(note):
        c = get_db()
        try:
            c.execute('UPDATE extractions SET note=? WHERE id=?', (note, ext_id))
            c.commit()
        finally:
            c.close()

    try:
        for seq, (rel_name, full_path) in enumerate(saved_files, start=1):
            write(f'Reading file {seq} of {total}…')
            if seq > 1 and throttle_seconds:
                time.sleep(throttle_seconds)
            result = extract_vbl_invoice(full_path)
            c = get_db()
            try:
                c.execute(
                    '''INSERT INTO extracted_invoices
                       (extraction_id, file_name, seq, raw_json, error)
                       VALUES (?,?,?,?,?)''',
                    (ext_id, rel_name, seq,
                     json.dumps(result['data']) if result['ok'] else result['raw'],
                     result.get('error'))
                )
                c.commit()
            finally:
                c.close()
        write('Done')
        c = get_db()
        try:
            c.execute('UPDATE extractions SET status=? WHERE id=?', ('extracted', ext_id))
            c.commit()
        finally:
            c.close()
    except Exception as e:
        c = get_db()
        try:
            c.execute('UPDATE extractions SET status=?, note=? WHERE id=?',
                      ('failed', f'Error: {e}'[:500], ext_id))
            c.commit()
        finally:
            c.close()


@app.route('/extract', methods=['GET', 'POST'])
def extract_upload():
    if request.method == 'POST':
        files = request.files.getlist('files')
        files = [f for f in files if f and f.filename]
        if not files:
            flash('Please choose at least one file.')
            return redirect(url_for('extract_upload'))

        # Validate
        for f in files:
            ext = Path(f.filename).suffix.lower()
            if ext not in ALLOWED_EXTS:
                flash(f'Unsupported file type: {f.filename}. Use JPG, PNG, PDF.')
                return redirect(url_for('extract_upload'))

        # Create extraction record (status='pending', progress note set)
        conn = get_db()
        conn.execute(
            'INSERT INTO extractions (created_at, mode, status, note) VALUES (?,?,?,?)',
            (datetime.now().isoformat(), 'combine', 'pending',
             f'Queued — uploading {len(files)} file(s)')
        )
        ext_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.commit()
        conn.close()

        # Save all files to disk synchronously (fast), then dispatch the AI work
        # to a background thread so the user gets an immediate redirect.
        ext_dir = os.path.join(UPLOAD_DIR, str(ext_id))
        os.makedirs(ext_dir, exist_ok=True)
        saved_files = []
        for seq, f in enumerate(files, start=1):
            safe_name = f'{seq:02d}_{uuid.uuid4().hex[:8]}{Path(f.filename).suffix.lower()}'
            full_path = os.path.join(ext_dir, safe_name)
            f.save(full_path)
            rel_name = f'{ext_id}/{safe_name}'
            saved_files.append((rel_name, full_path))

        # Kick off background extraction
        import threading
        t = threading.Thread(target=_run_extraction_async,
                             args=(ext_id, saved_files),
                             daemon=True)
        t.start()

        # Redirect immediately — the review page polls until status=='extracted'
        return redirect(url_for('extract_review', extraction_id=ext_id))

    # GET — local OCR is always available once deps are installed
    return render_template('extract_upload.html', config_error=None)


@app.route('/extract/<int:extraction_id>/status')
def extract_status(extraction_id):
    """JSON status endpoint polled by the extract_review page while extraction
       is still running in the background."""
    conn = get_db()
    row = conn.execute('SELECT status, note, created_at FROM extractions WHERE id=?',
                       (extraction_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'status': 'not_found'}), 404
    total = conn.execute(
        'SELECT COUNT(*) FROM extracted_invoices WHERE extraction_id=?',
        (extraction_id,)).fetchone()[0]
    conn.close()
    return jsonify({
        'status': row['status'],         # pending | extracted | reviewed | used | failed
        'note': row['note'] or '',       # human-readable progress message
        'files_done': total,
        'created_at': row['created_at'],
    })


@app.route('/extract/<int:extraction_id>', methods=['GET', 'POST'])
def extract_review(extraction_id):
    conn = get_db()
    extraction = conn.execute(
        'SELECT * FROM extractions WHERE id=?', (extraction_id,)
    ).fetchone()
    if not extraction:
        conn.close()
        flash('Extraction not found.')
        return redirect(url_for('extract_upload'))

    # If the background extraction is still running (or failed), render a
    # lightweight polling page instead of the review form.
    if extraction['status'] in ('pending', 'failed'):
        conn.close()
        return render_template('extract_processing.html',
                               extraction=dict(extraction))

    invs = conn.execute(
        'SELECT * FROM extracted_invoices WHERE extraction_id=? ORDER BY seq',
        (extraction_id,)
    ).fetchall()

    if request.method == 'POST':
        # User has reviewed/edited each invoice and chosen mode → save edits
        mode = request.form.get('mode', 'combine')
        for inv in invs:
            edited = {}
            prefix = f'inv_{inv["id"]}_'
            for key in ['doc_no', 'date',
                        'consignor_name', 'consignor_address', 'consignor_gstin',
                        'consignee_name', 'consignee_address',
                        'consignee_gstin', 'consignee_mobile', 'vehicle_reg_no',
                        'truck_capacity', 'trip_type', 'transporter',
                        'total_weight_kg', 'total_quantity', 'quantity_unit']:
                edited[key] = request.form.get(prefix + key, '').strip()
            conn.execute(
                'UPDATE extracted_invoices SET edited_json=? WHERE id=?',
                (json.dumps(edited), inv['id'])
            )
        conn.execute('UPDATE extractions SET mode=?, status=? WHERE id=?',
                     (mode, 'reviewed', extraction_id))
        conn.commit()
        conn.close()
        return redirect(url_for('new_bill', from_extraction=extraction_id))

    invoices = []
    for inv in invs:
        d = dict(inv)
        try:
            d['parsed'] = json.loads(d['edited_json'] or d['raw_json'] or '{}')
        except Exception:
            d['parsed'] = {}
        invoices.append(d)
    conn.close()
    return render_template('extract_review.html',
                           extraction=dict(extraction),
                           invoices=invoices,
                           default_consignor=get_setting('default_consignor_name') or '')


@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    return send_from_directory(UPLOAD_DIR, filename)


# ─────────────────────────────────────────────────────────────────────────────
# CHALLAN EXTRACTION (Phase B) — handwritten Goods Consignment Notes
# ─────────────────────────────────────────────────────────────────────────────

CHALLAN_EXTRACTION_PROMPT = """You are reading a handwritten Goods Consignment Note (challan / LR copy)
issued by the transporter. Most fields are HANDWRITTEN in pen on a pre-printed form. The form
header (the transporter's own pre-printed name, address, GSTIN, FSSAI) is printed and constant — ignore it.

Read carefully — the photo may be at an angle, fingers may obscure parts, and handwriting can be
sloppy. Where unsure, give your BEST guess and mark "low" or "medium" confidence.

Return ONLY a JSON object matching this schema (no prose, no markdown fences):

{
  "lr_no":             "string — top-right printed/inked LR No. (4-5 digit number, e.g. 13833)",
  "challan_date":      "YYYY-MM-DD — Date: field (DD/MM/YY or DD/MM/YYYY → YYYY-MM-DD)",
  "consignor_name":    "string — Consignor field (the consignor company name as handwritten)",
  "consignor_address": "string — extra lines under consignor name (e.g. GADAN KHEDA KANPUR)",
  "consignee_name":    "string — Consignee field (handwritten, e.g. MAHAK ENTERPRISES)",
  "consignee_address": "string — extra lines under consignee name (e.g. BINDKI)",
  "from_city_state":   "string — From City, State value (handwritten)",
  "to_city_state":     "string — To City, State value (handwritten)",
  "invoice_no":        "string — Invoice Number (handwritten, e.g. 880207639)",
  "invoice_date":      "YYYY-MM-DD — Date of Invoice value",
  "consignment_value": number,
  "gst_number":        "string — GST Number if filled (often blank)",
  "no_of_articles":    "string — value in 'No. of Articles' column (e.g. '610Nag', '1115', '1700')",
  "description":       "string — value in 'Description of Goods' column (e.g. 'AS Invoice', 'Nag-As Invoice')",
  "value_of_goods":    number,
  "weight_kg":         number,
  "del_no":            "string — Del. No. field (numeric handwritten, e.g. 893358002)",
  "shipment_no":       "string — Shipment No. (often blank)",
  "cost_no":           "string — Cost No. (often blank)",
  "seal_no":           "string — Seal No. (often blank or e.g. '9MT', '20MT')",
  "driver_name":       "string — handwritten driver name (may be blank)",
  "driver_mobile":     "string — Driver Mob. No. (10 digits, handwritten)",
  "truck_no":          "string — Truck No. uppercase, no spaces (e.g. UP71T8680, UP91T4333, UP78AT8509)",
  "gate_in_time":      "string — Gate In Time (HH:MM or as written, often blank)",
  "gate_out_time":     "string — Gate Out Time (often blank)",
  "lane_transit_time": "string — often blank",
  "expected_arrival":  "string — Expected Arrival and Time (often blank)",
  "confidence_per_field": {},
  "notes": "string — anything unusual"
}

Rules:
- HANDWRITING is expected — read every visible written character.
- If a field is blank/empty on the paper, return "" or 0 / null appropriately.
- Truck numbers: uppercase, no spaces. "UP 71 T 8680" → "UP71T8680".
- Driver mobile: 10 digits only, strip any spaces or dashes.
- Dates: convert to YYYY-MM-DD strictly.
- Confidence: only include fields where you're "low" or "medium" confidence (omit if high).
- LR No. is sometimes printed in the original book and sometimes overwritten — read what's visible.
"""


def extract_challan_image(file_path):
    """Send one challan image/PDF to Gemini and return parsed JSON dict."""
    client, err = _get_gemini_client()
    if err:
        return {"ok": False, "data": {}, "error": err, "raw": ""}

    try:
        pil_img = _load_image_for_gemini(file_path)
    except Exception as e:
        return {"ok": False, "data": {}, "error": f"Could not load image: {e}", "raw": ""}

    try:
        from google.genai import types
        response = _gemini_call_with_retry(lambda: client.models.generate_content(
            model=os.environ.get('GEMINI_MODEL', 'gemini-flash-latest'),
            contents=[pil_img, CHALLAN_EXTRACTION_PROMPT],
            config=types.GenerateContentConfig(
                response_mime_type='application/json',
                temperature=0.0,
                max_output_tokens=4096,
            ),
        ))
    except Exception as e:
        return {"ok": False, "data": {}, "error": _friendly_gemini_error(e), "raw": ""}

    raw = (response.text or '').strip()
    data, parse_warn = _try_parse_gemini_json(raw)
    if data is not None:
        return {"ok": True, "data": data, "error": parse_warn, "raw": raw}
    return {"ok": False, "data": {}, "error": parse_warn or "Could not parse Gemini output", "raw": raw}


def remember_driver(conn, name, mobile):
    if not mobile:
        return
    mobile = ''.join(c for c in str(mobile) if c.isdigit())
    if len(mobile) < 8:
        return
    conn.execute(
        '''INSERT INTO drivers (mobile, name, updated_at) VALUES (?,?,?)
           ON CONFLICT(mobile) DO UPDATE SET name=excluded.name, updated_at=excluded.updated_at''',
        (mobile, (name or '').strip(), datetime.now().isoformat())
    )


def get_drivers():
    conn = get_db()
    rows = conn.execute('SELECT mobile, name FROM drivers ORDER BY updated_at DESC').fetchall()
    conn.close()
    return [dict(r) for r in rows]


def next_lr_no():
    n = int(get_setting('next_lr_number') or 1)
    return str(n), n


def find_unique_lr_no(conn, preferred=''):
    """Return an LR number not yet present in challans.
       If `preferred` is given and free, use it. Otherwise scan from next_lr_number upward."""
    pref = (preferred or '').strip()
    if pref:
        if not conn.execute('SELECT 1 FROM challans WHERE lr_no=?', (pref,)).fetchone():
            return pref
    cur = int(get_setting('next_lr_number') or 1)
    for _ in range(10_000):   # safety cap
        candidate = str(cur)
        if not conn.execute('SELECT 1 FROM challans WHERE lr_no=?', (candidate,)).fetchone():
            return candidate
        cur += 1
    # Last-ditch: random suffix
    return f'{cur}-{uuid.uuid4().hex[:4]}'


# ─── /challan routes ─────────────────────────────────────────────────────────

@app.route('/challans')
def challans_index():
    conn = get_db()
    rows = conn.execute(
        '''SELECT id, lr_no, challan_date, consignor_name, consignee_name,
                  truck_no, driver_name, status, weight_kg
           FROM challans ORDER BY id DESC LIMIT 200'''
    ).fetchall()
    conn.close()
    return render_template('challans_index.html', challans=[dict(r) for r in rows])


@app.route('/challan/extract', methods=['GET', 'POST'])
def challan_extract_upload():
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename:
            flash('Please choose a challan photo or PDF.')
            return redirect(url_for('challan_extract_upload'))
        ext = Path(f.filename).suffix.lower()
        if ext not in ALLOWED_EXTS:
            flash(f'Unsupported file type: {f.filename}. Use JPG, PNG, PDF.')
            return redirect(url_for('challan_extract_upload'))

        # Save to uploads/challans/
        chl_dir = os.path.join(UPLOAD_DIR, 'challans')
        os.makedirs(chl_dir, exist_ok=True)
        safe_name = f'{datetime.now().strftime("%Y%m%d-%H%M%S")}_{uuid.uuid4().hex[:8]}{ext}'
        full_path = os.path.join(chl_dir, safe_name)
        f.save(full_path)
        rel_name = f'challans/{safe_name}'

        # Extract via Gemini
        result = extract_challan_image(full_path)
        if not result['ok']:
            flash(f'Extraction failed: {result["error"]}')
            return redirect(url_for('challan_extract_upload'))

        # Save as draft challan (status='draft' until reviewed & saved)
        d = result['data']
        extracted_lr = (d.get('lr_no') or '').strip()
        # Use extracted LR if numeric AND free; else auto-find a free one.
        candidate = extracted_lr if extracted_lr.isdigit() else ''
        conn = get_db()
        lr_no_val = find_unique_lr_no(conn, preferred=candidate)
        if extracted_lr and lr_no_val != extracted_lr:
            flash(f'Extracted LR-{extracted_lr} already exists, using LR-{lr_no_val} instead — change it on the next screen if needed.')
        conn.execute('''
            INSERT INTO challans (
                lr_no, challan_date, consignor_name, consignor_address,
                consignee_name, consignee_address, from_city_state, to_city_state,
                invoice_no, invoice_date, consignment_value, gst_number,
                no_of_articles, description, value_of_goods, weight_kg,
                del_no, shipment_no, cost_no, seal_no,
                driver_name, driver_mobile, truck_no,
                gate_in_time, gate_out_time, lane_transit_time, expected_arrival,
                source_image, raw_extraction, confidence_json, status, created_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            lr_no_val, d.get('challan_date'),
            d.get('consignor_name'), d.get('consignor_address'),
            d.get('consignee_name'), d.get('consignee_address'),
            d.get('from_city_state'), d.get('to_city_state'),
            d.get('invoice_no'), d.get('invoice_date'),
            _safe_num(d.get('consignment_value')), d.get('gst_number'),
            d.get('no_of_articles'), d.get('description'),
            _safe_num(d.get('value_of_goods')), _safe_num(d.get('weight_kg')),
            d.get('del_no'), d.get('shipment_no'), d.get('cost_no'), d.get('seal_no'),
            d.get('driver_name'), d.get('driver_mobile'), (d.get('truck_no') or '').upper().replace(' ', ''),
            d.get('gate_in_time'), d.get('gate_out_time'),
            d.get('lane_transit_time'), d.get('expected_arrival'),
            rel_name, result['raw'],
            json.dumps(d.get('confidence_per_field', {})),
            'draft', datetime.now().isoformat()
        ))
        chl_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.commit()
        conn.close()
        return redirect(url_for('challan_review', challan_id=chl_id))

    next_lr, _ = next_lr_no()
    return render_template('challan_upload.html', next_lr=next_lr)


def _safe_num(v):
    try:
        return float(v) if v not in (None, '', 'null') else None
    except (TypeError, ValueError):
        return None


@app.route('/challan/<int:challan_id>', methods=['GET', 'POST'])
def challan_review(challan_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM challans WHERE id=?', (challan_id,)).fetchone()
    if not row:
        conn.close()
        return 'Challan not found', 404

    if request.method == 'POST':
        f = request.form
        lr_val = (f.get('lr_no') or '').strip()
        # If user kept a fresh auto-LR, increment the counter
        prev_lr = row['lr_no']
        if row['status'] == 'draft' and lr_val == prev_lr and lr_val.isdigit():
            cur_n = int(get_setting('next_lr_number') or 1)
            if int(lr_val) >= cur_n:
                set_setting('next_lr_number', str(int(lr_val) + 1))

        conn.execute('''
            UPDATE challans SET
                lr_no=?, challan_date=?, consignor_name=?, consignor_address=?,
                consignee_name=?, consignee_address=?, from_city_state=?, to_city_state=?,
                invoice_no=?, invoice_date=?, consignment_value=?, gst_number=?,
                no_of_articles=?, description=?, value_of_goods=?, weight_kg=?,
                del_no=?, pod_doc_no=?, shipment_no=?, cost_no=?, seal_no=?,
                driver_name=?, driver_mobile=?, truck_no=?,
                gate_in_time=?, gate_out_time=?, lane_transit_time=?, expected_arrival=?,
                status='open', updated_at=?
            WHERE id=?''', (
            lr_val, f.get('challan_date'),
            f.get('consignor_name'), f.get('consignor_address'),
            f.get('consignee_name'), f.get('consignee_address'),
            f.get('from_city_state'), f.get('to_city_state'),
            f.get('invoice_no'), f.get('invoice_date'),
            _safe_num(f.get('consignment_value')), f.get('gst_number'),
            f.get('no_of_articles'), f.get('description'),
            _safe_num(f.get('value_of_goods')), _safe_num(f.get('weight_kg')),
            f.get('del_no'), f.get('pod_doc_no'),
            f.get('shipment_no'), f.get('cost_no'), f.get('seal_no'),
            f.get('driver_name'), f.get('driver_mobile'),
            (f.get('truck_no') or '').upper().replace(' ', ''),
            f.get('gate_in_time'), f.get('gate_out_time'),
            f.get('lane_transit_time'), f.get('expected_arrival'),
            datetime.now().isoformat(), challan_id
        ))
        # Remember driver + truck for autocomplete next time
        remember_driver(conn, f.get('driver_name'), f.get('driver_mobile'))
        remember_vehicle(conn, (f.get('truck_no') or '').upper().replace(' ', ''))

        # Audit: log challan update
        before = dict(row)
        after = {
            'lr_no': lr_val, 'challan_date': f.get('challan_date'),
            'consignor_name': f.get('consignor_name'),
            'consignee_name': f.get('consignee_name'),
            'invoice_no': f.get('invoice_no'),
            'truck_no': (f.get('truck_no') or '').upper().replace(' ', ''),
            'driver_name': f.get('driver_name'), 'driver_mobile': f.get('driver_mobile'),
            'weight_kg': _safe_num(f.get('weight_kg')), 'status': 'open',
        }
        changes = _diff_dict(before, after, fields=after.keys())
        action_type = 'create' if before.get('status') == 'draft' else 'update'
        log_audit(conn, action_type, 'challan', challan_id,
                  summary=f'{"Created" if action_type == "create" else "Edited"} challan LR-{lr_val} → {f.get("consignee_name") or ""}',
                  changes=changes)

        # ── Phase D: auto-create a draft ledger entry the first time this challan
        #   transitions out of 'draft' status. Skip if one already exists.
        existing_le_id = row['ledger_entry_id'] if 'ledger_entry_id' in row.keys() else None
        if not existing_le_id:
            conn.execute('''
              INSERT INTO ledger_entries (
                challan_id, entry_date, gr_no, vehicle_no, station, shipment_no, trip_type,
                mt_qty, freight, advance_cash, advance_account, diesel,
                remarks, created_at
              ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
              challan_id, f.get('challan_date'),
              lr_val, (f.get('truck_no') or '').upper().replace(' ', ''),
              f.get('to_city_state'), f.get('invoice_no'),
              'One Way',
              (_safe_num(f.get('weight_kg')) or 0) / 1000.0 if f.get('weight_kg') else None,
              0, 0, 0, 0,
              f'Auto-created from challan LR-{lr_val} · {f.get("consignee_name") or ""}',
              datetime.now().isoformat()
            ))
            new_le_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            conn.execute('UPDATE challans SET ledger_entry_id=? WHERE id=?', (new_le_id, challan_id))
            log_audit(conn, 'create', 'ledger_entry', new_le_id,
                      summary=f'Auto-created from challan LR-{lr_val}')

        conn.commit()
        conn.close()
        flash(f'Challan LR-{lr_val} saved. Ledger draft created — fill in money in the Ledger.')
        return redirect(url_for('challans_index'))

    challan = dict(row)
    try:
        challan['confidence'] = json.loads(challan.get('confidence_json') or '{}')
    except Exception:
        challan['confidence'] = {}
    conn.close()
    return render_template('challan_review.html',
                           challan=challan,
                           drivers=get_drivers(),
                           vehicles=get_vehicles(),
                           audit_entries=get_audit_for('challan', challan_id))


@app.route('/challan/<int:challan_id>/delete', methods=['POST'])
def challan_delete(challan_id):
    conn = get_db()
    row = conn.execute('SELECT lr_no, consignee_name FROM challans WHERE id=?', (challan_id,)).fetchone()
    if not row:
        conn.close()
        flash('Challan not found.')
        return redirect(url_for('challans_index'))
    sm = f'Deleted challan LR-{row["lr_no"]} → {row["consignee_name"] or ""}'
    conn.execute('INSERT INTO challans_archive SELECT * FROM challans WHERE id=?', (challan_id,))
    conn.execute('DELETE FROM challans WHERE id=?', (challan_id,))
    log_audit(conn, 'delete', 'challan', challan_id, summary=sm + ' (→ Recycle Bin)')
    conn.commit()
    conn.close()
    flash('Challan moved to the Recycle Bin.')
    return redirect(url_for('challans_index'))


# ─── End challan routes ──────────────────────────────────────────────────────


# ═════════════════════════════════════════════════════════════════════════════
# PHASE C — Ledger + POD + Master data (Transporters, Diesel Vendors)
# ═════════════════════════════════════════════════════════════════════════════

# ── Transporter helpers ──
def get_transporters():
    conn = get_db()
    rows = conn.execute('SELECT id, name, mobile, bank_details, notes FROM transporters ORDER BY name').fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_diesel_vendors():
    conn = get_db()
    rows = conn.execute('SELECT id, name, location, notes FROM diesel_vendors ORDER BY name').fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── /masters — combined Transporters + Diesel Vendors page ─────────────────
@app.route('/masters')
def masters_index():
    return render_template('masters.html',
                           transporters=get_transporters(),
                           diesel_vendors=get_diesel_vendors())


@app.route('/masters/transporter/add', methods=['POST'])
def transporter_add():
    f = request.form
    name = (f.get('name') or '').strip()
    if not name:
        flash('Transporter name required.')
        return redirect(url_for('masters_index'))
    conn = get_db()
    try:
        conn.execute(
            '''INSERT INTO transporters (name, mobile, bank_details, notes, created_at)
               VALUES (?,?,?,?,?)''',
            (name, f.get('mobile') or '', f.get('bank_details') or '',
             f.get('notes') or '', datetime.now().isoformat())
        )
        new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        log_audit(conn, 'create', 'transporter', new_id, summary=f'Added transporter "{name}"')
        conn.commit()
        flash(f'Transporter "{name}" added.')
    except sqlite3.IntegrityError:
        flash(f'Transporter "{name}" already exists.')
    conn.close()
    return redirect(url_for('masters_index'))


@app.route('/masters/transporter/<int:tid>/delete', methods=['POST'])
def transporter_delete(tid):
    conn = get_db()
    row = conn.execute('SELECT name FROM transporters WHERE id=?', (tid,)).fetchone()
    name = row['name'] if row else f'#{tid}'
    conn.execute('DELETE FROM transporters WHERE id=?', (tid,))
    log_audit(conn, 'delete', 'transporter', tid, summary=f'Removed transporter "{name}"')
    conn.commit()
    conn.close()
    flash('Transporter removed.')
    return redirect(url_for('masters_index'))


@app.route('/masters/diesel-vendor/add', methods=['POST'])
def diesel_vendor_add():
    f = request.form
    name = (f.get('name') or '').strip()
    if not name:
        flash('Vendor name required.')
        return redirect(url_for('masters_index'))
    conn = get_db()
    try:
        conn.execute(
            '''INSERT INTO diesel_vendors (name, location, notes, created_at)
               VALUES (?,?,?,?)''',
            (name, f.get('location') or '', f.get('notes') or '',
             datetime.now().isoformat())
        )
        new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        log_audit(conn, 'create', 'diesel_vendor', new_id, summary=f'Added diesel vendor "{name}"')
        conn.commit()
        flash(f'Diesel vendor "{name}" added.')
    except sqlite3.IntegrityError:
        flash(f'Vendor "{name}" already exists.')
    conn.close()
    return redirect(url_for('masters_index'))


@app.route('/masters/diesel-vendor/<int:vid>/delete', methods=['POST'])
def diesel_vendor_delete(vid):
    conn = get_db()
    row = conn.execute('SELECT name FROM diesel_vendors WHERE id=?', (vid,)).fetchone()
    name = row['name'] if row else f'#{vid}'
    conn.execute('DELETE FROM diesel_vendors WHERE id=?', (vid,))
    log_audit(conn, 'delete', 'diesel_vendor', vid, summary=f'Removed diesel vendor "{name}"')
    conn.commit()
    conn.close()
    flash('Diesel vendor removed.')
    return redirect(url_for('masters_index'))


# ── Ledger ──

def _ledger_balance(freight, adv_cash, adv_acct, diesel):
    return float(freight or 0) - float(adv_cash or 0) - float(adv_acct or 0) - float(diesel or 0)


def _is_overdue(entry_date_iso, pod_received, threshold_days):
    if pod_received:
        return False
    if not entry_date_iso:
        return False
    try:
        d = datetime.strptime(entry_date_iso[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return False
    return (datetime.now().date() - d).days > threshold_days


@app.route('/ledger')
def ledger_index():
    f = request.args
    where = ['1=1']
    params = []
    if f.get('q'):
        where.append('(le.gr_no LIKE ? OR le.vehicle_no LIKE ? OR le.station LIKE ?)')
        q = f'%{f["q"].strip()}%'
        params += [q, q, q]
    if f.get('from'):
        where.append('le.entry_date >= ?'); params.append(f['from'])
    if f.get('to'):
        where.append('le.entry_date <= ?'); params.append(f['to'])
    if f.get('status') == 'pod_pending':
        where.append('le.pod_received = 0')
    elif f.get('status') == 'unpaid':
        where.append('le.paid = 0')
    elif f.get('status') == 'paid':
        where.append('le.paid = 1')

    sql = f'''
      SELECT le.*, t.name AS transporter_name, dv.name AS diesel_vendor_name
      FROM ledger_entries le
      LEFT JOIN transporters     t  ON t.id = le.transporter_id
      LEFT JOIN diesel_vendors   dv ON dv.id = le.diesel_vendor_id
      WHERE {" AND ".join(where)}
      ORDER BY le.entry_date DESC, le.id DESC
      LIMIT 500
    '''
    conn = get_db()
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    overdue_days = int(get_setting('pod_overdue_days') or 10)
    entries = []
    for r in rows:
        e = dict(r)
        e['balance']   = _ledger_balance(e['freight'], e['advance_cash'], e['advance_account'], e['diesel'])
        e['overdue']   = _is_overdue(e['entry_date'], e['pod_received'], overdue_days)
        entries.append(e)

    return render_template('ledger_index.html',
                           entries=entries,
                           filters=dict(f),
                           overdue_days=overdue_days)


@app.route('/ledger/new', methods=['GET', 'POST'])
def ledger_new():
    if request.method == 'POST':
        return _ledger_save_new()
    return render_template('ledger_form.html',
                           entry=None,
                           transporters=get_transporters(),
                           diesel_vendors=get_diesel_vendors(),
                           vehicles=get_vehicles(),
                           today=datetime.now().strftime('%Y-%m-%d'))


def _ledger_save_new():
    f = request.form
    conn = get_db()
    conn.execute('''
        INSERT INTO ledger_entries (
          challan_id, entry_date, gr_no, vehicle_no, station, shipment_no, trip_type,
          mt_qty, freight, advance_cash, advance_account, diesel,
          diesel_vendor_id, transporter_id, remarks, created_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        f.get('challan_id') or None, f.get('entry_date'),
        (f.get('gr_no') or '').strip(), (f.get('vehicle_no') or '').strip().upper(),
        f.get('station'), f.get('shipment_no'), f.get('trip_type', 'One Way'),
        _safe_num(f.get('mt_qty')), _safe_num(f.get('freight')),
        _safe_num(f.get('advance_cash')) or 0, _safe_num(f.get('advance_account')) or 0,
        _safe_num(f.get('diesel')) or 0,
        f.get('diesel_vendor_id') or None, f.get('transporter_id') or None,
        f.get('remarks'), datetime.now().isoformat()
    ))
    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    remember_vehicle(conn, (f.get('vehicle_no') or '').strip().upper())
    conn.commit()
    conn.close()
    flash(f'Ledger entry GR-{f.get("gr_no") or new_id} saved.')
    return redirect(url_for('ledger_view', le_id=new_id))


@app.route('/ledger/<int:le_id>', methods=['GET', 'POST'])
def ledger_view(le_id):
    conn = get_db()
    if request.method == 'POST':
        f = request.form
        before_row = conn.execute('SELECT * FROM ledger_entries WHERE id=?', (le_id,)).fetchone()
        before = dict(before_row) if before_row else {}
        conn.execute('''
            UPDATE ledger_entries SET
              entry_date=?, gr_no=?, vehicle_no=?, station=?, shipment_no=?, trip_type=?,
              mt_qty=?, freight=?, advance_cash=?, advance_account=?, diesel=?,
              diesel_vendor_id=?, transporter_id=?, remarks=?, updated_at=?
            WHERE id=?''', (
            f.get('entry_date'), (f.get('gr_no') or '').strip(),
            (f.get('vehicle_no') or '').strip().upper(),
            f.get('station'), f.get('shipment_no'), f.get('trip_type', 'One Way'),
            _safe_num(f.get('mt_qty')), _safe_num(f.get('freight')),
            _safe_num(f.get('advance_cash')) or 0, _safe_num(f.get('advance_account')) or 0,
            _safe_num(f.get('diesel')) or 0,
            f.get('diesel_vendor_id') or None, f.get('transporter_id') or None,
            f.get('remarks'), datetime.now().isoformat(), le_id
        ))
        remember_vehicle(conn, (f.get('vehicle_no') or '').strip().upper())
        # Audit
        after = {
            'entry_date': f.get('entry_date'), 'gr_no': (f.get('gr_no') or '').strip(),
            'vehicle_no': (f.get('vehicle_no') or '').strip().upper(),
            'station': f.get('station'), 'shipment_no': f.get('shipment_no'),
            'trip_type': f.get('trip_type', 'One Way'),
            'freight': _safe_num(f.get('freight')) or 0,
            'advance_cash': _safe_num(f.get('advance_cash')) or 0,
            'advance_account': _safe_num(f.get('advance_account')) or 0,
            'diesel': _safe_num(f.get('diesel')) or 0,
            'remarks': f.get('remarks'),
        }
        changes = _diff_dict(before, after, fields=after.keys())
        if changes:
            log_audit(conn, 'update', 'ledger_entry', le_id,
                      summary=f'Edited GR-{after["gr_no"]} ({len(changes)} field{"s" if len(changes)!=1 else ""})',
                      changes=changes)
        conn.commit()
        conn.close()
        flash('Ledger entry updated.')
        return redirect(url_for('ledger_view', le_id=le_id))

    row = conn.execute('SELECT * FROM ledger_entries WHERE id=?', (le_id,)).fetchone()
    conn.close()
    if not row:
        return 'Ledger entry not found', 404
    entry = dict(row)
    entry['balance'] = _ledger_balance(entry['freight'], entry['advance_cash'],
                                       entry['advance_account'], entry['diesel'])
    return render_template('ledger_form.html',
                           entry=entry,
                           transporters=get_transporters(),
                           diesel_vendors=get_diesel_vendors(),
                           vehicles=get_vehicles(),
                           today=datetime.now().strftime('%Y-%m-%d'),
                           audit_entries=get_audit_for('ledger_entry', le_id))


@app.route('/ledger/<int:le_id>/pod', methods=['POST'])
def ledger_pod(le_id):
    f = request.form
    pod_received = 1 if f.get('pod_received') else 0
    pod_date = f.get('pod_date') or datetime.now().strftime('%Y-%m-%d')

    pod_image_rel = None
    file = request.files.get('pod_image')
    if file and file.filename:
        ext = Path(file.filename).suffix.lower()
        if ext in ALLOWED_EXTS:
            pod_dir = os.path.join(UPLOAD_DIR, 'pods')
            os.makedirs(pod_dir, exist_ok=True)
            safe_name = f'pod_{le_id}_{uuid.uuid4().hex[:8]}{ext}'
            file.save(os.path.join(pod_dir, safe_name))
            pod_image_rel = f'pods/{safe_name}'

    conn = get_db()
    before_row = conn.execute('SELECT pod_received, pod_image, gr_no FROM ledger_entries WHERE id=?',
                              (le_id,)).fetchone()
    before = dict(before_row) if before_row else {}
    if pod_image_rel:
        conn.execute('UPDATE ledger_entries SET pod_received=?, pod_date=?, pod_image=?, updated_at=? WHERE id=?',
                     (pod_received, pod_date, pod_image_rel, datetime.now().isoformat(), le_id))
    else:
        conn.execute('UPDATE ledger_entries SET pod_received=?, pod_date=?, updated_at=? WHERE id=?',
                     (pod_received, pod_date, datetime.now().isoformat(), le_id))
    # Audit
    if before.get('pod_received') != pod_received:
        action = 'POD received' if pod_received else 'POD un-marked'
        if pod_image_rel and pod_received:
            action += ' (with photo)'
        log_audit(conn, 'pod_mark', 'ledger_entry', le_id,
                  summary=f'{action} for GR-{before.get("gr_no", le_id)} on {pod_date}')
    elif pod_image_rel:
        log_audit(conn, 'pod_mark', 'ledger_entry', le_id,
                  summary=f'POD photo added for GR-{before.get("gr_no", le_id)}')
    conn.commit()
    conn.close()
    flash('POD status updated.')
    # If the click came from the list page, return there. Else stay on the entry detail.
    ref = request.referrer or ''
    if '/ledger' in ref and not f'/ledger/{le_id}' in ref:
        return redirect(url_for('ledger_index'))
    return redirect(url_for('ledger_view', le_id=le_id))


@app.route('/ledger/<int:le_id>/paid', methods=['POST'])
def ledger_paid(le_id):
    f = request.form
    is_paid = 1 if f.get('paid') else 0
    conn = get_db()
    before_row = conn.execute(
        '''SELECT paid, gr_no, transporter_id, freight, advance_cash, advance_account, diesel
           FROM ledger_entries WHERE id=?''', (le_id,)).fetchone()
    before = dict(before_row) if before_row else {}
    ref = f'auto-paid:ledger:{le_id}'
    if is_paid:
        conn.execute('''UPDATE ledger_entries SET
                          paid=1, paid_date=?, paid_mode=?, paid_amount=?, paid_reference=?,
                          updated_at=?
                        WHERE id=?''',
                     (f.get('paid_date') or datetime.now().strftime('%Y-%m-%d'),
                      f.get('paid_mode'), _safe_num(f.get('paid_amount')),
                      f.get('paid_reference'),
                      datetime.now().isoformat(), le_id))
        # Mirror the flag into the payments ledger (the balance source of truth).
        net = (before.get('freight') or 0) - (before.get('advance_cash') or 0) \
            - (before.get('advance_account') or 0) - (before.get('diesel') or 0)
        amt = _safe_num(f.get('paid_amount')) or net
        if before.get('transporter_id'):
            _auto_payment_upsert(conn, 'transporter', before.get('transporter_id'),
                                 amt, ref,
                                 when=f.get('paid_date'),
                                 mode=f.get('paid_mode'),
                                 created_by=current_user())
    else:
        conn.execute('''UPDATE ledger_entries SET
                          paid=0, paid_date=NULL, paid_mode=NULL, paid_amount=NULL,
                          paid_reference=NULL, updated_at=?
                        WHERE id=?''',
                     (datetime.now().isoformat(), le_id))
        _auto_payment_remove(conn, ref)
    if before.get('paid') != is_paid:
        action = 'Marked paid' if is_paid else 'Un-marked paid'
        details = []
        if is_paid and f.get('paid_mode'):   details.append(f.get('paid_mode'))
        if is_paid and f.get('paid_amount'): details.append(f'₹{f.get("paid_amount")}')
        suffix = f' ({" · ".join(details)})' if details else ''
        log_audit(conn, 'paid_mark', 'ledger_entry', le_id,
                  summary=f'{action}{suffix} for GR-{before.get("gr_no", le_id)}')
    conn.commit()
    conn.close()
    flash('Payment status updated.')
    return redirect(url_for('ledger_view', le_id=le_id))


@app.route('/ledger/<int:le_id>/delete', methods=['POST'])
def ledger_delete(le_id):
    conn = get_db()
    row = conn.execute('SELECT gr_no FROM ledger_entries WHERE id=?', (le_id,)).fetchone()
    if not row:
        conn.close()
        flash('Ledger entry not found.')
        return redirect(url_for('ledger_index'))
    sm = f'Deleted ledger entry GR-{row["gr_no"]}'
    conn.execute('INSERT INTO ledger_entries_archive SELECT * FROM ledger_entries WHERE id=?', (le_id,))
    conn.execute('DELETE FROM ledger_entries WHERE id=?', (le_id,))
    log_audit(conn, 'delete', 'ledger_entry', le_id, summary=sm + ' (→ Recycle Bin)')
    conn.commit()
    conn.close()
    flash('Ledger entry moved to the Recycle Bin.')
    return redirect(url_for('ledger_index'))


def _parse_id_list(form_field):
    """Parse a comma-separated or repeated 'ids' field into a list of ints."""
    raw = request.form.getlist(form_field) or []
    if len(raw) == 1 and ',' in raw[0]:
        raw = raw[0].split(',')
    out = []
    for x in raw:
        try:
            n = int(str(x).strip())
            if n > 0:
                out.append(n)
        except ValueError:
            pass
    return out


@app.route('/ledger/bulk-pod', methods=['POST'])
def ledger_bulk_pod():
    ids = _parse_id_list('ids')
    if not ids:
        flash('Select at least one entry first.')
        return redirect(url_for('ledger_index'))
    today = datetime.now().strftime('%Y-%m-%d')
    conn = get_db()
    placeholders = ','.join(['?'] * len(ids))
    conn.execute(
        f'''UPDATE ledger_entries
            SET pod_received=1, pod_date=COALESCE(pod_date, ?), updated_at=?
            WHERE id IN ({placeholders}) AND COALESCE(pod_received,0)=0''',
        [today, datetime.now().isoformat()] + ids)
    affected = conn.total_changes
    log_audit(conn, 'bulk_pod', 'ledger_entry', 0,
              summary=f'Marked POD received for {affected} entries')
    conn.commit()
    conn.close()
    flash(f'Marked {affected} ledger entr{"y" if affected == 1 else "ies"} as POD received.')
    return redirect(url_for('ledger_index'))


@app.route('/ledger/bulk-paid', methods=['POST'])
def ledger_bulk_paid():
    ids = _parse_id_list('ids')
    mode = request.form.get('mode') or 'Cash'
    if not ids:
        flash('Select at least one entry first.')
        return redirect(url_for('ledger_index'))
    today = datetime.now().strftime('%Y-%m-%d')
    conn = get_db()
    placeholders = ','.join(['?'] * len(ids))
    # Capture the rows we're about to flip (still unpaid) so we can write one
    # payments row per newly-settled trip — the payments table is the balance
    # source of truth, so the flag alone is no longer enough.
    to_pay = conn.execute(
        f'''SELECT id, transporter_id, freight, advance_cash, advance_account, diesel
            FROM ledger_entries
            WHERE id IN ({placeholders}) AND COALESCE(paid,0)=0''',
        ids).fetchall()
    conn.execute(
        f'''UPDATE ledger_entries
            SET paid=1, paid_date=COALESCE(paid_date, ?), paid_mode=COALESCE(paid_mode, ?), updated_at=?
            WHERE id IN ({placeholders}) AND COALESCE(paid,0)=0''',
        [today, mode, datetime.now().isoformat()] + ids)
    affected = conn.total_changes
    for r in to_pay:
        if not r['transporter_id']:
            continue
        net = (r['freight'] or 0) - (r['advance_cash'] or 0) \
            - (r['advance_account'] or 0) - (r['diesel'] or 0)
        _auto_payment_upsert(conn, 'transporter', r['transporter_id'], net,
                             f'auto-paid:ledger:{r["id"]}',
                             when=today, mode=mode, created_by=current_user())
    log_audit(conn, 'bulk_paid', 'ledger_entry', 0,
              summary=f'Marked {affected} entries paid ({mode})')
    conn.commit()
    conn.close()
    flash(f'Marked {affected} ledger entr{"y" if affected == 1 else "ies"} as paid ({mode}).')
    return redirect(url_for('ledger_index'))


@app.route('/ledger/<int:le_id>/duplicate', methods=['POST'])
def ledger_duplicate(le_id):
    """Clone a ledger row for a recurring trip. Resets identifiers and POD/paid
       state, keeps freight/advance/diesel/vehicle/station/transporter."""
    conn = get_db()
    row = conn.execute('SELECT * FROM ledger_entries WHERE id=?', (le_id,)).fetchone()
    if not row:
        conn.close()
        flash('Ledger entry not found.')
        return redirect(url_for('ledger_index'))
    src = dict(row)
    today = datetime.now().strftime('%Y-%m-%d')
    conn.execute('''
        INSERT INTO ledger_entries (
            entry_date, gr_no, vehicle_no, station, shipment_no, trip_type,
            mt_qty, freight, advance_cash, advance_account, diesel,
            diesel_vendor_id, transporter_id, weight_kg,
            remarks, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        today,
        '',  # GR No. — leave blank for user to fill in fresh
        src.get('vehicle_no'), src.get('station'), src.get('shipment_no'),
        src.get('trip_type') or 'One Way',
        src.get('mt_qty'), src.get('freight'),
        src.get('advance_cash') or 0, src.get('advance_account') or 0,
        src.get('diesel') or 0,
        src.get('diesel_vendor_id'), src.get('transporter_id'),
        src.get('weight_kg'),
        f'Duplicated from GR-{src.get("gr_no") or src["id"]}',
        datetime.now().isoformat()))
    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    log_audit(conn, 'create', 'ledger_entry', new_id,
              summary=f'Duplicated from GR-{src.get("gr_no") or src["id"]}')
    conn.commit()
    conn.close()
    flash(f'Trip duplicated. Fill in the new GR No. and any changes.')
    return redirect(url_for('ledger_view', le_id=new_id))


# ── Ledger extraction (handwritten page → multiple rows) ─────────────────────

LEDGER_EXTRACTION_PROMPT = """You are reading a handwritten Daily Ledger / Bill Register page from
the transporter's own records. The page has a single date heading at the top (often DD/MM/YY)
and multiple ledger rows below in a table.

Each row's columns (left to right):
  • GR No.            — 4–5 digit number (e.g. 12571)
  • Vehicle No.       — like UP71AN0R41 (uppercase, no spaces)
  • Station           — destination town/city, e.g. ORAI, BINDKI, SISOLAR
  • Shipment / Invoice — long numeric like 3625003039, often followed by (OW), (BW), or (TW) in
                         parentheses. (OW)=One Way, (TW) or (BW)=Two Way / Both Ways.
  • MT Qty            — like "9MT", "9.5MT" — extract just the number
  • Freight           — ₹ amount (no currency symbol)
  • Advance           — cash advance to driver
  • Advance In A/c    — bank/account advance to vehicle owner
  • Diesel            — ₹ amount of diesel given
  • Balance           — computed (Freight − all advances). Read the written value if visible.
  • Remark            — short code, often a number like "9271" or text like "JL", "SDRAJ"
  • Signature         — initials or short name (JL, RAJKUMAR, KAHIVALYA, etc.)

CRITICAL — Vehicle numbers are split across TWO LINES:
  The main row shows the FIRST PART of the vehicle no., e.g. "UP71BCN", "UP91T", "UP51AT".
  The sub-row directly BELOW shows the LAST 4 DIGITS, e.g. "1294", "2403", "1928".
  → Combine them into the full vehicle_no: "UP71BCN" + "1294" = "UP71BCN1294".
  Indian commercial vehicle format is: 2-letter state + 2-digit district + 1-3 letter series
  + 4-digit number. So if the suffix isn't on the main row, pull it from the line below.

Each row's SUB-ROW (the line directly below the main entry) contains, in order:
  • A 4-digit number → this is the vehicle-no suffix described above (NOT a consignee reference)
  • A consignee/agency name (SHREE BALAJI AGENCY, MAHAK ENTERPRISES, MAA ANNAPURNA TRADERS, etc.)
  • An additional weight figure (e.g. 9022, 9050.275, 9038.840) — capture as "consignee_weight" if present.

Top-of-page totals like "Balance / Advance / Total: 30,612" are page summaries — IGNORE them.

Return ONLY a JSON object (no prose, no markdown fences):

{
  "page_date": "YYYY-MM-DD",
  "rows": [
    {
      "gr_no":             "12571",
      "vehicle_no":        "UP71AN3117",     // main + sub-row 4-digit suffix combined
      "station":           "ORAI",
      "shipment_no":       "3625003039",
      "trip_type":         "One Way" | "Two Way",
      "consignee_name":    "SHREE BALAJI AGENCY",
      "consignee_weight":  9022,             // optional, from sub-row if visible
      "mt_qty":            9,
      "freight":           6000,
      "advance_cash":      1500,
      "advance_account":   0,
      "diesel":            3000,
      "balance":           1500,
      "remark":            "9271",
      "signature":         "JL",
      "confidence":        "high" | "medium" | "low"
    }
  ],
  "notes": "anything unusual"
}

Rules:
- Vehicle numbers: uppercase, strip all spaces.
- Numbers: never include thousand separators or units inside the number value.
- Date: convert DD/MM/YY → YYYY-MM-DD (assume 20YY for the year).
- Trip type: if the shipment number is followed by "(OW)" → "One Way", "(TW)" or "(BW)" → "Two Way".
- If a column is blank for a row, return 0 for numeric or "" for string.
- Mark a row "low" confidence if multiple fields are unclear; "medium" if 1–2 fields.
- Read EVERY row visible on the page, in top-to-bottom order.
"""


def extract_ledger_image(file_path):
    """Send a ledger-page photo to Gemini and return parsed JSON."""
    client, err = _get_gemini_client()
    if err:
        return {"ok": False, "data": {}, "error": err, "raw": ""}
    try:
        pil_img = _load_image_for_gemini(file_path)
    except Exception as e:
        return {"ok": False, "data": {}, "error": f"Could not load image: {e}", "raw": ""}
    try:
        from google.genai import types
        response = _gemini_call_with_retry(lambda: client.models.generate_content(
            model=os.environ.get('GEMINI_MODEL', 'gemini-flash-latest'),
            contents=[pil_img, LEDGER_EXTRACTION_PROMPT],
            config=types.GenerateContentConfig(
                response_mime_type='application/json',
                temperature=0.0,
                max_output_tokens=8192,
            ),
        ))
    except Exception as e:
        return {"ok": False, "data": {}, "error": _friendly_gemini_error(e), "raw": ""}

    raw = (response.text or '').strip()
    data, parse_warn = _try_parse_gemini_json(raw)
    if data is not None:
        return {"ok": True, "data": data, "error": parse_warn, "raw": raw}
    return {"ok": False, "data": {}, "error": parse_warn or "Could not parse Gemini output", "raw": raw}


@app.route('/ledger/extract', methods=['GET', 'POST'])
def ledger_extract_upload():
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename:
            flash('Please choose a ledger photo or PDF.')
            return redirect(url_for('ledger_extract_upload'))
        ext = Path(f.filename).suffix.lower()
        if ext not in ALLOWED_EXTS:
            flash(f'Unsupported file type: {f.filename}.')
            return redirect(url_for('ledger_extract_upload'))

        led_dir = os.path.join(UPLOAD_DIR, 'ledger-pages')
        os.makedirs(led_dir, exist_ok=True)
        safe_name = f'{datetime.now().strftime("%Y%m%d-%H%M%S")}_{uuid.uuid4().hex[:8]}{ext}'
        full_path = os.path.join(led_dir, safe_name)
        f.save(full_path)
        rel_name = f'ledger-pages/{safe_name}'

        result = extract_ledger_image(full_path)
        if not result['ok']:
            flash(f'Extraction failed: {result["error"]}')
            return redirect(url_for('ledger_extract_upload'))

        data = result['data']
        conn = get_db()
        conn.execute(
            '''INSERT INTO ledger_extractions (source_image, page_date, raw_json, status, created_at)
               VALUES (?,?,?,?,?)''',
            (rel_name, data.get('page_date'), result['raw'], 'pending', datetime.now().isoformat())
        )
        ext_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        conn.commit()
        conn.close()
        return redirect(url_for('ledger_extract_review', le_id=ext_id))

    return render_template('ledger_extract_upload.html')


@app.route('/ledger/extract/<int:le_id>', methods=['GET', 'POST'])
def ledger_extract_review(le_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM ledger_extractions WHERE id=?', (le_id,)).fetchone()
    if not row:
        conn.close()
        return 'Extraction not found', 404

    if request.method == 'POST':
        # Bulk save reviewed rows into ledger_entries
        f = request.form
        n = max(0, min(100, _safe_int(f.get('row_count')) or 0))   # clamped 0..100
        page_date = f.get('page_date') or datetime.now().strftime('%Y-%m-%d')
        saved_count = 0
        for i in range(n):
            include = f.get(f'r_{i}_include')
            if not include:
                continue   # row was unchecked
            conn.execute('''
                INSERT INTO ledger_entries (
                  entry_date, gr_no, vehicle_no, station, shipment_no, trip_type,
                  mt_qty, freight, advance_cash, advance_account, diesel,
                  diesel_vendor_id, transporter_id, remarks, created_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                page_date,
                (f.get(f'r_{i}_gr_no') or '').strip(),
                (f.get(f'r_{i}_vehicle_no') or '').strip().upper(),
                f.get(f'r_{i}_station'),
                f.get(f'r_{i}_shipment_no'),
                f.get(f'r_{i}_trip_type', 'One Way'),
                _safe_num(f.get(f'r_{i}_mt_qty')),
                _safe_num(f.get(f'r_{i}_freight')) or 0,
                _safe_num(f.get(f'r_{i}_advance_cash')) or 0,
                _safe_num(f.get(f'r_{i}_advance_account')) or 0,
                _safe_num(f.get(f'r_{i}_diesel')) or 0,
                None, None,    # diesel_vendor_id, transporter_id — set on individual rows later
                f.get(f'r_{i}_remarks'),
                datetime.now().isoformat()
            ))
            remember_vehicle(conn, (f.get(f'r_{i}_vehicle_no') or '').strip().upper())
            saved_count += 1
        conn.execute('UPDATE ledger_extractions SET status=?, edited_json=? WHERE id=?',
                     ('used', json.dumps(dict(f)), le_id))
        conn.commit()
        conn.close()
        flash(f'{saved_count} ledger row{"s" if saved_count != 1 else ""} saved.')
        return redirect(url_for('ledger_index'))

    extraction = dict(row)
    try:
        parsed = json.loads(extraction['raw_json'])
    except Exception:
        parsed = {}
    extraction['parsed'] = parsed
    conn.close()
    return render_template('ledger_extract_review.html', extraction=extraction)


# ─── End Phase C routes ──────────────────────────────────────────────────────


# ═════════════════════════════════════════════════════════════════════════════
# PHASE D — Auto-link Challan ↔ Ledger ↔ Bill
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/to-bill')
def ready_to_bill():
    """Ledger entries that are POD-received and not yet billed."""
    conn = get_db()
    rows = conn.execute('''
      SELECT le.*, c.lr_no AS challan_lr, c.consignee_name AS challan_consignee,
             c.invoice_no AS challan_invoice, c.invoice_date AS challan_inv_date,
             c.weight_kg AS challan_weight
      FROM ledger_entries le
      LEFT JOIN challans c ON c.id = le.challan_id
      WHERE le.pod_received = 1 AND (le.bill_id IS NULL OR le.bill_id = 0)
      ORDER BY le.entry_date DESC, le.id DESC
    ''').fetchall()
    conn.close()
    return render_template('to_bill.html', entries=[dict(r) for r in rows])


def _build_prefill_from_ledger(le_id):
    """Read a ledger entry + its challan and build the same (bill_dict, deliveries) shape
       that _build_prefill_from_extraction produces, so /bill/new can render with bill=bill_dict."""
    conn = get_db()
    le_row = conn.execute('SELECT * FROM ledger_entries WHERE id=?', (le_id,)).fetchone()
    if not le_row:
        conn.close()
        return None, []
    le = dict(le_row)
    challan = None
    if le.get('challan_id'):
        c_row = conn.execute('SELECT * FROM challans WHERE id=?', (le['challan_id'],)).fetchone()
        if c_row:
            challan = dict(c_row)
    # Try to find a saved recipient for full address/GSTIN
    recipient = None
    consignee_name = (challan.get('consignee_name') if challan else '') or ''
    if consignee_name:
        r = conn.execute('SELECT * FROM recipients WHERE LOWER(name)=LOWER(?)',
                         (consignee_name,)).fetchone()
        if r:
            recipient = dict(r)
    conn.close()

    weight_kg = le.get('mt_qty', 0) or 0
    if weight_kg and weight_kg < 1000:
        weight_kg = weight_kg * 1000.0   # mt_qty is in MT, convert to kg
    if not weight_kg and challan and challan.get('weight_kg'):
        weight_kg = challan['weight_kg']

    bill_pre = {
        'bill_date': le.get('entry_date') or datetime.now().strftime('%Y-%m-%d'),
        'recipient_name': consignee_name,
        'recipient_address': (recipient or {}).get('address', '') or
                             (challan or {}).get('consignee_address', ''),
        'recipient_gstin': (recipient or {}).get('gstin', ''),
        'state_code': (recipient or {}).get('state_code', '')
                       or ((recipient or {}).get('gstin', '') or '')[:2],
        'trip_type': le.get('trip_type', 'One Way'),
        'vehicle_no': le.get('vehicle_no') or (challan or {}).get('truck_no', ''),
        'delivery_month': '',
        'delivery_month_select': '',
    }

    delivery = {
        'sr_no':            1,
        'gr_no':            le.get('gr_no') or (challan or {}).get('lr_no', ''),
        'outward_no':       le.get('shipment_no') or (challan or {}).get('invoice_no', ''),
        'outward_date':     (challan or {}).get('invoice_date', '') or le.get('entry_date', ''),
        'inward_no':        '',
        'inward_date':      '',
        'location':         le.get('station', ''),
        'consignee':        consignee_name,
        'delivery_qty':     '',
        'converted_case':   '',
        'inward_qty':       '',
        'empty_qty':        '',
        'weight':           weight_kg,
        'freight_rate':     '',     # auto-fills via JS rate-card lookup on page load
        'overload':         '',
        'toll_tax':         '',
        'excess_km':        '',
        'detention':        '',
        'unloading':        '',
        'value_of_supply':  '',
    }

    return bill_pre, [delivery]


# ─── End Phase D routes ──────────────────────────────────────────────────────


# ═════════════════════════════════════════════════════════════════════════════
# PHASE E — Dashboard + Reports + Client Payment
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/bill/<int:bill_id>/client-paid', methods=['POST'])
def bill_client_paid(bill_id):
    f = request.form
    is_paid = 1 if f.get('client_paid') else 0
    conn = get_db()
    before_row = conn.execute(
        'SELECT client_paid, bill_no, recipient_name, total_amount FROM bills WHERE id=?',
        (bill_id,)).fetchone()
    before = dict(before_row) if before_row else {}
    ref = f'auto-paid:bill:{bill_id}'
    if is_paid:
        conn.execute('''UPDATE bills SET client_paid=1, client_paid_date=?, client_paid_mode=?,
                                          client_paid_amount=?, client_paid_reference=?
                        WHERE id=?''',
                     (f.get('client_paid_date') or datetime.now().strftime('%Y-%m-%d'),
                      f.get('client_paid_mode'), _safe_num(f.get('client_paid_amount')),
                      f.get('client_paid_reference'), bill_id))
        # Mirror the flag into the payments ledger (the balance source of truth).
        amt = _safe_num(f.get('client_paid_amount')) or (before.get('total_amount') or 0)
        _auto_payment_upsert(conn, 'client', (before.get('recipient_name') or '').strip(),
                             amt, ref,
                             when=f.get('client_paid_date'),
                             mode=f.get('client_paid_mode'),
                             created_by=current_user())
    else:
        conn.execute('''UPDATE bills SET client_paid=0, client_paid_date=NULL, client_paid_mode=NULL,
                                          client_paid_amount=NULL, client_paid_reference=NULL
                        WHERE id=?''', (bill_id,))
        _auto_payment_remove(conn, ref)
    if before.get('client_paid') != is_paid:
        action = 'Client marked paid' if is_paid else 'Client payment un-marked'
        details = []
        if is_paid and f.get('client_paid_mode'):   details.append(f.get('client_paid_mode'))
        if is_paid and f.get('client_paid_amount'): details.append(f'₹{f.get("client_paid_amount")}')
        suffix = f' ({" · ".join(details)})' if details else ''
        log_audit(conn, 'client_paid', 'bill', bill_id,
                  summary=f'{action}{suffix} for {before.get("bill_no", bill_id)}')
    conn.commit()
    conn.close()
    flash('Client payment status updated.')
    return redirect(url_for('view_bill', bill_id=bill_id))


@app.route('/dashboard')
def dashboard():
    """Owner overview: this week's activity + outstanding amounts."""
    from datetime import timedelta
    today = datetime.now().date()
    week_ago = (today - timedelta(days=7)).isoformat()
    month_ago = (today - timedelta(days=30)).isoformat()
    overdue_days = int(get_setting('pod_overdue_days') or 10)
    overdue_cutoff = (today - timedelta(days=overdue_days)).isoformat()

    conn = get_db()
    def one(sql, *params):
        r = conn.execute(sql, params).fetchone()
        return r[0] if r else 0

    # ── Outstanding balances: computed the SAME way as the Payments Hub ──
    # (charges minus payments, via get_party_balance) so the dashboard and the
    # hub can never disagree. The payments table is the single source of truth;
    # we no longer read the per-row client_paid / paid flags for money.
    _clients_bal = list_clients_with_balance()
    client_outstanding = sum(c['balance'] for c in _clients_bal if c['balance'] > 0.005)
    client_outstanding_count = sum(1 for c in _clients_bal if c['balance'] > 0.005)
    transporter_balance_owed = 0.0
    for _t in get_transporters():
        _b = get_party_balance('transporter', _t['id'])
        if _b > 0.005:
            transporter_balance_owed += _b

    kpis = {
        # This week
        'trips_dispatched_week': one(
            "SELECT COUNT(*) FROM challans WHERE challan_date >= ? AND status != 'draft'",
            week_ago),
        'bills_generated_week': one(
            "SELECT COUNT(*) FROM bills WHERE bill_date >= ?", week_ago),
        'freight_billed_week':   one(
            "SELECT COALESCE(SUM(total_amount),0) FROM bills WHERE bill_date >= ?", week_ago),
        'advances_paid_week':    one(
            """SELECT COALESCE(SUM(advance_cash + advance_account + diesel),0)
               FROM ledger_entries WHERE entry_date >= ?""", week_ago),

        # This month
        'freight_billed_month':  one(
            "SELECT COALESCE(SUM(total_amount),0) FROM bills WHERE bill_date >= ?", month_ago),
        'advances_paid_month':   one(
            """SELECT COALESCE(SUM(advance_cash + advance_account + diesel),0)
               FROM ledger_entries WHERE entry_date >= ?""", month_ago),

        # Outstanding (right now)
        'pod_pending_count': one(
            "SELECT COUNT(*) FROM ledger_entries WHERE pod_received = 0"),
        'pod_overdue_count': one(
            "SELECT COUNT(*) FROM ledger_entries WHERE pod_received = 0 AND entry_date < ?",
            overdue_cutoff),
        'ready_to_bill_count': one(
            "SELECT COUNT(*) FROM ledger_entries WHERE pod_received = 1 AND (bill_id IS NULL OR bill_id = 0)"),
        'transporter_balance_owed': transporter_balance_owed,
        'client_outstanding':       client_outstanding,
        'client_outstanding_count': client_outstanding_count,
    }

    # Recent activity feed (last 10 events across challans/bills/POD)
    recent = []
    for r in conn.execute(
        """SELECT 'challan' AS kind, id, lr_no AS ref, challan_date AS dt, consignee_name AS detail
           FROM challans WHERE status != 'draft' ORDER BY id DESC LIMIT 5""").fetchall():
        recent.append(dict(r))
    for r in conn.execute(
        """SELECT 'bill' AS kind, id, bill_no AS ref, bill_date AS dt, recipient_name AS detail
           FROM bills ORDER BY id DESC LIMIT 5""").fetchall():
        recent.append(dict(r))
    recent.sort(key=lambda x: x.get('dt') or '', reverse=True)
    recent = recent[:10]

    conn.close()
    return render_template('dashboard.html', kpis=kpis, recent=recent,
                           overdue_days=overdue_days, vault=_backup_health())


@app.route('/reports/diesel')
def report_diesel():
    """Diesel given, grouped by vendor and by vehicle, over a date range."""
    from datetime import timedelta
    f = request.args
    today = datetime.now().date()
    default_from = (today - timedelta(days=30)).isoformat()
    df = f.get('from') or default_from
    dt = f.get('to')   or today.isoformat()

    conn = get_db()
    by_vendor = conn.execute("""
      SELECT COALESCE(dv.name, '— No vendor —') AS vendor,
             COUNT(le.id)  AS trips,
             COALESCE(SUM(le.diesel), 0) AS total_diesel
      FROM ledger_entries le
      LEFT JOIN diesel_vendors dv ON dv.id = le.diesel_vendor_id
      WHERE le.entry_date BETWEEN ? AND ? AND le.diesel > 0
      GROUP BY le.diesel_vendor_id
      ORDER BY total_diesel DESC
    """, (df, dt)).fetchall()

    by_vehicle = conn.execute("""
      SELECT COALESCE(le.vehicle_no, '—') AS vehicle,
             COUNT(*) AS trips,
             COALESCE(SUM(le.diesel), 0) AS total_diesel
      FROM ledger_entries le
      WHERE le.entry_date BETWEEN ? AND ? AND le.diesel > 0
      GROUP BY le.vehicle_no
      ORDER BY total_diesel DESC LIMIT 30
    """, (df, dt)).fetchall()

    grand_total = sum(r['total_diesel'] for r in by_vendor)
    grand_trips = sum(r['trips'] for r in by_vendor)
    conn.close()

    return render_template('report_diesel.html',
                           by_vendor=[dict(r) for r in by_vendor],
                           by_vehicle=[dict(r) for r in by_vehicle],
                           grand_total=grand_total, grand_trips=grand_trips,
                           date_from=df, date_to=dt)


@app.route('/reports/transporters')
def report_transporters():
    """Per-transporter rollup: trips, advances paid, balance owed."""
    f = request.args
    df = f.get('from') or ''
    dt = f.get('to')   or ''
    where, params = ['1=1'], []
    if df:
        where.append('le.entry_date >= ?'); params.append(df)
    if dt:
        where.append('le.entry_date <= ?'); params.append(dt)

    conn = get_db()
    rows = conn.execute(f"""
      SELECT t.id, t.name, t.mobile,
             COUNT(le.id) AS trip_count,
             COALESCE(SUM(le.freight), 0)         AS total_freight,
             COALESCE(SUM(le.advance_cash), 0)    AS total_cash,
             COALESCE(SUM(le.advance_account), 0) AS total_account,
             COALESCE(SUM(le.diesel), 0)          AS total_diesel
      FROM transporters t
      LEFT JOIN ledger_entries le
             ON le.transporter_id = t.id
            AND ({" AND ".join(where)})
      GROUP BY t.id
      ORDER BY trip_count DESC
    """, params).fetchall()
    # Balance owed comes from the payments ledger (single source of truth), the
    # same figure the Payments Hub shows — not the per-row `paid` flag. This is a
    # lifetime balance (charges minus payments), independent of the date filter.
    rows = [dict(r) for r in rows]
    for r in rows:
        r['balance_owed'] = get_party_balance('transporter', r['id'])
    rows.sort(key=lambda r: (r['balance_owed'], r['trip_count']), reverse=True)

    # Also include "no-transporter" trips so they don't disappear
    unassigned = conn.execute(f"""
      SELECT COUNT(*) AS trip_count,
             COALESCE(SUM(freight), 0)         AS total_freight,
             COALESCE(SUM(advance_cash), 0)    AS total_cash,
             COALESCE(SUM(advance_account), 0) AS total_account,
             COALESCE(SUM(diesel), 0)          AS total_diesel
      FROM ledger_entries le
      WHERE le.transporter_id IS NULL AND ({" AND ".join(where)})
    """, params).fetchone()
    conn.close()

    return render_template('report_transporters.html',
                           rows=[dict(r) for r in rows],
                           unassigned=dict(unassigned) if unassigned else None,
                           date_from=df, date_to=dt)


@app.route('/reports/transporter/<int:tid>')
def report_transporter_detail(tid):
    f = request.args
    df = f.get('from') or ''
    dt = f.get('to')   or ''
    conn = get_db()
    transporter = conn.execute('SELECT * FROM transporters WHERE id=?', (tid,)).fetchone()
    if not transporter:
        conn.close()
        return 'Transporter not found', 404

    where, params = ['transporter_id = ?'], [tid]
    if df: where.append('entry_date >= ?'); params.append(df)
    if dt: where.append('entry_date <= ?'); params.append(dt)

    rows = conn.execute(
        f"SELECT * FROM ledger_entries WHERE {' AND '.join(where)} ORDER BY entry_date DESC",
        params
    ).fetchall()
    conn.close()

    entries = []
    total_freight = total_cash = total_account = total_diesel = 0.0
    for r in rows:
        e = dict(r)
        f_amt  = e['freight'] or 0
        c_amt  = e['advance_cash'] or 0
        ac_amt = e['advance_account'] or 0
        d_amt  = e['diesel'] or 0
        e['balance'] = f_amt - c_amt - ac_amt - d_amt
        total_freight += f_amt
        total_cash    += c_amt
        total_account += ac_amt
        total_diesel  += d_amt
        entries.append(e)
    # Balance owed = charges minus payments (Payments-Hub math, single source of
    # truth) rather than summing the per-row `paid` flag. Lifetime figure.
    balance_owed_total = get_party_balance('transporter', tid)

    return render_template('report_transporter_detail.html',
                           transporter=dict(transporter),
                           entries=entries,
                           date_from=df, date_to=dt,
                           total_freight=total_freight, total_cash=total_cash,
                           total_account=total_account, total_diesel=total_diesel,
                           balance_owed_total=balance_owed_total)


# ─── End Phase E routes ──────────────────────────────────────────────────────


# ── Payments hub + per-party detail pages ───────────────────────────────────

@app.route('/payments')
def payments_hub():
    """Single page with three tabs: receivables, transporter payables, diesel payables."""
    clients = list_clients_with_balance()
    receivable_total = sum(c['balance'] for c in clients if c['balance'] > 0.005)

    transporters = get_transporters()
    for t in transporters:
        t['balance'] = get_party_balance('transporter', t['id'])
    transporter_payable = sum(t['balance'] for t in transporters if t['balance'] > 0.005)

    diesel_vendors = get_diesel_vendors()
    for v in diesel_vendors:
        v['balance'] = get_party_balance('diesel_vendor', v['id'])
    diesel_payable = sum(v['balance'] for v in diesel_vendors if v['balance'] > 0.005)

    return render_template('payments_hub.html',
        clients=clients, transporters=transporters, diesel_vendors=diesel_vendors,
        receivable_total=receivable_total,
        transporter_payable=transporter_payable,
        diesel_payable=diesel_payable,
        net_position=receivable_total - transporter_payable - diesel_payable,
    )


@app.route('/payments/<party_type>/<path:party_key>')
def party_detail(party_type, party_key):
    """Per-party ledger card: balance + recent transactions + record-payment form."""
    if party_type not in ('client', 'transporter', 'diesel_vendor'):
        return 'Unknown party type', 404

    display_name = party_key
    extra_info = {}
    if party_type == 'transporter':
        conn = get_db()
        r = conn.execute('SELECT * FROM transporters WHERE id=?', (party_key,)).fetchone()
        conn.close()
        if r:
            display_name = r['name']
            extra_info = {'mobile': r['mobile'], 'bank_details': r['bank_details']}
    elif party_type == 'diesel_vendor':
        conn = get_db()
        r = conn.execute('SELECT * FROM diesel_vendors WHERE id=?', (party_key,)).fetchone()
        conn.close()
        if r:
            display_name = r['name']
            extra_info = {'location': r['location']}

    balance = get_party_balance(party_type, party_key)
    transactions = get_party_transactions(party_type, party_key)

    # Compute charges + payments totals for the summary strip
    conn = get_db()
    if party_type == 'client':
        total_charges = _client_charges(conn, party_key)
    elif party_type == 'transporter':
        total_charges = _transporter_charges_net(conn, party_key)
    else:
        total_charges = _diesel_vendor_charges(conn, party_key)
    total_paid = _payments_total(conn, party_type, party_key)
    conn.close()

    return render_template('party_detail.html',
        party_type=party_type, party_key=party_key, display_name=display_name,
        extra_info=extra_info,
        balance=balance, total_charges=total_charges, total_paid=total_paid,
        transactions=transactions,
        today=datetime.now().strftime('%Y-%m-%d'),
    )


@app.route('/payments/add', methods=['POST'])
def payment_add():
    f = request.form
    party_type = f.get('party_type')
    party_key  = (f.get('party_key') or '').strip()
    if party_type not in ('client', 'transporter', 'diesel_vendor') or not party_key:
        flash('Missing party info.')
        return redirect(url_for('payments_hub'))
    amount = _safe_num(f.get('amount'))
    if not amount or amount <= 0:
        flash('Amount must be greater than zero.')
        return redirect(url_for('party_detail', party_type=party_type, party_key=party_key))

    conn = get_db()
    conn.execute('''
        INSERT INTO payments (party_type, party_key, payment_date, amount,
                              mode, reference, notes, source, created_at, created_by)
        VALUES (?,?,?,?,?,?,?, 'manual', ?, ?)
    ''', (party_type, party_key,
          f.get('payment_date') or datetime.now().strftime('%Y-%m-%d'),
          amount, f.get('mode'), f.get('reference'), f.get('notes'),
          datetime.now().isoformat(), current_user()))
    pid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    direction = 'received from' if party_type == 'client' else 'paid to'
    log_audit(conn, 'create', 'payment', pid,
              summary=f'₹{amount:,.2f} {direction} {party_type} {party_key}'
                      + (f' ({f.get("mode")})' if f.get('mode') else ''))
    conn.commit()
    conn.close()
    flash(f'Payment of ₹{amount:,.2f} recorded.')
    return redirect(url_for('party_detail', party_type=party_type, party_key=party_key))


@app.route('/payments/<int:pid>/delete', methods=['POST'])
def payment_delete(pid):
    conn = get_db()
    row = conn.execute('SELECT * FROM payments WHERE id=?', (pid,)).fetchone()
    if not row:
        conn.close()
        flash('Payment not found.')
        return redirect(url_for('payments_hub'))
    conn.execute('INSERT INTO payments_archive SELECT * FROM payments WHERE id=?', (pid,))
    conn.execute('DELETE FROM payments WHERE id=?', (pid,))
    log_audit(conn, 'delete', 'payment', pid,
              summary=f'Reverted ₹{row["amount"]:,.2f} payment for {row["party_type"]} {row["party_key"]} (→ Recycle Bin)')
    conn.commit()
    conn.close()
    flash('Payment moved to the Recycle Bin.')
    return redirect(url_for('party_detail', party_type=row['party_type'], party_key=row['party_key']))


# ── Recycle Bin (soft-delete restore / purge) ────────────────────────────────
# Deleted bills/challans/ledger entries/payments are archived in <table>_archive
# by their delete routes. Here the user can RESTORE them (move back to the live
# table with the original id, so FK links re-validate automatically) or PURGE
# them forever (which also nulls the now-permanently-broken FK links so freed
# records can be reused — e.g. ledger trips go back to "ready to bill").
#
# Table names below come from this fixed whitelist (never from user input), and
# <int:rid> is bound as a SQL parameter, so there is no injection surface.
_RECYCLE = {
    'bill':    {'table': 'bills',          'label': 'Bill',         'order': 'bill_no',
                'redirect': 'index',
                'purge_null': [('ledger_entries', 'bill_id')]},
    'challan': {'table': 'challans',       'label': 'Challan',      'order': 'lr_no',
                'redirect': 'challans_index',
                'purge_null': [('ledger_entries', 'challan_id')]},
    'ledger':  {'table': 'ledger_entries', 'label': 'Ledger entry', 'order': 'gr_no',
                'redirect': 'ledger_index',
                'purge_null': [('bills', 'ledger_entry_id'), ('challans', 'ledger_entry_id')]},
    'payment': {'table': 'payments',       'label': 'Payment',      'order': 'payment_date',
                'redirect': 'payments_hub',
                'purge_null': []},
}


@app.route('/recycle-bin')
def recycle_bin():
    conn = get_db()
    bins = {
        'bills': [dict(r) for r in conn.execute(
            'SELECT id, bill_no, bill_date, recipient_name, total_amount, created_at '
            'FROM bills_archive ORDER BY bill_no').fetchall()],
        'challans': [dict(r) for r in conn.execute(
            'SELECT id, lr_no, challan_date, consignee_name, truck_no, created_at '
            'FROM challans_archive ORDER BY lr_no').fetchall()],
        'ledger_entries': [dict(r) for r in conn.execute(
            'SELECT id, gr_no, entry_date, station, vehicle_no, freight, created_at '
            'FROM ledger_entries_archive ORDER BY entry_date DESC, id DESC').fetchall()],
        'payments': [dict(r) for r in conn.execute(
            'SELECT id, payment_date, party_type, party_key, amount, mode, created_at '
            'FROM payments_archive ORDER BY payment_date DESC, id DESC').fetchall()],
    }
    conn.close()
    total = sum(len(v) for v in bins.values())
    return render_template('recycle_bin.html', bins=bins, total=total)


@app.route('/recycle-bin/<entity>/<int:rid>/restore', methods=['POST'])
def recycle_restore(entity, rid):
    info = _RECYCLE.get(entity)
    if not info:
        flash('Unknown item type.')
        return redirect(url_for('recycle_bin'))
    conn = get_db()
    arch = info['table'] + '_archive'
    # Don't clobber a live row that somehow re-took this id.
    if conn.execute(f'SELECT 1 FROM {info["table"]} WHERE id=?', (rid,)).fetchone():
        conn.close()
        flash(f'{info["label"]} #{rid} already exists in the live list — not restored.')
        return redirect(url_for('recycle_bin'))
    conn.execute(f'INSERT INTO {info["table"]} SELECT * FROM {arch} WHERE id=?', (rid,))
    conn.execute(f'DELETE FROM {arch} WHERE id=?', (rid,))
    log_audit(conn, 'restore', entity, rid,
              summary=f'Restored {info["label"].lower()} #{rid} from Recycle Bin')
    conn.commit()
    conn.close()
    flash(f'{info["label"]} restored.')
    return redirect(url_for(info['redirect']))


@app.route('/recycle-bin/<entity>/<int:rid>/purge', methods=['POST'])
def recycle_purge(entity, rid):
    if current_user_role() != 'admin':
        flash('Only admins can permanently delete from the Recycle Bin.')
        return redirect(url_for('recycle_bin'))
    info = _RECYCLE.get(entity)
    if not info:
        flash('Unknown item type.')
        return redirect(url_for('recycle_bin'))
    conn = get_db()
    arch = info['table'] + '_archive'
    # Null the FK links that pointed at this row, so the freed records can be
    # reused (e.g. ledger trips go back to "ready to bill").
    for (ftable, fcol) in info['purge_null']:
        conn.execute(f'UPDATE {ftable} SET {fcol}=NULL WHERE {fcol}=?', (rid,))
    conn.execute(f'DELETE FROM {arch} WHERE id=?', (rid,))
    log_audit(conn, 'purge', entity, rid,
              summary=f'Permanently deleted {info["label"].lower()} #{rid}')
    conn.commit()
    conn.close()
    flash(f'{info["label"]} permanently deleted.')
    return redirect(url_for('recycle_bin'))


@app.route('/audit')
def audit_log_view():
    """Filterable global activity log."""
    f = request.args
    where, params = ['1=1'], []
    if f.get('user'):    where.append('user_name = ?'); params.append(f['user'])
    if f.get('action'):  where.append('action = ?'); params.append(f['action'])
    if f.get('entity'):  where.append('entity = ?'); params.append(f['entity'])
    if f.get('from'):    where.append('occurred_at >= ?'); params.append(f['from'])
    if f.get('to'):      where.append('occurred_at <= ?'); params.append(f['to'] + 'T23:59:59')
    if f.get('q'):
        where.append('(summary LIKE ? OR changes LIKE ?)')
        q = f'%{f["q"]}%'; params += [q, q]

    sql = f'SELECT * FROM audit_log WHERE {" AND ".join(where)} ORDER BY occurred_at DESC LIMIT 500'
    conn = get_db()
    rows = conn.execute(sql, params).fetchall()
    distinct_users    = [r[0] for r in conn.execute('SELECT DISTINCT user_name FROM audit_log ORDER BY user_name').fetchall() if r[0]]
    distinct_actions  = [r[0] for r in conn.execute('SELECT DISTINCT action FROM audit_log ORDER BY action').fetchall() if r[0]]
    distinct_entities = [r[0] for r in conn.execute('SELECT DISTINCT entity FROM audit_log ORDER BY entity').fetchall() if r[0]]
    conn.close()

    entries = []
    for r in rows:
        e = dict(r)
        try:
            e['changes_parsed'] = json.loads(e.get('changes') or 'null')
        except Exception:
            e['changes_parsed'] = None
        entries.append(e)

    return render_template('audit.html',
                           entries=entries,
                           filters=dict(f),
                           distinct_users=distinct_users,
                           distinct_actions=distinct_actions,
                           distinct_entities=distinct_entities)


# ── Automatic daily database backup ──────────────────────────────────────────
def backup_db_if_needed():
    """Create a dated copy of bills.db once per day. Keep last 14 local backups,
       and (if Google Drive is connected) upload the same file to the customer's
       own Drive folder in a background daemon thread."""
    if not os.path.exists(DB_PATH):
        return
    today = datetime.now().strftime('%Y-%m-%d')
    backup_path = os.path.join(BACKUP_DIR, f'bills-{today}.db')
    fresh_backup = False
    if not os.path.exists(backup_path):
        try:
            # Use SQLite's online backup API (safe even with the app running)
            import sqlite3 as _sql
            src = _sql.connect(DB_PATH)
            dst = _sql.connect(backup_path)
            with dst:
                src.backup(dst)
            src.close(); dst.close()
            fresh_backup = True
            # Record backup health so the dashboard can show a calm green/red
            # "Data Vault" status without re-scanning the folder.
            try:
                set_setting('last_backup_at', datetime.now().isoformat(timespec='seconds'))
                set_setting('last_backup_error', '')
            except Exception:
                pass
            # Prune local: keep newest 14
            backups = sorted(
                f for f in os.listdir(BACKUP_DIR)
                if f.startswith('bills-') and f.endswith('.db')
            )
            for old in backups[:-14]:
                try:
                    os.remove(os.path.join(BACKUP_DIR, old))
                except OSError:
                    pass
        except Exception as e:
            app.logger.warning(f'DB backup failed (non-fatal): {e}')
            try:
                set_setting('last_backup_error', str(e))
            except Exception:
                pass
            return

    # Google Drive sync: only if Drive is connected. Daemon-thread so the
    # foreground request isn't blocked on network I/O.
    try:
        if _drive_is_connected():
            state = _get_drive_state()
            already_synced_today = (state.get('last_uploaded_file') ==
                                    f'bills-{today}.db'
                                    and state.get('last_sync_status') == 'ok')
            # Trigger sync if we just made a fresh backup, OR if today's file
            # hasn't been uploaded yet (covers Flask restart mid-day).
            if fresh_backup or not already_synced_today:
                import threading
                threading.Thread(target=_drive_sync_now, daemon=True).start()
    except Exception as e:
        app.logger.warning(f'Drive sync dispatch failed (non-fatal): {e}')


def _backup_health():
    """Plain-English backup status for the dashboard 'Data Vault' tile.

    Returns a dict:
      last_backup_at : ISO string of the last successful backup (or '')
      latest_file    : newest bills-*.db filename in BACKUP_DIR (or None)
      latest_dt      : datetime of the newest backup file (or None)
      stale          : True if the newest backup is missing or older than ~2 days
      days_old       : whole days since the newest backup (or None)
      error          : last recorded backup error string (or '')
    """
    info = {
        'last_backup_at': '',
        'latest_file': None,
        'latest_dt': None,
        'stale': True,
        'days_old': None,
        'error': '',
    }
    try:
        info['last_backup_at'] = get_setting('last_backup_at') or ''
        info['error'] = get_setting('last_backup_error') or ''
    except Exception:
        pass
    try:
        files = sorted(
            f for f in os.listdir(BACKUP_DIR)
            if f.startswith('bills-') and f.endswith('.db')
        )
        if files:
            latest = files[-1]
            info['latest_file'] = latest
            mtime = os.path.getmtime(os.path.join(BACKUP_DIR, latest))
            latest_dt = datetime.fromtimestamp(mtime)
            info['latest_dt'] = latest_dt
            age_days = (datetime.now() - latest_dt).total_seconds() / 86400.0
            info['days_old'] = int(age_days)
            info['stale'] = age_days > 2
        else:
            info['stale'] = True
    except Exception as e:
        info['error'] = info['error'] or str(e)
        info['stale'] = True
    return info


# ── Health / status endpoint (so the user can verify config quickly) ─────────
@app.route('/health')
def health():
    """Lightweight status check — DB reachable, AI key configured, recent backup, etc."""
    status = {'ok': True, 'checks': {}}
    # DB reachable
    try:
        conn = get_db()
        bill_count = conn.execute('SELECT COUNT(*) FROM bills').fetchone()[0]
        ledger_count = conn.execute('SELECT COUNT(*) FROM ledger_entries').fetchone()[0]
        rate_count = conn.execute('SELECT COUNT(*) FROM freight_rates').fetchone()[0]
        conn.close()
        status['checks']['db'] = {'ok': True, 'bills': bill_count,
                                   'ledger_entries': ledger_count, 'rate_rows': rate_count}
    except Exception as e:
        status['ok'] = False
        status['checks']['db'] = {'ok': False, 'error': str(e)[:200]}

    # Gemini key configured
    key = os.environ.get('GOOGLE_API_KEY', '').strip()
    status['checks']['ai'] = {
        'configured': bool(key) and not key.startswith('PASTE_'),
        'model': os.environ.get('GEMINI_MODEL', 'gemini-flash-latest'),
    }

    # Latest backup
    try:
        backups = sorted(f for f in os.listdir(BACKUP_DIR) if f.startswith('bills-'))
        status['checks']['backups'] = {
            'count': len(backups),
            'latest': backups[-1] if backups else None,
        }
    except Exception:
        status['checks']['backups'] = {'count': 0}

    return jsonify(status)


# ── Restore from a local backup (admin only) ─────────────────────────────────
def _list_backup_files():
    """Newest-first list of {name, dt, date, size_kb} for the Restore screen."""
    out = []
    try:
        for f in os.listdir(BACKUP_DIR):
            if not (f.endswith('.db')):
                continue
            # Show the daily backups and any safety copies we made before a restore.
            if not (f.startswith('bills-') or f.startswith('bills_before')):
                continue
            full = os.path.join(BACKUP_DIR, f)
            if not os.path.isfile(full):
                continue
            st = os.stat(full)
            dt = datetime.fromtimestamp(st.st_mtime)
            out.append({
                'name': f,
                'dt': dt,
                'date': dt.strftime('%d %b %Y, %I:%M %p'),
                'size_kb': max(1, round(st.st_size / 1024)),
            })
    except Exception as e:
        app.logger.warning(f'Listing backups failed: {e}')
    out.sort(key=lambda x: x['dt'], reverse=True)
    return out


@app.route('/restore')
def restore_page():
    """List local backups so the owner can roll back to an earlier copy."""
    guard = _require_admin()
    if guard:
        return guard
    return render_template('restore.html',
                           backups=_list_backup_files(),
                           health=_backup_health())


@app.route('/restore', methods=['POST'])
def restore_do():
    """Restore a chosen backup over the live bills.db, after making a safety copy.

    Validates the filename is a real file that lives directly inside BACKUP_DIR
    (no path traversal, no symlinks pointing elsewhere)."""
    guard = _require_admin()
    if guard:
        return guard
    chosen = (request.form.get('filename') or '').strip()
    try:
        # Reject anything that isn't a bare filename.
        if not chosen or chosen != os.path.basename(chosen) or chosen.startswith('.'):
            flash('That backup name is not valid.')
            return redirect(url_for('restore_page'))
        target = os.path.join(BACKUP_DIR, chosen)
        # Confirm the resolved path is really inside BACKUP_DIR and is a file.
        backup_dir_real = os.path.realpath(BACKUP_DIR)
        target_real = os.path.realpath(target)
        if os.path.dirname(target_real) != backup_dir_real or not os.path.isfile(target_real):
            flash('That backup could not be found.')
            return redirect(url_for('restore_page'))

        # 1) Consistent safety copy of the CURRENT live database. We use the
        #    SQLite backup API (WAL-aware) instead of a raw file copy, because
        #    a raw copy of bills.db alone would MISS any recent committed writes
        #    still sitting in the -wal sidecar — i.e. the safety copy could be
        #    silently incomplete right when the user needs it most.
        if os.path.exists(DB_PATH):
            ts = datetime.now().strftime('%Y-%m-%d_%H%M%S')
            safety = os.path.join(BACKUP_DIR, f'bills_before_restore-{ts}.db')
            src = sqlite3.connect(DB_PATH)
            try:
                # Fold outstanding WAL frames into the main file first.
                try:
                    src.execute('PRAGMA wal_checkpoint(TRUNCATE)')
                except sqlite3.Error:
                    pass
                dst = sqlite3.connect(safety)
                try:
                    with dst:
                        src.backup(dst)
                finally:
                    dst.close()
            finally:
                src.close()

        # 2) Restore by copying the backup THROUGH SQLite (backup API) INTO the
        #    live database, not with a raw file copy. This overwrites every page
        #    inside the live DB's own transaction/WAL, so:
        #      • there are no stale -wal frames from the old data to replay
        #        (the root cause of "successful" restores that silently corrupt
        #        or revert), and
        #      • there are no sidecar files to delete — which on Windows can fail
        #        with a sharing violation whenever any connection is still open.
        #    It is safe even while the database is in use.
        srcconn = sqlite3.connect(target_real)
        dstconn = sqlite3.connect(DB_PATH)
        try:
            # Start from a clean WAL, overwrite all pages from the backup, then
            # fold the restored pages back into the main file.
            try:
                dstconn.execute('PRAGMA wal_checkpoint(TRUNCATE)')
            except sqlite3.Error:
                pass
            with dstconn:
                srcconn.backup(dstconn)
            try:
                dstconn.execute('PRAGMA wal_checkpoint(TRUNCATE)')
            except sqlite3.Error:
                pass
        finally:
            srcconn.close()
            dstconn.close()

        flash(f'Restore ho gaya — data ab "{chosen}" wali copy se chal raha hai. '
              'Purani copy safety ke liye backups folder me rakh li gayi hai. '
              'Behtar hoga ki Munshi ko ek baar band karke dobara kholein (restart).')
    except Exception as e:
        app.logger.warning(f'Restore failed: {e}')
        flash(f'Restore nahi ho paya: {e}')
    return redirect(url_for('restore_page'))


if __name__ == '__main__':
    # Friendly startup guard: turn the two most likely "won't start" failures
    # (a damaged data file, or the app already running) into plain-English
    # guidance in the terminal instead of a raw Python traceback.
    try:
        init_db()
    except sqlite3.DatabaseError as e:
        print('\n' + '=' * 64)
        print(' Munshi could not open your data file (bills.db).')
        print(' It looks damaged or is not a valid database file.')
        print(f' (technical detail: {e})')
        print('')
        print(' HOW TO FIX (your backups are safe — nothing was deleted):')
        print('  1. Close this window.')
        print('  2. Open the "backups" folder next to Munshi.')
        print('  3. Copy the newest  bills-YYYY-MM-DD.db  file.')
        print('  4. Rename the copy to  bills.db  and replace the damaged one.')
        print('  5. Start Munshi again.')
        print('=' * 64 + '\n')
        sys.exit(1)

    backup_db_if_needed()
    port = int(os.environ.get('PORT', 5056))

    # Pre-flight: is the port already taken? (Almost always means Munshi is
    # already running.) We check before app.run so we can show plain-English
    # guidance instead of Werkzeug's technical "Address already in use" trace.
    import socket
    _probe = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        _probe.bind(('127.0.0.1', port))
    except OSError:
        print('\n' + '=' * 64)
        print(f' Munshi is already running (port {port} is in use).')
        print(f' Just open your web browser to:   http://127.0.0.1:{port}')
        print('')
        print(' If that page does not load, close the OTHER Munshi terminal')
        print(' window first, then start Munshi again.')
        print('=' * 64 + '\n')
        sys.exit(1)
    finally:
        _probe.close()

    def open_browser():
        webbrowser.open(f'http://127.0.0.1:{port}')
    if not os.environ.get('PORT'):  # don't auto-open in preview/dev mode
        Timer(1.2, open_browser).start()
    app.run(host='127.0.0.1', port=port, debug=False)
